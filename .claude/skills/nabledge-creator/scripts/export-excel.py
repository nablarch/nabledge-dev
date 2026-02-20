#!/usr/bin/env python3
"""
Export mapping Markdown to Excel format.

Exit codes:
  0: Success
  1: Error
"""

import sys
import re
from pathlib import Path


def parse_mapping_file(file_path: str):
    """Parse mapping Markdown file."""
    rows = []

    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # Find table
    in_table = False
    for line in lines:
        if line.startswith('|') and 'Source Path' in line:
            in_table = True
            continue
        elif in_table and line.startswith('|---'):
            continue
        elif in_table and line.startswith('|'):
            cols = [c.strip() for c in line.split('|')[1:-1]]
            if len(cols) == 8:
                # Extract URL from markdown link
                url_match = re.search(r'\[🔗\]\((.*?)\)', cols[3])
                url = url_match.group(1) if url_match else ''

                rows.append({
                    'source_path': cols[0],
                    'title': cols[1],
                    'title_ja': cols[2],
                    'official_url': url,
                    'type': cols[4],
                    'category': cols[5],
                    'pp': cols[6],
                    'target_path': cols[7],
                })
        elif in_table and not line.startswith('|'):
            break

    return rows


def export_to_excel(rows, output_path: str):
    """Export rows to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Error: openpyxl not installed. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping v6"

    # Header
    headers = ['Source Path', 'Title', 'Title (ja)', 'Official URL', 'Type', 'Category ID', 'Processing Pattern', 'Target Path']
    ws.append(headers)

    # Make header bold
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    for row in rows:
        ws.append([
            row['source_path'],
            row['title'],
            row['title_ja'],
            row['official_url'],
            row['type'],
            row['category'],
            row['pp'],
            row['target_path'],
        ])

    # Set hyperlinks for URL column
    for row_idx, row in enumerate(rows, 2):  # Start from 2 (after header)
        url = row['official_url']
        if url:
            cell = ws[f'D{row_idx}']
            cell.hyperlink = url
            cell.value = '🔗'
            cell.font = Font(color="0000FF", underline="single")

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(header)

        # Check data for max length (sample first 100 rows for performance)
        for row_idx in range(2, min(102, len(rows) + 2)):
            cell_value = ws[f'{col_letter}{row_idx}'].value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        adjusted_width = min(max_length + 2, 60)  # Cap at 60
        ws.column_dimensions[col_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Enable auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Save
    wb.save(output_path)
    print(f"Exported {len(rows)} rows to {output_path}", file=sys.stderr)


def main():
    if len(sys.argv) < 2:
        print("Usage: export-excel.py MAPPING_FILE [--output PATH]", file=sys.stderr)
        sys.exit(1)

    mapping_file = sys.argv[1]
    output_path = mapping_file.replace('.md', '.xlsx')

    if '--output' in sys.argv:
        idx = sys.argv.index('--output')
        if idx + 1 < len(sys.argv):
            output_path = sys.argv[idx + 1]

    # Parse and export
    print(f"Parsing {mapping_file}...", file=sys.stderr)
    rows = parse_mapping_file(mapping_file)

    print(f"Exporting to {output_path}...", file=sys.stderr)
    export_to_excel(rows, output_path)

    sys.exit(0)


if __name__ == '__main__':
    main()
