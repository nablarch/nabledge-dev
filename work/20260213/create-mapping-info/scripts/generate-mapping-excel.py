#!/usr/bin/env python3
"""
Generate Excel format from mapping files.

Outputs:
- mapping-v6.xlsx: Excel workbook with multiple sheets
- mapping-v5.xlsx: Excel workbook with multiple sheets
"""

import json
import sys
from pathlib import Path
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def load_json(file_path: Path) -> Dict:
    """Load JSON file."""
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def create_excel_workbook(mapping: Dict, version: str) -> Workbook:
    """Create Excel workbook with multiple sheets."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary")
    create_summary_sheet(ws_summary, mapping, version)

    # Sheet 2: All Files
    ws_all = wb.create_sheet("All Files")
    create_all_files_sheet(ws_all, mapping)

    # Sheet 3: In Scope
    ws_in = wb.create_sheet("In Scope")
    create_in_scope_sheet(ws_in, mapping)

    # Sheet 4: Out of Scope
    ws_out = wb.create_sheet("Out of Scope")
    create_out_of_scope_sheet(ws_out, mapping)

    return wb

def create_summary_sheet(ws, mapping: Dict, version: str):
    """Create summary statistics sheet."""
    stats = mapping['statistics']

    # Title
    ws['A1'] = f"Mapping Summary - V{version}"
    ws['A1'].font = Font(size=16, bold=True)

    # Statistics
    ws['A3'] = "Metric"
    ws['B3'] = "Count"
    ws['C3'] = "Percentage"

    # Header style
    for cell in ['A3', 'B3', 'C3']:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # Data
    total = stats['total']
    in_scope = stats['in_scope']
    out_scope = stats['out_of_scope']

    ws['A4'] = "Total Files"
    ws['B4'] = total
    ws['C4'] = "100%"

    ws['A5'] = "In Scope"
    ws['B5'] = in_scope
    ws['C5'] = f"{in_scope/total*100:.1f}%"
    ws['A5'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws['B5'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws['C5'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ws['A6'] = "Out of Scope"
    ws['B6'] = out_scope
    ws['C6'] = f"{out_scope/total*100:.1f}%"
    ws['A6'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws['B6'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws['C6'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

def create_all_files_sheet(ws, mapping: Dict):
    """Create sheet with all files."""
    mappings = sorted(mapping['mappings'], key=lambda x: x['source_file'])

    # Headers
    headers = ['#', 'Source File', 'Title', 'In Scope', 'Categories', 'Reason for Exclusion', 'Target Files']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # Data
    for idx, entry in enumerate(mappings, 2):
        ws.cell(idx, 1, idx - 1)
        ws.cell(idx, 2, entry['source_file'])
        ws.cell(idx, 3, entry['title'])
        ws.cell(idx, 4, "✓" if entry['in_scope'] else "✗")
        ws.cell(idx, 5, ", ".join(entry['categories']) if entry['categories'] else "")
        ws.cell(idx, 6, entry['reason_for_exclusion'] or "")
        ws.cell(idx, 7, ", ".join(entry['target_files']) if entry['target_files'] else "")

        # Color coding
        if entry['in_scope']:
            ws.cell(idx, 4).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            ws.cell(idx, 4).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 50

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Auto filter
    ws.auto_filter.ref = f"A1:G{len(mappings) + 1}"

def create_in_scope_sheet(ws, mapping: Dict):
    """Create sheet with in-scope files only."""
    mappings = [m for m in mapping['mappings'] if m['in_scope']]
    mappings = sorted(mappings, key=lambda x: x['source_file'])

    # Headers
    headers = ['#', 'Source File', 'Title', 'Categories', 'Target Files']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Data
    for idx, entry in enumerate(mappings, 2):
        ws.cell(idx, 1, idx - 1)
        ws.cell(idx, 2, entry['source_file'])
        ws.cell(idx, 3, entry['title'])
        ws.cell(idx, 4, ", ".join(entry['categories']) if entry['categories'] else "")
        ws.cell(idx, 5, ", ".join(entry['target_files']) if entry['target_files'] else "")

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 50

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Auto filter
    ws.auto_filter.ref = f"A1:E{len(mappings) + 1}"

def create_out_of_scope_sheet(ws, mapping: Dict):
    """Create sheet with out-of-scope files only."""
    mappings = [m for m in mapping['mappings'] if not m['in_scope']]
    mappings = sorted(mappings, key=lambda x: (x['reason_for_exclusion'], x['source_file']))

    # Headers
    headers = ['#', 'Source File', 'Title', 'Reason for Exclusion']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Data
    for idx, entry in enumerate(mappings, 2):
        ws.cell(idx, 1, idx - 1)
        ws.cell(idx, 2, entry['source_file'])
        ws.cell(idx, 3, entry['title'])
        ws.cell(idx, 4, entry['reason_for_exclusion'] or "")

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 40

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Auto filter
    ws.auto_filter.ref = f"A1:D{len(mappings) + 1}"

def main():
    """Main entry point."""
    script_dir = Path(__file__).parent
    work_dir = script_dir.parent

    # Process V6
    print("Generating V6 Excel file...")
    mapping_v6_file = work_dir / 'mapping-v6.json'
    mapping_v6 = load_json(mapping_v6_file)

    wb_v6 = create_excel_workbook(mapping_v6, '6')
    output_v6 = work_dir / 'mapping-v6.xlsx'
    wb_v6.save(output_v6)

    print(f"Wrote {output_v6}")
    print(f"  Sheets: Summary, All Files, In Scope, Out of Scope")

    # Process V5
    print("\nGenerating V5 Excel file...")
    mapping_v5_file = work_dir / 'mapping-v5.json'
    mapping_v5 = load_json(mapping_v5_file)

    wb_v5 = create_excel_workbook(mapping_v5, '5')
    output_v5 = work_dir / 'mapping-v5.xlsx'
    wb_v5.save(output_v5)

    print(f"Wrote {output_v5}")
    print(f"  Sheets: Summary, All Files, In Scope, Out of Scope")

    print("\nExcel files generated!")
    print("\nFeatures:")
    print("  - Multiple sheets for easy navigation")
    print("  - Color coding (green=in scope, red=out of scope)")
    print("  - Auto filters on all data sheets")
    print("  - Frozen header rows")
    print("  - Optimized column widths")

if __name__ == '__main__':
    main()
