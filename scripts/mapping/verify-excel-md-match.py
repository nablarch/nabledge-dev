#!/usr/bin/env python3
"""
Verify that Excel and Markdown mapping files match completely.

This script:
1. Reads mapping-v6.md and parses the table
2. Reads mapping-v6.xlsx and extracts data including hyperlinks
3. Compares all columns row by row
4. Reports any mismatches

Usage:
    python scripts/mapping/verify-excel-md-match.py
"""

import re
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# File paths
REPO_ROOT = Path(__file__).parent.parent.parent
MAPPING_MD = REPO_ROOT / "doc/mapping/mapping-v6.md"
MAPPING_XLSX = REPO_ROOT / "doc/mapping/mapping-v6.xlsx"

def parse_markdown_table(file_path: Path) -> pd.DataFrame:
    """Parse mapping-v6.md Markdown table into DataFrame."""
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # Find table start
    table_start = None
    for i, line in enumerate(lines):
        if line.strip().startswith('| Source Path |'):
            table_start = i + 2  # Skip header and separator
            break

    if table_start is None:
        raise ValueError("Table header not found")

    # Parse rows
    rows = []
    for line in lines[table_start:]:
        line = line.strip()
        if not line or not line.startswith('|'):
            break

        cells = [cell.strip() for cell in line.split('|')[1:-1]]
        if len(cells) == 8:
            rows.append(cells)

    columns = [
        'Source Path',
        'Title',
        'Title (ja)',
        'Official URL',
        'Type',
        'Category ID',
        'Processing Pattern',
        'Target Path'
    ]

    df = pd.DataFrame(rows, columns=columns)

    # Extract URL from Markdown link format [🔗](url)
    df['Official URL MD'] = df['Official URL'].apply(extract_url_from_markdown)

    return df

def extract_url_from_markdown(md_link: str) -> str:
    """Extract URL from Markdown link format [🔗](url)."""
    match = re.search(r'\[.*?\]\((https://.*?)\)', md_link)
    if match:
        return match.group(1)
    return md_link

def read_excel_with_hyperlinks(file_path: Path) -> pd.DataFrame:
    """Read Excel file and extract hyperlinks from Official URL column."""
    # Read basic data (keep_default_na=False to preserve empty strings)
    df = pd.read_excel(file_path, keep_default_na=False)

    # Load workbook to extract hyperlinks
    wb = load_workbook(file_path)
    ws = wb.active

    # Extract hyperlinks from column D (Official URL)
    url_col = 4  # Column D
    hyperlinks = []

    for row in range(2, ws.max_row + 1):  # Skip header
        cell = ws.cell(row=row, column=url_col)
        if cell.hyperlink:
            hyperlinks.append(cell.hyperlink.target)
        else:
            hyperlinks.append(None)

    # Add hyperlinks to DataFrame
    df['Official URL Excel'] = hyperlinks

    return df

def compare_dataframes(md_df: pd.DataFrame, excel_df: pd.DataFrame) -> bool:
    """Compare Markdown and Excel DataFrames."""
    print("=" * 70)
    print("VERIFICATION REPORT: Markdown vs Excel")
    print("=" * 70)
    print()

    # Check row counts
    md_rows = len(md_df)
    excel_rows = len(excel_df)

    print(f"📊 Row Count Check")
    print(f"   Markdown: {md_rows} rows")
    print(f"   Excel:    {excel_rows} rows")

    if md_rows != excel_rows:
        print(f"   ❌ Row count mismatch!")
        return False
    else:
        print(f"   ✅ Row counts match")
    print()

    # Compare each column
    columns_to_check = [
        'Source Path',
        'Title',
        'Title (ja)',
        'Type',
        'Category ID',
        'Processing Pattern',
        'Target Path'
    ]

    all_match = True
    mismatches = []

    for col in columns_to_check:
        matches = (md_df[col] == excel_df[col]).all()
        if matches:
            print(f"   ✅ {col}: All {md_rows} rows match")
        else:
            print(f"   ❌ {col}: Mismatch found")
            all_match = False

            # Find mismatched rows
            for idx in range(len(md_df)):
                if md_df[col].iloc[idx] != excel_df[col].iloc[idx]:
                    mismatches.append({
                        'row': idx + 1,
                        'column': col,
                        'md_value': md_df[col].iloc[idx],
                        'excel_value': excel_df[col].iloc[idx]
                    })

    # Compare URLs (Markdown link vs Excel hyperlink)
    print()
    print(f"🔗 Official URL Check")
    url_matches = (md_df['Official URL MD'] == excel_df['Official URL Excel']).all()

    if url_matches:
        print(f"   ✅ All {md_rows} URLs match")
    else:
        print(f"   ❌ URL mismatch found")
        all_match = False

        # Find mismatched URLs
        for idx in range(len(md_df)):
            md_url = md_df['Official URL MD'].iloc[idx]
            excel_url = excel_df['Official URL Excel'].iloc[idx]
            if md_url != excel_url:
                mismatches.append({
                    'row': idx + 1,
                    'column': 'Official URL',
                    'md_value': md_url,
                    'excel_value': excel_url
                })

    print()
    print("=" * 70)

    if all_match:
        print("✅ VERIFICATION PASSED: Excel and Markdown are identical")
        return True
    else:
        print("❌ VERIFICATION FAILED: Mismatches found")
        print()
        print("Mismatches:")
        for i, mismatch in enumerate(mismatches[:10], 1):  # Show first 10
            print(f"{i}. Row {mismatch['row']}, Column '{mismatch['column']}':")
            print(f"   MD:    {mismatch['md_value']}")
            print(f"   Excel: {mismatch['excel_value']}")

        if len(mismatches) > 10:
            print(f"... and {len(mismatches) - 10} more mismatches")

        return False

def main():
    """Main execution"""
    print(f"📖 Reading {MAPPING_MD}")
    md_df = parse_markdown_table(MAPPING_MD)

    print(f"📖 Reading {MAPPING_XLSX}")
    excel_df = read_excel_with_hyperlinks(MAPPING_XLSX)

    print()
    success = compare_dataframes(md_df, excel_df)

    return 0 if success else 1

if __name__ == '__main__':
    sys.exit(main())
