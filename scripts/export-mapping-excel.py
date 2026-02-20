#!/usr/bin/env python3
"""
Export mapping-v6.md to Excel format for human review.

This script:
1. Parses mapping-v6.md Markdown table
2. Extracts Official URLs from Markdown links
3. Creates Excel file with:
   - Proper column widths
   - Clickable hyperlinks for Official URL
   - Filters on all columns
   - Frozen header row
4. Outputs to doc/mapping/mapping-v6.xlsx

Usage:
    python scripts/export-mapping-excel.py
"""

import re
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# File paths
REPO_ROOT = Path(__file__).parent.parent
MAPPING_MD = REPO_ROOT / "doc/mapping/mapping-v6.md"
OUTPUT_XLSX = REPO_ROOT / "doc/mapping/mapping-v6.xlsx"

def parse_markdown_table(file_path: Path) -> pd.DataFrame:
    """
    Parse mapping-v6.md Markdown table into pandas DataFrame.
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # Find table start (after header line with columns)
    table_start = None
    for i, line in enumerate(lines):
        if line.strip().startswith('| Source Path |'):
            table_start = i + 2  # Skip header and separator line
            break

    if table_start is None:
        raise ValueError("Table header not found in mapping file")

    # Parse table rows
    rows = []
    for line in lines[table_start:]:
        line = line.strip()
        if not line or not line.startswith('|'):
            break

        # Split by | and strip whitespace
        cells = [cell.strip() for cell in line.split('|')[1:-1]]

        if len(cells) == 8:
            rows.append(cells)

    # Create DataFrame
    columns = [
        'Source Path',
        'Title',
        'Title (ja)',
        'Official URL',
        'Type',
        'Category ID',
        'Processing Pattern',
        'Target Path'
    ]

    df = pd.DataFrame(rows, columns=columns)

    # Extract URL from Markdown link format [🔗](url)
    df['Official URL'] = df['Official URL'].apply(extract_url_from_markdown)

    return df

def extract_url_from_markdown(md_link: str) -> str:
    """
    Extract URL from Markdown link format [🔗](url).
    Returns the URL string for Excel hyperlink.
    """
    match = re.search(r'\[.*?\]\((https://.*?)\)', md_link)
    if match:
        return match.group(1)
    return md_link

def create_excel(df: pd.DataFrame, output_path: Path):
    """
    Create Excel file with proper formatting:
    - Hyperlinks in Official URL column
    - Column width auto-adjustment
    - Filters on all columns
    - Frozen header row
    """
    # Write DataFrame to Excel
    df.to_excel(output_path, index=False, sheet_name='Mapping v6')

    # Load workbook for formatting
    wb = load_workbook(output_path)
    ws = wb.active

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Enable filters on all columns
    ws.auto_filter.ref = ws.dimensions

    # Format header row
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment

    # Convert Official URL column to hyperlinks
    url_col = 4  # Column D (Official URL)
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=url_col)
        url = cell.value
        if url and url.startswith('http'):
            # Create hyperlink with 🔗 as display text
            cell.hyperlink = url
            cell.value = '🔗'
            cell.font = Font(color='0563C1', underline='single')
            cell.alignment = Alignment(horizontal='center')

    # Auto-adjust column widths
    column_widths = {
        'A': 50,  # Source Path
        'B': 40,  # Title
        'C': 40,  # Title (ja)
        'D': 8,   # Official URL (just 🔗)
        'E': 20,  # Type
        'F': 25,  # Category ID
        'G': 22,  # Processing Pattern
        'H': 50,  # Target Path
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Align all cells to top-left with wrap text
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            if cell.column != 4:  # Skip Official URL column
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Save workbook
    wb.save(output_path)

def main():
    """Main execution"""
    print(f"📖 Reading {MAPPING_MD}")

    try:
        df = parse_markdown_table(MAPPING_MD)
        print(f"✅ Parsed {len(df)} rows")

        print(f"📝 Creating Excel file...")
        create_excel(df, OUTPUT_XLSX)

        print(f"✅ Excel file created: {OUTPUT_XLSX}")
        print(f"   Total rows: {len(df)}")
        print(f"   Columns: {', '.join(df.columns)}")

        return 0

    except Exception as e:
        print(f"❌ Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1

if __name__ == '__main__':
    sys.exit(main())
