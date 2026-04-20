"""Excel release note converter for RBKC.

Converts Nablarch official release note Excel files into section-split form
suitable for knowledge JSON files.  No AI, no external API calls.

**Format**: Each data row (one release note entry) becomes one Section.
Category rows (e.g. 'アプリケーションフレームワーク') are skipped.

**Column layout** (consistent across v5/v6 releases):
  A (0): category name for category rows, or blank/type for data rows
  B (1): content type (optional, e.g. 'オブジェクトコード、ソースコード')
  C (2): No. (integer) — present for data rows, None for category rows
  D (3): 分類
  E (4): リリース区分 (変更/追加/削除/不具合)
  F (5): タイトル
  G (6): 概要
  H (7): 修正後のバージョン
  I (8): 不具合の起因バージョン (モジュール)
  J (9): 不具合の起因バージョン (Nablarch)
  K (10): システムへの影響の可能性
  L (11): システムへの影響の可能性の内容と対処
  M (12): 参照先

Public API:
    convert(path: Path, file_id: str = "") -> RSTResult
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from scripts.converters.rst import RSTResult, Section


# Column indices (0-based)
_COL_CATEGORY = 0    # A: category label row marker
_COL_NO = 2          # C: No. (integer for data rows)
_COL_BUNRUI = 3      # D: 分類
_COL_KUBUN = 4       # E: リリース区分
_COL_TITLE = 5       # F: タイトル
_COL_OVERVIEW = 6    # G: 概要
_COL_REF = 12        # M: 参照先


def _cell(row: tuple, idx: int) -> str | None:
    """Return cell value at *idx*, or None if out of range or None."""
    if idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    return str(val).strip() or None


def _is_data_row(row: tuple) -> bool:
    """True if this row is a release note data entry (No. column = integer)."""
    if _COL_NO >= len(row):
        return False
    return isinstance(row[_COL_NO], int)


def _build_content(row: tuple) -> str:
    """Build Markdown content for one release note row."""
    parts: list[str] = []

    kubun = _cell(row, _COL_KUBUN)
    bunrui = _cell(row, _COL_BUNRUI)

    meta: list[str] = []
    if kubun:
        meta.append(f"**リリース区分**: {kubun}")
    if bunrui:
        meta.append(f"**分類**: {bunrui}")
    if meta:
        parts.append("  ".join(meta))

    overview = _cell(row, _COL_OVERVIEW)
    if overview:
        parts.append(overview)

    ref = _cell(row, _COL_REF)
    if ref and ref != "-":
        parts.append(f"参照先: {ref}")

    return "\n\n".join(parts)


def _iter_rows_xls(path: Path):
    """Yield rows as tuples of cell values from a .xls file using xlrd."""
    import xlrd
    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)
    for rx in range(ws.nrows):
        yield tuple(ws.cell_value(rx, cx) for cx in range(ws.ncols))


# XLS column indices (0-based) — v1.x legacy release note format
_XLS_COL_NO = 0       # A: №
_XLS_COL_BUNRUI = 1   # B: 分類
_XLS_COL_TITLE = 2    # C: タイトル
_XLS_COL_OVERVIEW = 3 # D: 変更概要
_XLS_COL_KUBUN = 4    # E: 変更区分
_XLS_COL_IMPACT = 6   # G: 業務アプリへの影響
_XLS_COL_DETAIL = 7   # H: 影響時の内容と対処


def _is_xls_data_row(row: tuple) -> bool:
    """True if row is a data row: A column is a number (int or float with integer value)."""
    if not row:
        return False
    val = row[_XLS_COL_NO]
    if isinstance(val, float):
        return val > 0
    if isinstance(val, int):
        return val > 0
    return False


def _xls_cell(row: tuple, idx: int) -> str:
    """Return cell value as stripped string, or '' if absent/None."""
    if idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    return str(val).strip()


_IMPACT_NO_EFFECT = {"なし", "-", "－", ""}
_DETAIL_PLACEHOLDER = {"-", "－", ""}


def _is_no_effect(text: str) -> bool:
    """True if impact text indicates no effect (including multiline 'なし...' variants)."""
    return text.split("\n")[0].strip() in _IMPACT_NO_EFFECT


def _build_xls_content(row: tuple) -> str:
    """Build Markdown content for one xls release note row."""
    parts: list[str] = []
    kubun = _xls_cell(row, _XLS_COL_KUBUN)
    bunrui = _xls_cell(row, _XLS_COL_BUNRUI)
    meta: list[str] = []
    if kubun:
        meta.append(f"**変更区分**: {kubun}")
    if bunrui:
        meta.append(f"**分類**: {bunrui}")
    if meta:
        parts.append("  ".join(meta))
    overview = _xls_cell(row, _XLS_COL_OVERVIEW)
    if overview:
        parts.append(overview)
    impact = _xls_cell(row, _XLS_COL_IMPACT)
    detail = _xls_cell(row, _XLS_COL_DETAIL)
    if impact and not _is_no_effect(impact):
        parts.append(f"影響: {impact}")
    if detail and detail not in _DETAIL_PLACEHOLDER:
        parts.append(detail)
    return "\n\n".join(parts)


def _convert_xls(path: Path) -> RSTResult:
    """Convert a legacy .xls release note file."""
    rows = list(_iter_rows_xls(path))
    doc_title = ""
    for row in rows:
        val = row[0] if row else None
        if val and str(val).strip():
            doc_title = str(val).lstrip("■").strip()
            break
    sections: list[Section] = []
    for row in rows:
        if not _is_xls_data_row(row):
            continue
        no_val = row[_XLS_COL_NO]
        no_str = str(int(no_val)) if isinstance(no_val, float) else str(no_val)
        title_text = _xls_cell(row, _XLS_COL_TITLE)
        section_title = f"No.{no_str} {title_text}".strip()
        content = _build_xls_content(row)
        sections.append(Section(title=section_title, content=content))
    return RSTResult(title=doc_title, no_knowledge_content=False, sections=sections)


def convert(path: Path, file_id: str = "") -> RSTResult:
    """Convert a release note Excel file at *path* to :class:`RSTResult`.

    Supports both .xlsx (openpyxl) and .xls (xlrd) formats.

    Args:
        path: Path to the ``.xlsx`` or ``.xls`` file.
        file_id: Unused; present for API consistency with other converters.

    Returns:
        :class:`RSTResult` with one :class:`Section` per data row.
    """
    if path.suffix.lower() == ".xls":
        return _convert_xls(path)

    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active

    # Extract document title from the first non-empty row (row 1)
    doc_title = ""
    for row in ws.iter_rows(values_only=True):
        val = row[0] if row else None
        if val:
            doc_title = str(val).lstrip("■").strip()
            break

    sections: list[Section] = []
    for row in ws.iter_rows(values_only=True):
        if not _is_data_row(row):
            continue

        no = row[_COL_NO]
        title_text = _cell(row, _COL_TITLE) or ""
        section_title = f"No.{no} {title_text}".strip()
        content = _build_content(row)
        sections.append(Section(title=section_title, content=content))

    return RSTResult(
        title=doc_title,
        no_knowledge_content=False,
        sections=sections,
    )
