"""Shared Excel parsing utilities for RBKC converters.

Implements Phase 22-B sheet-level file split: each worksheet produces one
:class:`RSTResult` (→ 1 JSON + 1 docs MD).  Release-note and security-table
converters share this logic; they differ only in the mapping entries that
select them.

Design reference: ``tools/rbkc/docs/rbkc-converter-design.md`` §8.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from scripts.create.converters.rst import RSTResult, Section


# ---------------------------------------------------------------------------
# Raw sheet extraction (format-agnostic)
# ---------------------------------------------------------------------------

@dataclass
class RawSheet:
    """Sheet contents normalised to a 2-D list of stripped strings.

    Empty cells become ``""``.  Row lengths are uniform (padded with "").
    """
    name: str
    rows: list[list[str]]  # rows[r][c] -- all strings, "" when empty

    @property
    def nrows(self) -> int:
        return len(self.rows)

    @property
    def ncols(self) -> int:
        return len(self.rows[0]) if self.rows else 0


def list_sheet_names(path: Path) -> list[str]:
    """Return worksheet names in file order."""
    ext = path.suffix.lower()
    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(str(path))
        return [s.name for s in wb.sheets()]
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    return list(wb.sheetnames)


def read_sheet(path: Path, sheet_name: str) -> RawSheet:
    """Read one worksheet into :class:`RawSheet` (stripped strings, "" for empty)."""
    ext = path.suffix.lower()
    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(str(path))
        sheet = wb.sheet_by_name(sheet_name)
        rows: list[list[str]] = []
        for rx in range(sheet.nrows):
            row = []
            for cx in range(sheet.ncols):
                v = sheet.cell_value(rx, cx)
                row.append(str(v).strip() if v is not None else "")
            rows.append(row)
    else:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), data_only=True)
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([
                str(v).strip() if v is not None else ""
                for v in row
            ])
    # Trim trailing fully-empty rows (common in xlsx).  Keep leading empties —
    # they preserve row numbering for reference.
    while rows and all(c == "" for c in rows[-1]):
        rows.pop()
    # Normalise row widths (openpyxl already does this; xlrd may vary).
    width = max((len(r) for r in rows), default=0)
    for r in rows:
        if len(r) < width:
            r.extend([""] * (width - len(r)))
    # Trim trailing all-empty columns.
    while width > 0 and all(r[width - 1] == "" for r in rows):
        for r in rows:
            r.pop()
        width -= 1
    return RawSheet(name=sheet_name, rows=rows)


# ---------------------------------------------------------------------------
# Title / preamble extraction
# ---------------------------------------------------------------------------

def _collect_preamble(
    rows: list[list[str]],
    start: int,
    header_row_idx: int,
    title_row_extras: list[str] | None = None,
) -> str:
    """Collect preamble text from rows[start:header_row_idx] (P1 only).

    Every non-empty cell in the range becomes one preamble line
    (flattened).  Empty rows are skipped.  Callers use this only when a
    header row was actually located (P1); for P2 there is no preamble.

    ``title_row_extras``: optional non-title cells from the title row
    (row 0) that are prepended to the preamble.  The v5/v6 releasenote
    pattern keeps the ``■…`` title in col 0 and scatters ※-annotations
    in cols 7/9/11/… within the same row — these must be captured.
    """
    lines: list[str] = []
    if title_row_extras:
        for c in title_row_extras:
            lines.append(" ".join(c.split()))
    for i in range(start, header_row_idx):
        for c in rows[i]:
            if c:
                lines.append(" ".join(c.split()))
    return "\n".join(lines)


def _extract_title(sheet: RawSheet) -> tuple[str, int, list[str]]:
    """Return (title, next_row_index, title_row_extras).

    * title: row-1 ``■...`` text if present, else sheet name.
    * next_row_index: 1 when a ``■…`` row is consumed, else 0.
    * title_row_extras: when a ``■…`` row is consumed, every OTHER
      non-empty cell on that row (annotations scattered alongside the
      title in corpus releasenotes).
    """
    rows = sheet.rows
    title = sheet.name
    if not rows:
        return title, 0, []
    row0 = rows[0]
    first_idx = next((i for i, v in enumerate(row0) if v), -1)
    if first_idx < 0 or not row0[first_idx].startswith("■"):
        return title, 0, []
    title = " ".join(row0[first_idx].split())
    extras = [v for i, v in enumerate(row0) if v and i != first_idx]
    return title, 1, extras


# ---------------------------------------------------------------------------
# Header detection (P1 candidate)
# ---------------------------------------------------------------------------

def _run_length(row: list[str]) -> int:
    """Longest run of contiguous non-empty cells in *row*."""
    best = cur = 0
    for v in row:
        if v:
            cur += 1
            best = max(best, cur)
        else:
            cur = 0
    return best


#: Separator for span-inherit composed column names (spec §8-3, Phase
#: 22-B-12).  Kept in sync with ``scripts/verify/verify.SEP`` — both sides
#: derive independently from the same spec constant.
SEP = " / "


def _looks_like_header_cell(v: str) -> bool:
    """Heuristic mirror of ``scripts.verify.verify._looks_like_header_cell``.

    A header cell is a short (≤ 40 chars), non-numeric label.  Applied
    to reject data rows (numeric IDs, long prose) from being mis-picked
    as header rows.
    """
    if not v:
        return False
    s = v.strip()
    if not s or len(s) > 40:
        return False
    try:
        float(s.replace(",", ""))
        return False
    except ValueError:
        pass
    return True


def _looks_like_sub_header(row_h: list[str], row_h1: list[str]) -> bool:
    """Spec §8-3 span-inherit parent/leaf detection (Phase 22-B-12).

    Independently derived from the spec, identical semantics to the
    verify-side helper.  A multi-row header is recognised iff some
    parent cell genuinely spans ≥ 2 leaf cells (i.e. there is a parent
    column p such that the leaf row has ≥ 2 non-empty cells in
    [p, next_parent_column)).  Additionally:
      - all non-empty leaf cells must look like header labels
      - all non-empty parent cells must look like header labels (long
        free-text cells are preamble, not parents)
    """
    h_non_empty_cols = [cx for cx, v in enumerate(row_h) if v]
    h1_non_empty_cols = [cx for cx, v in enumerate(row_h1) if v]
    if not h_non_empty_cols or not h1_non_empty_cols:
        return False
    if not all(_looks_like_header_cell(row_h1[c]) for c in h1_non_empty_cols):
        return False
    if not all(_looks_like_header_cell(row_h[c]) for c in h_non_empty_cols):
        return False
    h1_col_set = set(h1_non_empty_cols)
    max_c = max(len(row_h), len(row_h1))
    for i, p_col in enumerate(h_non_empty_cols):
        next_p = h_non_empty_cols[i + 1] if i + 1 < len(h_non_empty_cols) else max_c
        leaf_in_span = sum(1 for lc in h1_col_set if p_col <= lc < next_p)
        if leaf_in_span >= 2:
            return True
    return False


def _compose_header_columns(header_rows: list[list[str]], col_count: int) -> list[str]:
    """Span-inherit composition (spec §8-3, Phase 22-B-12).

    Two distinct roles:

    * **The top row** (``header_rows[0]``) is the primary header.  Each
      non-empty cell spans rightward until the next non-empty cell in
      the same row — but only over columns that have no value in their
      own top-row cell.  This handles the v5 bessatsu pattern where the
      top row has only scattered parent labels and the bottom row
      carries the per-column leaves.
    * **Rows below the top** are sub-headers that split specific
      top-row spans.  A sub-header cell at column ``c`` appends to the
      top-row label that covers column ``c``.

    Composed column = ``SEP.join([top_label, sub_label_1, sub_label_2, ...])``
    with empty parts dropped.  A column where the top row is empty AND
    no sub-header covers it yields "".
    """
    if not header_rows:
        return [""] * col_count
    top_row = header_rows[0]
    sub_rows = header_rows[1:]

    # Step 1: compute the primary label per column.
    #
    # A top-row value V at column t covers columns [t, next_top_col).
    # Within that span, we inherit V to column c ONLY IF at least one
    # sub-header cell exists in [t, next_top_col) — i.e. V genuinely
    # spans over sub-cells.  Otherwise V stays at column t only (it's
    # a standalone single-row header cell with an empty neighbour, not
    # a span parent).
    top_cols = [c for c in range(col_count)
                if c < len(top_row) and top_row[c]]
    # Columns that any sub row has a value at.
    sub_cols = set()
    for sr in sub_rows:
        for c in range(min(col_count, len(sr))):
            if sr[c]:
                sub_cols.add(c)

    primary: list[str] = [""] * col_count
    for i, t in enumerate(top_cols):
        next_t = top_cols[i + 1] if i + 1 < len(top_cols) else col_count
        label = _flatten_ws(top_row[t])
        # Does this top cell span over ≥ 2 sub-cells?  If not, it stays
        # at column t alone (no inherit).
        sub_in_span = [c for c in sub_cols if t <= c < next_t]
        if len(sub_in_span) >= 2:
            for c in range(t, next_t):
                primary[c] = label
        else:
            primary[t] = label

    # Step 2: attach sub-header labels per column.  A sub-header value
    # at column c becomes a part appended after the primary label for
    # column c.
    composed: list[str] = []
    for c in range(col_count):
        parts: list[str] = []
        if primary[c]:
            parts.append(primary[c])
        for sr in sub_rows:
            v = _flatten_ws(sr[c]) if c < len(sr) and sr[c] else ""
            if v:
                parts.append(v)
        # Dedup consecutive identical parts (e.g. single-row header
        # where top == sub would duplicate).
        dedup: list[str] = []
        for p in parts:
            if not dedup or dedup[-1] != p:
                dedup.append(p)
        composed.append(SEP.join(dedup))
    return composed


def _detect_header(sheet: RawSheet, body_start: int) -> tuple[int, list[str]] | None:
    """Find the (possibly multi-row) header starting at/around ``body_start``.

    Algorithm (spec §8-3 Phase 22-B-12):
    1. Scan downward for the first row with ``_run_length ≥ 3`` — this
       is the primary header candidate.
    2. Walk downward while the next row qualifies as a span-inherit child.
    3. Walk upward while the previous row qualifies as a span-inherit
       parent (corpus: v5 bessatsu has a sparse parent row above a
       dense leaf row, not detected by the downward scan alone).
    4. Compose column names via :func:`_compose_header_columns`.

    Returns ``(data_start, column_names)`` or ``None``.
    """
    rows = sheet.rows
    n = len(rows)
    for h in range(body_start, min(body_start + 20, n)):
        row_h = rows[h]
        if _run_length(row_h) < 3:
            continue
        header_rows_idx = [h]
        # Walk downward: extend with leaf-ier rows.
        while (
            header_rows_idx[-1] + 1 < n
            and _looks_like_sub_header(
                rows[header_rows_idx[-1]], rows[header_rows_idx[-1] + 1]
            )
        ):
            header_rows_idx.append(header_rows_idx[-1] + 1)
        # Walk upward: insert sparse parent rows above the current top.
        while (
            header_rows_idx[0] - 1 >= body_start
            and _looks_like_sub_header(
                rows[header_rows_idx[0] - 1], rows[header_rows_idx[0]]
            )
        ):
            header_rows_idx.insert(0, header_rows_idx[0] - 1)
        data_start = header_rows_idx[-1] + 1
        # Need ≥ 2 data rows
        data_rows_available = 0
        for r in rows[data_start:]:
            if any(v for v in r):
                data_rows_available += 1
                if data_rows_available >= 2:
                    break
        if data_rows_available < 2:
            continue
        col_count = max((len(rows[r]) for r in header_rows_idx), default=0)
        header_matrix = [rows[r] for r in header_rows_idx]
        columns = _compose_header_columns(header_matrix, col_count)
        return data_start, columns
    return None


def _flatten_ws(s: str) -> str:
    """Collapse embedded whitespace/newlines to single spaces."""
    return " ".join(s.split())


# ---------------------------------------------------------------------------
# Sheet → RSTResult conversion
# ---------------------------------------------------------------------------

def sheet_to_result(sheet: RawSheet) -> tuple[RSTResult, dict]:
    """Convert one :class:`RawSheet` to an :class:`RSTResult`.

    Returns ``(result, meta)`` where ``meta`` carries the ``sheet_type``
    ("P1" | "P2") and, for P1, the table body (``columns`` + ``data_rows``)
    needed by docs.py to restore a MD table.  ``meta`` is serialised onto
    the output JSON verbatim via ``run._convert_and_write``.
    """
    title, body_start, title_row_extras = _extract_title(sheet)

    # Column-count ≤ 2 anywhere below the title is forced P2 (§8-2).
    useful_width = _useful_width(sheet, body_start)
    header = _detect_header(sheet, body_start) if useful_width > 2 else None

    if header is not None:
        data_start, columns = header
        header_start = _find_header_start(sheet.rows, body_start, data_start)
        preamble = _collect_preamble(
            sheet.rows, body_start, header_start, title_row_extras
        )
        sections, data_rows = _build_p1_sections(sheet, data_start, columns)
        result = RSTResult(
            title=title,
            no_knowledge_content=False,
            content=preamble,
            sections=sections,
        )
        meta = {
            "sheet_type": "P1",
            "columns": columns,
            "data_rows": data_rows,
        }
        return result, meta

    # P2: keep every body row verbatim (no flatten) so verify's raw-token
    # match succeeds.
    content = _build_p2_content(sheet, body_start, "")
    result = RSTResult(
        title=title,
        no_knowledge_content=False,
        content=content,
        sections=[],
    )
    meta = {"sheet_type": "P2"}
    return result, meta


def _find_header_start(rows: list[list[str]], body_start: int, data_start: int) -> int:
    """Locate the first row of the header block within [body_start, data_start).

    The header block ends at data_start - 1 (the leaf row).  Walking
    upward from that row, a predecessor qualifies as part of the header
    iff ``_looks_like_sub_header(pred, successor)`` — same span-inherit
    rule ``_detect_header`` uses to extend the header upward.  Rows
    above the header block are preamble.
    """
    if data_start <= body_start:
        return body_start
    header_top = data_start - 1
    while header_top - 1 >= body_start and _looks_like_sub_header(
        rows[header_top - 1], rows[header_top]
    ):
        header_top -= 1
    return header_top


def _useful_width(sheet: RawSheet, body_start: int) -> int:
    """Columns that actually contain data in the body."""
    width = sheet.ncols
    used = [False] * width
    for r in sheet.rows[body_start:]:
        for cx in range(min(width, len(r))):
            if r[cx]:
                used[cx] = True
    return sum(1 for u in used if u)


def _build_p1_sections(
    sheet: RawSheet,
    data_start: int,
    columns: list[str],
) -> tuple[list[Section], list[list[str]]]:
    """Emit one section per data row; return also the raw row matrix."""
    sections: list[Section] = []
    data_rows: list[list[str]] = []
    # Identify the "title column" — either a column named "タイトル", or
    # the first column that is not "No" / "No." / "№" / empty.
    title_col = _pick_title_col(columns)

    width = len(columns)
    for r in sheet.rows[data_start:]:
        cells = [c for c in r[:width]] + [""] * max(0, width - len(r))
        if not any(cells):
            continue
        # Forward-fill row-spanning cells: in merged-cell sheets, only the
        # top-left holds the value.  We currently do NOT forward-fill
        # (`openpyxl` already returns None for merged slaves), so each row
        # reflects what a reader sees.  This keeps verify's cell-tokens
        # model (tokens appear only in the top-left row) consistent.
        data_rows.append(cells)
        # Section title = value at title_col, with fallback.  Both branches
        # must apply `_flatten_ws` so the title matches verify's flattened
        # source token (verify emits `" ".join(title_val.split())` for the
        # same cell per spec §8-4).  Missing flatten on the fallback branch
        # surfaces as QC1/QC2 FAIL on multiline cells in spacer columns
        # (e.g. v1.4 1.4.3-releasenote `UI開発基盤\n※…`).
        raw_title = cells[title_col] if 0 <= title_col < width else ""
        section_title = _flatten_ws(raw_title or _first_non_empty(cells))
        # Section content = {col}: {val} vertical list (all non-empty cells).
        # Flatten embedded newlines in values so each `{列名}: {値}` line
        # is parseable by verify's line-split (§8-4 is line-based).
        content_lines = []
        for cx, col in enumerate(columns):
            if cx >= len(cells):
                continue
            val = cells[cx]
            if not val:
                continue
            if not col:
                # Spacer column (empty header) — skip.  Any non-empty
                # cell here is flagged by verify QP as "unexpected
                # column" or QC1 for the raw cell token, whichever
                # applies; we do not emit an ambiguous ": {val}" line.
                continue
            flat_val = _flatten_ws(val)
            content_lines.append(f"{col}: {flat_val}")
        sections.append(Section(
            title=section_title,
            content="\n".join(content_lines),
        ))
    return sections, data_rows


def _pick_title_col(columns: list[str]) -> int:
    """Choose the column index whose value we use as section title."""
    for cx, name in enumerate(columns):
        if name == "タイトル":
            return cx
    # Fallback: first column that is not a row-number column.
    for cx, name in enumerate(columns):
        if not name:
            continue
        if name in ("No", "No.", "№", "#"):
            continue
        return cx
    return 0


def _first_non_empty(cells: list[str]) -> str:
    return next((c for c in cells if c), "")


def _build_p2_content(sheet: RawSheet, body_start: int, preamble: str) -> str:
    """Flatten all body rows to text.

    Cell values are whitespace-flattened (embedded newlines / repeated
    spaces collapsed to single spaces) to match the verify-side token
    flatten.  Excel cells routinely carry display-only newlines that
    have no semantic meaning — collapsing them keeps JSON/source
    comparable.
    """
    lines: list[str] = []
    if preamble:
        lines.append(preamble)
    for r in sheet.rows[body_start:]:
        non_empty = [_flatten_ws(c) for c in r if c]
        if not non_empty:
            continue
        lines.append("  ".join(non_empty))
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Convenience: per-file convert (used by run.py when sheet_name is given)
# ---------------------------------------------------------------------------

def convert_sheet(path: Path, sheet_name: str) -> tuple[RSTResult, dict]:
    sheet = read_sheet(path, sheet_name)
    return sheet_to_result(sheet)
