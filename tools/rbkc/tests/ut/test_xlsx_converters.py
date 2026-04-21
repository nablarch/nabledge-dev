"""Unit tests for xlsx converters schema (Phase 21-D).

title="" and sections=[]. All cell content is placed in top-level content.
"""
from __future__ import annotations

from pathlib import Path

import pytest

import openpyxl


@pytest.fixture
def releasenote_xlsx(tmp_path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "バージョン"
    ws["B1"] = "変更点"
    ws["A2"] = "6.0.0"
    ws["B2"] = "初回リリース"
    path = tmp_path / "rn.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def security_xlsx(tmp_path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "対応表"
    ws["A1"] = "機能"
    ws["B1"] = "対策"
    ws["A2"] = "ログイン"
    ws["B2"] = "CSRF トークン"
    path = tmp_path / "sec.xlsx"
    wb.save(path)
    return path


class TestReleaseNoteTopLevelContent:
    def test_title_is_empty(self, releasenote_xlsx):
        from scripts.create.converters.xlsx_releasenote import convert
        result = convert(releasenote_xlsx, "releasenote")
        assert result.title == ""

    def test_sections_is_empty(self, releasenote_xlsx):
        from scripts.create.converters.xlsx_releasenote import convert
        result = convert(releasenote_xlsx, "releasenote")
        assert result.sections == []

    def test_all_cells_in_top_level_content(self, releasenote_xlsx):
        from scripts.create.converters.xlsx_releasenote import convert
        result = convert(releasenote_xlsx, "releasenote")
        assert "バージョン" in result.content
        assert "変更点" in result.content
        assert "6.0.0" in result.content
        assert "初回リリース" in result.content


class TestSecurityTopLevelContent:
    def test_title_is_empty(self, security_xlsx):
        from scripts.create.converters.xlsx_security import convert
        result = convert(security_xlsx, "security")
        assert result.title == ""

    def test_sections_is_empty(self, security_xlsx):
        from scripts.create.converters.xlsx_security import convert
        result = convert(security_xlsx, "security")
        assert result.sections == []

    def test_all_cells_in_top_level_content(self, security_xlsx):
        from scripts.create.converters.xlsx_security import convert
        result = convert(security_xlsx, "security")
        assert "ログイン" in result.content
        assert "CSRF トークン" in result.content
