"""Unit tests for ``scripts.common.sources``.

TDD: These tests define the public API contract for the new
``scripts/common/sources.py`` module.  All tests fail (ImportError) until
the module is created (Task 2).
"""
from __future__ import annotations

from pathlib import Path

import pytest


def _repo_root() -> Path:
    # tests/ut/test_sources.py → repo root is 4 levels up
    return Path(__file__).resolve().parents[4]


class TestSourceFileDataclass:
    """SourceFile must be importable from scripts.common.sources."""

    def test_sourcefile_importable(self):
        from scripts.common.sources import SourceFile  # noqa: F401

    def test_sourcefile_is_frozen_dataclass(self):
        from scripts.common.sources import SourceFile

        sf = SourceFile(
            path=Path("/tmp/foo.rst"),
            format="rst",
            filename="foo.rst",
        )
        assert sf.path == Path("/tmp/foo.rst")
        assert sf.format == "rst"
        assert sf.filename == "foo.rst"

    def test_sourcefile_is_hashable(self):
        from scripts.common.sources import SourceFile

        sf = SourceFile(path=Path("/tmp/a.rst"), format="rst", filename="a.rst")
        # frozen dataclasses are hashable
        assert hash(sf) is not None


class TestFileInfoDataclass:
    """FileInfo must be importable from scripts.common.sources."""

    def test_fileinfo_importable(self):
        from scripts.common.sources import FileInfo  # noqa: F401

    def test_fileinfo_fields(self):
        from scripts.common.sources import FileInfo

        fi = FileInfo(
            source_path=Path("/tmp/foo.rst"),
            format="rst",
            file_id="libraries-foo",
            type="component",
            category="libraries",
            output_path="component/libraries/libraries-foo.json",
        )
        assert fi.source_path == Path("/tmp/foo.rst")
        assert fi.format == "rst"
        assert fi.file_id == "libraries-foo"
        assert fi.type == "component"
        assert fi.category == "libraries"
        assert fi.output_path == "component/libraries/libraries-foo.json"
        assert fi.sheet_name is None

    def test_fileinfo_sheet_name_optional(self):
        from scripts.common.sources import FileInfo

        fi = FileInfo(
            source_path=Path("/tmp/foo.xlsx"),
            format="xlsx",
            file_id="releases-foo",
            type="releases",
            category="releases",
            output_path="releases/releases/releases-foo.json",
            sheet_name="Sheet1",
        )
        assert fi.sheet_name == "Sheet1"


class TestScanSources:
    """scan_sources must be importable from scripts.common.sources and
    produce the same results as the current scripts.create.scan.scan_sources."""

    def test_scan_sources_importable(self):
        from scripts.common.sources import scan_sources  # noqa: F401

    def test_scan_sources_files_list_returns_sourcefile_objects(self, tmp_path):
        from scripts.common.sources import scan_sources, SourceFile

        rst_file = tmp_path / "doc.rst"
        rst_file.touch()
        md_file = tmp_path / "guide.md"
        md_file.touch()
        xlsx_file = tmp_path / "data.xlsx"
        xlsx_file.touch()

        results = scan_sources("6", tmp_path, files=[
            str(rst_file.relative_to(tmp_path)),
            str(md_file.relative_to(tmp_path)),
            str(xlsx_file.relative_to(tmp_path)),
        ])

        assert len(results) == 3
        assert all(isinstance(r, SourceFile) for r in results)
        formats = {r.format for r in results}
        assert formats == {"rst", "md", "xlsx"}

    def test_scan_sources_skips_unsupported_extensions(self, tmp_path):
        from scripts.common.sources import scan_sources

        txt_file = tmp_path / "readme.txt"
        txt_file.touch()

        results = scan_sources("6", tmp_path, files=[str(txt_file)])

        assert results == []

    def test_scan_sources_full_scan_returns_sourcefile_list(self):
        """Full directory scan should return SourceFile objects for v6."""
        from scripts.common.sources import scan_sources, SourceFile

        repo = _repo_root()
        results = scan_sources("6", repo)

        assert isinstance(results, list)
        assert len(results) > 0
        assert all(isinstance(r, SourceFile) for r in results)

    def test_scan_sources_result_matches_create_scan(self):
        """common.sources.scan_sources must return same result as create.scan.scan_sources."""
        from scripts.common.sources import scan_sources as common_scan
        from scripts.create.scan import scan_sources as create_scan

        repo = _repo_root()
        common_result = sorted(common_scan("6", repo), key=lambda sf: str(sf.path))
        create_result = sorted(create_scan("6", repo), key=lambda sf: str(sf.path))

        assert len(common_result) == len(create_result)
        for c, cr in zip(common_result, create_result):
            assert c.path == cr.path
            assert c.format == cr.format
            assert c.filename == cr.filename


class TestClassifySources:
    """classify_sources must be importable from scripts.common.sources and
    produce the same results as the current scripts.create.classify.classify_sources."""

    def test_classify_sources_importable(self):
        from scripts.common.sources import classify_sources  # noqa: F401

    def test_classify_sources_returns_fileinfo_list(self):
        from scripts.common.sources import scan_sources, classify_sources, FileInfo

        repo = _repo_root()
        sources = scan_sources("6", repo)
        results = classify_sources(sources, "6", repo)

        assert isinstance(results, list)
        assert len(results) > 0
        assert all(isinstance(fi, FileInfo) for fi in results)

    def test_classify_sources_result_matches_create_classify(self):
        """common.sources.classify_sources must return same result as
        create.classify.classify_sources."""
        from scripts.common.sources import (
            scan_sources as common_scan,
            classify_sources as common_classify,
        )
        from scripts.create.scan import scan_sources as create_scan
        from scripts.create.classify import classify_sources as create_classify

        repo = _repo_root()

        common_sources = common_scan("6", repo)
        create_sources = create_scan("6", repo)

        common_result = sorted(
            common_classify(common_sources, "6", repo),
            key=lambda fi: fi.output_path,
        )
        create_result = sorted(
            create_classify(create_sources, "6", repo),
            key=lambda fi: fi.output_path,
        )

        assert len(common_result) == len(create_result)
        for c, cr in zip(common_result, create_result):
            assert c.source_path == cr.source_path
            assert c.format == cr.format
            assert c.file_id == cr.file_id
            assert c.type == cr.type
            assert c.category == cr.category
            assert c.output_path == cr.output_path
            assert c.sheet_name == cr.sheet_name


class TestListSheetNames:
    """SE-F1: list_sheet_names must be importable from scripts.common.sources (not only from
    scripts.create.converters.xlsx_common) to avoid §2-2 layering violation."""

    def test_list_sheet_names_importable_from_common_sources(self):
        from scripts.common.sources import list_sheet_names  # noqa: F401

    def test_list_sheet_names_returns_sheet_list(self, tmp_path):
        """list_sheet_names returns sheet names for a real xlsx file."""
        import openpyxl
        from scripts.common.sources import list_sheet_names

        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.create_sheet("Sheet2")
        path = tmp_path / "test.xlsx"
        wb.save(str(path))

        names = list_sheet_names(path)
        assert names == ["Sheet1", "Sheet2"]

    def test_list_sheet_names_single_sheet(self, tmp_path):
        """list_sheet_names returns a list with one entry for a single-sheet workbook."""
        import openpyxl
        from scripts.common.sources import list_sheet_names

        wb = openpyxl.Workbook()
        wb.active.title = "Only"
        path = tmp_path / "single.xlsx"
        wb.save(str(path))

        names = list_sheet_names(path)
        assert names == ["Only"]


class TestClassifySourcesEdgeCases:
    """QA-F1〜F4: edge-case tests for classify_sources and scan_sources."""

    def test_classify_sources_empty_input(self, monkeypatch):
        """QA-F2: classify_sources([]) must return []."""
        from scripts.common.sources import classify_sources
        from pathlib import Path

        monkeypatch.setattr("scripts.common.sources._load_mappings", lambda v, r: {})

        result = classify_sources([], "6", Path("."))
        assert result == []

    def test_classify_sources_single_sheet_xlsx_no_sheet_suffix(self, tmp_path, monkeypatch):
        """QA-F1: xlsx with exactly one sheet → file_id must NOT get a sheet-name suffix."""
        import openpyxl
        from scripts.common.sources import SourceFile, classify_sources
        from scripts.common.file_id import FileClass

        # Create a real xlsx file with one sheet
        wb = openpyxl.Workbook()
        wb.active.title = "Only"
        xlsx_path = tmp_path / "spec.xlsx"
        wb.save(str(xlsx_path))

        src = SourceFile(path=xlsx_path, format="xlsx", filename="spec.xlsx")

        def fake_derive(path, fmt, version, repo_root, mappings=None):
            return FileClass(
                source_path=path,
                format=fmt,
                type="component",
                category="test",
                file_id="test-spec",
                matched_pattern="",
            )

        monkeypatch.setattr("scripts.common.sources._load_mappings", lambda v, r: {})
        monkeypatch.setattr("scripts.common.sources.derive_file_id", fake_derive)

        results = classify_sources([src], "6", tmp_path)

        assert len(results) == 1
        fi = results[0]
        # Single sheet: file_id must NOT end with "-Only"
        assert fi.file_id == "test-spec"
        assert fi.sheet_name == "Only"

    def test_classify_sources_multi_sheet_xlsx_gets_sheet_suffix(self, tmp_path, monkeypatch):
        """QA-F1 complement: xlsx with 2 sheets → each FileInfo gets a sheet-name suffix."""
        import openpyxl
        from scripts.common.sources import SourceFile, classify_sources
        from scripts.common.file_id import FileClass

        wb = openpyxl.Workbook()
        wb.active.title = "A"
        wb.create_sheet("B")
        xlsx_path = tmp_path / "multi.xlsx"
        wb.save(str(xlsx_path))

        src = SourceFile(path=xlsx_path, format="xlsx", filename="multi.xlsx")

        def fake_derive(path, fmt, version, repo_root, mappings=None):
            return FileClass(
                source_path=path,
                format=fmt,
                type="component",
                category="test",
                file_id="test-multi",
                matched_pattern="",
            )

        monkeypatch.setattr("scripts.common.sources._load_mappings", lambda v, r: {})
        monkeypatch.setattr("scripts.common.sources.derive_file_id", fake_derive)

        results = classify_sources([src], "6", tmp_path)

        assert len(results) == 2
        ids = {fi.file_id for fi in results}
        assert ids == {"test-multi-A", "test-multi-B"}

    def test_scan_sources_absolute_path_input(self, tmp_path):
        """QA-F4: scan_sources with files= as absolute paths must work."""
        from scripts.common.sources import scan_sources, SourceFile

        rst_file = tmp_path / "doc.rst"
        rst_file.touch()

        results = scan_sources("6", tmp_path, files=[str(rst_file)])  # absolute path

        assert len(results) == 1
        assert results[0].format == "rst"
        assert results[0].path == rst_file

    def test_scan_sources_v1x_iterdir_path(self, tmp_path):
        """QA-F3: scan_sources for v1.x version uses iterdir() of v_dir."""
        from scripts.common.sources import scan_sources

        # Build a minimal v1.4-style directory tree
        v_dir = tmp_path / ".lw" / "nab-official" / "v1.4"
        subdir = v_dir / "nablarch-fw-web"
        subdir.mkdir(parents=True)
        rst_file = subdir / "feature.rst"
        rst_file.touch()

        # scan_sources with a non-existent mapping (files= path, so no mapping needed)
        results = scan_sources("1.4", tmp_path, files=[
            str(rst_file.relative_to(tmp_path)),
        ])
        assert len(results) == 1
        assert results[0].format == "rst"


class TestSheetSlug:
    """_sheet_slug must be importable from scripts.common.sources."""

    def test_sheet_slug_importable(self):
        from scripts.common.sources import _sheet_slug  # noqa: F401

    def test_sheet_slug_replaces_path_separators(self):
        from scripts.common.sources import _sheet_slug

        assert "/" not in _sheet_slug("Sheet/Name")
        assert "\\" not in _sheet_slug("Sheet\\Name")

    def test_sheet_slug_preserves_japanese(self):
        from scripts.common.sources import _sheet_slug

        result = _sheet_slug("設計書")
        assert result == "設計書"

    def test_sheet_slug_empty_string_returns_sheet(self):
        from scripts.common.sources import _sheet_slug

        # All-whitespace input → strip() → empty string → "sheet" fallback
        result = _sheet_slug("   ")
        assert result == "sheet"

    def test_sheet_slug_slash_becomes_underscore(self):
        from scripts.common.sources import _sheet_slug

        # "/" is replaced with "_" — not stripped, so result is "_" not "sheet"
        result = _sheet_slug("/")
        assert result == "_"
