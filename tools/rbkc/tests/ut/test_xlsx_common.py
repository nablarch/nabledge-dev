"""Tests for scripts/create/converters/xlsx_common.py — P1-merged grouping.

Spec reference: tools/rbkc/docs/rbkc-converter-design.md §8-4 P1-merged
"""
from __future__ import annotations

import pytest

from scripts.create.converters.xlsx_common import RawSheet, _build_p1_sections


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _make_sheet(rows: list[list[str]], name: str = "Test") -> RawSheet:
    """Build a RawSheet from a 2D string list."""
    width = max((len(r) for r in rows), default=0)
    padded = [r + [""] * (width - len(r)) for r in rows]
    return RawSheet(name=name, rows=padded)


# ---------------------------------------------------------------------------
# P1-merged: basic grouping
# ---------------------------------------------------------------------------

class TestBuildP1SectionsP1Merged:
    """_build_p1_sections with merge_groups produces group-level sections."""

    def test_two_groups_two_rows_each(self):
        """Two merge groups of 2 rows each → 2 sections, 2 data_rows entries."""
        # data rows start at index 0 in the sheet (header already stripped)
        # columns: ["タイトル", "説明", "チェック"]
        # group A: rows 0-1 (merged title "脆弱性A")
        # group B: rows 2-3 (merged title "脆弱性B")
        rows = [
            ["脆弱性A", "対策1", "済"],
            ["",        "対策2", "未"],
            ["脆弱性B", "対策3", "済"],
            ["",        "対策4", "未"],
        ]
        sheet = _make_sheet(rows)
        columns = ["タイトル", "説明", "チェック"]
        merge_groups = [(0, 1), (2, 3)]  # (start_row, end_row) 0-indexed in sheet.rows

        sections, data_rows = _build_p1_sections(
            sheet, data_start=0, columns=columns, merge_groups=merge_groups
        )

        assert len(sections) == 2
        assert len(data_rows) == 2

    def test_group_title_is_head_row_title(self):
        """section.title = タイトル列 value of the first row in the group."""
        rows = [
            ["脆弱性A", "対策1", "済"],
            ["",        "対策2", "未"],
            ["脆弱性B", "対策3", "済"],
            ["",        "対策4", "未"],
        ]
        sheet = _make_sheet(rows)
        columns = ["タイトル", "説明", "チェック"]
        merge_groups = [(0, 1), (2, 3)]

        sections, _ = _build_p1_sections(
            sheet, data_start=0, columns=columns, merge_groups=merge_groups
        )

        assert sections[0].title == "脆弱性A"
        assert sections[1].title == "脆弱性B"

    def test_section_content_includes_all_rows_in_group(self):
        """section.content must include {col}: {val} for every row in the group."""
        rows = [
            ["脆弱性A", "対策1", "済"],
            ["",        "対策2", "未"],
        ]
        sheet = _make_sheet(rows)
        columns = ["タイトル", "説明", "チェック"]
        merge_groups = [(0, 1)]

        sections, _ = _build_p1_sections(
            sheet, data_start=0, columns=columns, merge_groups=merge_groups
        )

        content = sections[0].content
        assert "説明: 対策1" in content
        assert "説明: 対策2" in content
        assert "チェック: 済" in content
        assert "チェック: 未" in content

    def test_data_rows_contains_only_head_rows(self):
        """data_rows must contain only the first row of each group."""
        rows = [
            ["脆弱性A", "対策1", "済"],
            ["",        "対策2", "未"],
            ["脆弱性B", "対策3", "済"],
            ["",        "対策4", "未"],
        ]
        sheet = _make_sheet(rows)
        columns = ["タイトル", "説明", "チェック"]
        merge_groups = [(0, 1), (2, 3)]

        _, data_rows = _build_p1_sections(
            sheet, data_start=0, columns=columns, merge_groups=merge_groups
        )

        # data_rows must have 2 entries, each being the head row of the group
        assert len(data_rows) == 2
        assert data_rows[0][0] == "脆弱性A"
        assert data_rows[1][0] == "脆弱性B"
        # Tail rows must NOT appear as separate data_rows entries
        assert all(r[0] != "" for r in data_rows)


# ---------------------------------------------------------------------------
# P1-merged: edge cases
# ---------------------------------------------------------------------------

class TestBuildP1SectionsP1MergedEdgeCases:
    """Edge cases for P1-merged grouping."""

    def test_single_row_group(self):
        """A group with only 1 row (no merge) → still 1 section."""
        rows = [
            ["脆弱性A", "対策1", "済"],
        ]
        sheet = _make_sheet(rows)
        columns = ["タイトル", "説明", "チェック"]
        merge_groups = [(0, 0)]

        sections, data_rows = _build_p1_sections(
            sheet, data_start=0, columns=columns, merge_groups=merge_groups
        )

        assert len(sections) == 1
        assert sections[0].title == "脆弱性A"
        assert len(data_rows) == 1

    def test_mixed_group_sizes(self):
        """Groups of different sizes are handled independently."""
        # group A: 1 row; group B: 3 rows; group C: 2 rows
        rows = [
            ["A", "a1", "x"],
            ["B", "b1", "x"],
            ["",  "b2", "y"],
            ["",  "b3", "z"],
            ["C", "c1", "x"],
            ["",  "c2", "y"],
        ]
        sheet = _make_sheet(rows)
        columns = ["タイトル", "詳細", "チェック"]
        merge_groups = [(0, 0), (1, 3), (4, 5)]

        sections, data_rows = _build_p1_sections(
            sheet, data_start=0, columns=columns, merge_groups=merge_groups
        )

        assert len(sections) == 3
        assert sections[0].title == "A"
        assert sections[1].title == "B"
        assert sections[2].title == "C"
        assert len(data_rows) == 3

        # Group B content must include all 3 rows
        b_content = sections[1].content
        assert "詳細: b1" in b_content
        assert "詳細: b2" in b_content
        assert "詳細: b3" in b_content

    def test_empty_tail_rows_in_group_are_skipped(self):
        """All-empty rows within a group must not produce {col}: lines."""
        rows = [
            ["脆弱性A", "対策1", "済"],
            ["",        "",      ""],   # all-empty tail row
        ]
        sheet = _make_sheet(rows)
        columns = ["タイトル", "説明", "チェック"]
        merge_groups = [(0, 1)]

        sections, data_rows = _build_p1_sections(
            sheet, data_start=0, columns=columns, merge_groups=merge_groups
        )

        # Still 1 section, 1 data_rows entry
        assert len(sections) == 1
        assert len(data_rows) == 1
        # Content only from non-empty cells
        content = sections[0].content
        assert "タイトル: 脆弱性A" in content
        assert "説明: 対策1" in content
        assert "チェック: 済" in content

    def test_no_merge_groups_falls_back_to_row_per_section(self):
        """merge_groups=None → original P1 behaviour (1 row = 1 section)."""
        rows = [
            ["脆弱性A", "対策1"],
            ["脆弱性B", "対策2"],
        ]
        sheet = _make_sheet(rows)
        columns = ["タイトル", "説明"]

        sections, data_rows = _build_p1_sections(
            sheet, data_start=0, columns=columns
        )

        assert len(sections) == 2
        assert len(data_rows) == 2
        assert sections[0].title == "脆弱性A"
        assert sections[1].title == "脆弱性B"


# ---------------------------------------------------------------------------
# load_sheet_subtype_map: P1-merged entries are returned
# ---------------------------------------------------------------------------

class TestLoadSheetSubtypeMapP1Merged:
    """load_sheet_subtype_map must parse P1-merged entries from mapping file."""

    def test_p1_merged_entry_returned(self, tmp_path):
        """A P1-merged row in the mapping file must be included in the result."""
        from scripts.create.converters.xlsx_common import load_sheet_subtype_map

        mapping_content = (
            "| ファイル名 | シート名 | サブタイプ |\n"
            "|-----------|----------|------------|\n"
            "| security-check-2.xlsx | 2.チェックリスト | P1-merged |\n"
            "| release-note.xlsx     | 1.リリースノート | P2-1      |\n"
        )
        mapping_file = tmp_path / "xlsx-sheet-mapping.md"
        mapping_file.write_text(mapping_content, encoding="utf-8")

        result = load_sheet_subtype_map(mapping_file)

        assert ("security-check-2.xlsx", "2.チェックリスト") in result
        assert result[("security-check-2.xlsx", "2.チェックリスト")] == "P1-merged"

    def test_p2_1_entry_still_returned(self, tmp_path):
        """Existing P2-1 entries must remain in the result alongside P1-merged."""
        from scripts.create.converters.xlsx_common import load_sheet_subtype_map

        mapping_content = (
            "| ファイル名 | シート名 | サブタイプ |\n"
            "|-----------|----------|------------|\n"
            "| security-check-2.xlsx | 2.チェックリスト | P1-merged |\n"
            "| release-note.xlsx     | 1.リリースノート | P2-1      |\n"
        )
        mapping_file = tmp_path / "xlsx-sheet-mapping.md"
        mapping_file.write_text(mapping_content, encoding="utf-8")

        result = load_sheet_subtype_map(mapping_file)

        assert ("release-note.xlsx", "1.リリースノート") in result
        assert result[("release-note.xlsx", "1.リリースノート")] == "P2-1"


# ---------------------------------------------------------------------------
# sheet_to_result: P1-merged subtype flows through to meta
# ---------------------------------------------------------------------------

class TestSheetToResultP1Merged:
    """sheet_to_result with sheet_subtype='P1-merged' must produce correct meta."""

    def _make_p1_merged_xlsx(self, tmp_path) -> "Path":
        """Create a minimal xlsx with merged title column cells."""
        import openpyxl
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "2.チェックリスト"

        # Row 1: ■タイトル行
        ws.cell(1, 1, "■セキュリティチェックリスト")

        # Row 2: ヘッダ行
        ws.cell(2, 1, "タイトル")
        ws.cell(2, 2, "対策")
        ws.cell(2, 3, "チェック")

        # Rows 3-4: group A (merged タイトル column A3:A4)
        ws.cell(3, 1, "脆弱性A")
        ws.cell(3, 2, "対策1")
        ws.cell(3, 3, "済")
        ws.cell(4, 1, "")
        ws.cell(4, 2, "対策2")
        ws.cell(4, 3, "未")
        ws.merge_cells("A3:A4")

        # Rows 5-6: group B (merged タイトル column A5:A6)
        ws.cell(5, 1, "脆弱性B")
        ws.cell(5, 2, "対策3")
        ws.cell(5, 3, "済")
        ws.cell(6, 1, "")
        ws.cell(6, 2, "対策4")
        ws.cell(6, 3, "未")
        ws.merge_cells("A5:A6")

        path = tmp_path / "security-check-2.xlsx"
        wb.save(str(path))
        return path

    def test_meta_sheet_subtype_is_p1_merged(self, tmp_path):
        """meta['sheet_subtype'] must be 'P1-merged' when subtype is passed."""
        from scripts.create.converters.xlsx_common import read_sheet, sheet_to_result

        path = self._make_p1_merged_xlsx(tmp_path)
        sheet = read_sheet(path, "2.チェックリスト")
        _, meta = sheet_to_result(sheet, sheet_subtype="P1-merged")

        assert meta.get("sheet_subtype") == "P1-merged"

    def test_meta_sheet_type_is_p1(self, tmp_path):
        """meta['sheet_type'] must remain 'P1' for P1-merged sheets."""
        from scripts.create.converters.xlsx_common import read_sheet, sheet_to_result

        path = self._make_p1_merged_xlsx(tmp_path)
        sheet = read_sheet(path, "2.チェックリスト")
        _, meta = sheet_to_result(sheet, sheet_subtype="P1-merged")

        assert meta.get("sheet_type") == "P1"

    def test_sections_count_equals_group_count(self, tmp_path):
        """sections count must equal number of merge groups (2), not data rows (4)."""
        from scripts.create.converters.xlsx_common import read_sheet, sheet_to_result

        path = self._make_p1_merged_xlsx(tmp_path)
        sheet = read_sheet(path, "2.チェックリスト")
        result, meta = sheet_to_result(sheet, sheet_subtype="P1-merged")

        assert len(result.sections) == 2

    def test_data_rows_count_equals_group_count(self, tmp_path):
        """data_rows count must equal sections count (groups), not total data rows."""
        from scripts.create.converters.xlsx_common import read_sheet, sheet_to_result

        path = self._make_p1_merged_xlsx(tmp_path)
        sheet = read_sheet(path, "2.チェックリスト")
        _, meta = sheet_to_result(sheet, sheet_subtype="P1-merged")

        assert len(meta["data_rows"]) == 2

    def test_section_titles_from_head_rows(self, tmp_path):
        """section titles must come from the head row of each merge group."""
        from scripts.create.converters.xlsx_common import read_sheet, sheet_to_result

        path = self._make_p1_merged_xlsx(tmp_path)
        sheet = read_sheet(path, "2.チェックリスト")
        result, _ = sheet_to_result(sheet, sheet_subtype="P1-merged")

        titles = [s.title for s in result.sections]
        assert titles == ["脆弱性A", "脆弱性B"]

    def test_section_content_includes_all_group_rows(self, tmp_path):
        """section.content must include tokens from every row in the group."""
        from scripts.create.converters.xlsx_common import read_sheet, sheet_to_result

        path = self._make_p1_merged_xlsx(tmp_path)
        sheet = read_sheet(path, "2.チェックリスト")
        result, _ = sheet_to_result(sheet, sheet_subtype="P1-merged")

        # Section 0 (脆弱性A) must contain both row-3 and row-4 tokens
        content_a = result.sections[0].content
        assert "対策: 対策1" in content_a
        assert "対策: 対策2" in content_a
        assert "チェック: 済" in content_a
        assert "チェック: 未" in content_a
