"""Tests for scripts/verify/verify.py — rebuilt per rbkc-verify-quality-design.md."""
from __future__ import annotations

import json
import textwrap
from pathlib import Path

import pytest


# ---------------------------------------------------------------------------
# QO1: docs MD 構造整合性
# ---------------------------------------------------------------------------

class TestCheckJsonDocsMdConsistency_QO1:
    """QO1: docs MD structure (title, section titles, order) must match JSON."""

    def _check(self, data, docs_md_text):
        from scripts.verify.verify import check_json_docs_md_consistency
        return check_json_docs_md_consistency(data, docs_md_text)

    def test_pass_title_and_sections_match(self):
        data = {
            "id": "f", "title": "タイトル", "content": "", "sections": [
                {"id": "s1", "title": "概要", "content": "説明"},
                {"id": "s2", "title": "設定", "content": "設定内容"},
            ]
        }
        docs = "# タイトル\n\n## 概要\n\n説明\n\n## 設定\n\n設定内容\n"
        assert self._check(data, docs) == []

    def test_fail_title_mismatch(self):
        data = {"id": "f", "title": "正しいタイトル", "content": "", "sections": []}
        docs = "# 別のタイトル\n\n"
        issues = self._check(data, docs)
        assert any("QO1" in i and "title" in i for i in issues)

    def test_fail_section_title_missing(self):
        data = {
            "id": "f", "title": "T", "content": "", "sections": [
                {"id": "s1", "title": "存在しないセクション", "content": "c"},
            ]
        }
        docs = "# T\n\n## 別のセクション\n\nc\n"
        issues = self._check(data, docs)
        assert any("QO1" in i for i in issues)

    def test_fail_section_order_reversed(self):
        data = {
            "id": "f", "title": "T", "content": "", "sections": [
                {"id": "s1", "title": "A", "content": "a"},
                {"id": "s2", "title": "B", "content": "b"},
            ]
        }
        # B appears before A in docs MD — order mismatch
        docs = "# T\n\n## B\n\nb\n\n## A\n\na\n"
        issues = self._check(data, docs)
        assert any("QO1" in i for i in issues)

    def test_pass_no_knowledge_content_skipped(self):
        data = {"id": "f", "title": "T", "no_knowledge_content": True, "sections": []}
        assert self._check(data, "# 全然違う\n") == []

    def test_pass_no_sections_no_h2(self):
        """sections=[] means no ## headings expected in docs MD."""
        data = {"id": "f", "title": "T", "content": "本文です。", "sections": []}
        docs = "# T\n\n本文です。\n"
        assert self._check(data, docs) == []

    def test_fail_extra_h2_in_docs_md(self):
        """docs MD has ## heading but JSON has no sections → QO1 FAIL."""
        data = {"id": "f", "title": "T", "content": "本文です。", "sections": []}
        docs = "# T\n\n## 余計なセクション\n\n本文です。\n"
        issues = self._check(data, docs)
        assert any("QO1" in i for i in issues)

    # --- QO1 Z-1 gap fill -------------------------------------------------

    def test_fail_h1_missing(self):
        data = {"id": "f", "title": "タイトル", "content": "", "sections": []}
        docs = "本文だけで h1 が無い。\n"
        issues = self._check(data, docs)
        assert any("QO1" in i for i in issues)

    def test_fail_top_level_content_below_first_h2_not_directly_under_h1(self):
        """Spec §3-3 QO2: top-level content must sit between `#` and the
        first `##`. Content appearing only *after* the first `##` FAILs."""
        data = {"id": "f", "title": "T", "content": "トップ本文。",
                "sections": [{"id": "s1", "title": "概要", "content": "概要本文"}]}
        # Top-level content is below the first ## — wrong location.
        docs = "# T\n\n## 概要\n\nトップ本文。\n\n概要本文\n"
        issues = self._check(data, docs)
        assert any("QO2" in i and "top-level" in i for i in issues)

    def test_fail_extra_h2_when_sections_present(self):
        """Spec §3-3 QO1: docs MD section titles must match JSON order and
        count. Extra H2 in docs MD beyond JSON sections FAILs."""
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "A", "content": "a"},
        ]}
        docs = "# T\n\n## A\n\na\n\n## 余計\n\nextra\n"
        issues = self._check(data, docs)
        assert any("QO1" in i for i in issues)

    def test_fail_duplicate_h2_order_violation(self):
        """Spec §3-3 QO1 section order requirement: JSON [A,B] vs docs
        [A,B,A] — the second A appears out of order relative to B. The
        current greedy matcher passes this; this test flags the gap."""
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "B", "content": "b"},
            {"id": "s2", "title": "A", "content": "a2"},
        ]}
        docs = "# T\n\n## A\n\na1\n\n## B\n\nb\n"
        # JSON requires B before A, docs shows A before B → QO1 FAIL
        issues = self._check(data, docs)
        assert any("QO1" in i for i in issues)

    def test_fail_readme_missing_page_declaration(self, tmp_path):
        """Spec §3-3 QO3: README.md must contain 'N ページ' declaration."""
        # Reuse the QO3 test class convenience — inline setup.
        from scripts.verify.verify import check_docs_coverage
        kdir = tmp_path / "knowledge"; kdir.mkdir()
        ddir = tmp_path / "docs"; ddir.mkdir()
        (kdir / "a.json").write_text(json.dumps({"title": "A"}))
        (ddir / "a.md").write_text("# A\n")
        (ddir / "README.md").write_text("目次\n---\n- [A](a.md)\n")  # no 'N ページ'
        issues = check_docs_coverage(kdir, ddir)
        assert any("QO3" in i and "ページ" in i for i in issues)

    def test_fail_empty_title_vs_nonempty_docs_h1(self):
        """Spec §3-3: JSON title must equal docs MD H1. An empty JSON title
        against a non-empty H1 does not match → FAIL."""
        data = {"id": "f", "title": "", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": ""}
        ]}
        docs = "# 何か\n\n## 概要\n"
        issues = self._check(data, docs)
        assert any("QO1" in i and "title mismatch" in i for i in issues)

    def test_title_with_markdown_special_chars_exact_match(self):
        data = {"id": "f", "title": "A `code` & B", "content": "", "sections": []}
        docs = "# A `code` & B\n"
        assert self._check(data, docs) == []

    def test_multiple_h1_in_docs_md_first_wins(self):
        """Multiple `#` lines in docs MD: matcher should use the first one."""
        data = {"id": "f", "title": "一番目", "content": "", "sections": []}
        docs = "# 一番目\n\n追加段落\n\n# 二番目\n"
        # First `#` matches the JSON title — this is the spec's happy path.
        assert self._check(data, docs) == []

    def test_pass_tilde_fenced_code_block_with_heading_inside(self):
        """CommonMark tilde-fenced code blocks (~~~) must also be stripped
        before scanning section titles — `##` inside a ~~~ fence is
        content, not a section marker."""
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "A", "content": "a"},
        ]}
        docs = (
            "# T\n\n"
            "## A\n\n"
            "a\n\n"
            "~~~md\n"
            "## fake heading inside tilde fence\n"
            "~~~\n"
        )
        # `## fake ...` sits inside a ~~~ fenced block so it is content;
        # QO1 must not report an extra section.
        assert self._check(data, docs) == []

    def test_pass_backtick_fenced_code_block_with_heading_inside(self):
        """Same invariant for triple-backtick fences (regression guard)."""
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "A", "content": "a"},
        ]}
        docs = (
            "# T\n\n"
            "## A\n\n"
            "a\n\n"
            "```md\n"
            "## fake heading inside backtick fence\n"
            "```\n"
        )
        assert self._check(data, docs) == []

    # --- Z-1 r7 Findings: QO1 regex refinement ---------------------------

    def test_pass_section_with_h3_subheading_in_content(self):
        """Z-1 r7 QO1 F1: `###` inside a section's content is a valid
        CommonMark subheading, not a section title. The 'extra direction'
        check must use `##` only so content-level `###` does not produce
        a spurious QO1 FAIL."""
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "セクション A",
             "content": "本文。\n\n### 小見出し\n\n詳細。"},
        ]}
        docs = "# T\n\n## セクション A\n\n本文。\n\n### 小見出し\n\n詳細。\n"
        assert self._check(data, docs) == []

    def test_pass_sections_empty_with_h3_in_top_content(self):
        """Z-1 r7 QO1 F2: sections=[] + top content containing `### foo`.
        The empty-section guard must not fire on content-level subheadings."""
        data = {"id": "f", "title": "T",
                "content": "前文。\n\n### 注記\n\n注記本文。", "sections": []}
        docs = "# T\n\n前文。\n\n### 注記\n\n注記本文。\n"
        assert self._check(data, docs) == []

    def test_pass_section_title_rendered_at_h3(self):
        """Z-1 r8 QO1 F1: spec §3-3 "JSON 各セクションのタイトルが docs MD
        の `##`/`###` に存在し、かつ JSON と同じ順序で並んでいる" permits
        a section title at `###`. When docs.py (or a future emitter)
        renders a section at `###`, QO1 must accept it. The previous
        `##`-only equality emitted a spurious 'order differs' FAIL."""
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "セクション A", "content": "a"},
            {"id": "s2", "title": "セクション B", "content": "b"},
        ]}
        # Second section rendered at ### (spec-sanctioned); docs.py's
        # current emitter uses ## only, but the verify check must accept
        # either level per the spec wording.
        docs = (
            "# T\n\n"
            "## セクション A\n\n"
            "a\n\n"
            "### セクション B\n\n"
            "b\n"
        )
        assert self._check(data, docs) == []


# ---------------------------------------------------------------------------
# QO2: docs MD 本文整合性
# ---------------------------------------------------------------------------

class TestCheckJsonDocsMdConsistency_QO2:
    """QO2: JSON content must appear verbatim in docs MD."""

    def _check(self, data, docs_md_text):
        from scripts.verify.verify import check_json_docs_md_consistency
        return check_json_docs_md_consistency(data, docs_md_text)

    def test_pass_top_content_in_docs(self):
        data = {"id": "f", "title": "T", "content": "トップレベル本文。", "sections": []}
        docs = "# T\n\nトップレベル本文。\n"
        assert self._check(data, docs) == []

    def test_fail_top_content_missing(self):
        data = {"id": "f", "title": "T", "content": "重要な本文。", "sections": []}
        docs = "# T\n\n全然違う内容。\n"
        issues = self._check(data, docs)
        assert any("QO2" in i for i in issues)

    def test_pass_section_content_in_docs(self):
        data = {
            "id": "f", "title": "T", "content": "", "sections": [
                {"id": "s1", "title": "概要", "content": "概要の説明。"},
            ]
        }
        docs = "# T\n\n## 概要\n\n概要の説明。\n"
        assert self._check(data, docs) == []

    def test_fail_section_content_missing(self):
        data = {
            "id": "f", "title": "T", "content": "", "sections": [
                {"id": "s1", "title": "概要", "content": "欠落する説明。"},
            ]
        }
        docs = "# T\n\n## 概要\n\n別の説明。\n"
        issues = self._check(data, docs)
        assert any("QO2" in i and "概要" in i for i in issues)

    def test_pass_assets_link_rewrite_symmetric(self, tmp_path):
        """JSON `assets/...` links are rewritten to `<rel>/assets/...` by
        docs.py. QO2's verbatim check applies the SAME transformation
        and then asserts the rewritten string appears in docs MD.

        Z-1 r7 QO2 F1: the expected docs body is GENERATED by calling
        docs._rewrite_asset_links — not hand-coded. Hand-coding the
        expected string makes the test circular (the same human authored
        both the fixture and the expected; drift between verify's rewrite
        copy and docs.py's rewrite goes undetected). Generating via the
        docs-side implementation closes that loop.
        """
        kdir = tmp_path / "knowledge"
        kdir.mkdir()
        docs_dir = tmp_path / "docs" / "guide"
        docs_dir.mkdir(parents=True)
        docs_md_path = docs_dir / "x.md"
        data = {
            "id": "f", "title": "T", "content": "", "sections": [
                {"id": "s1", "title": "図", "content": "![図](assets/img.png)"},
            ]
        }
        # Generate the expected docs body via docs._rewrite_asset_links so
        # drift between the two rewrite implementations surfaces here.
        from scripts.create.docs import _rewrite_asset_links
        rewritten = _rewrite_asset_links(
            data["sections"][0]["content"], docs_md_path, kdir,
        )
        docs = f"# T\n\n## 図\n\n{rewritten}\n"
        from scripts.verify.verify import check_json_docs_md_consistency
        issues = check_json_docs_md_consistency(
            data, docs, docs_md_path=docs_md_path, knowledge_dir=kdir,
        )
        assert issues == []

    def test_verify_and_docs_rewrite_agree_on_matrix(self, tmp_path):
        """Z-1 r7 QO2 F4: verify._apply_asset_link_rewrite and
        docs._rewrite_asset_links are independent copies (import across
        them is forbidden by the rbkc independence principle). This test
        imports docs._rewrite_asset_links ONLY inside the test, runs both
        on a matrix of inputs, and asserts equal output — catches drift
        between the two copies."""
        from scripts.verify.verify import _apply_asset_link_rewrite
        from scripts.create.docs import _rewrite_asset_links
        kdir = tmp_path / "knowledge"
        kdir.mkdir()
        docs_dir = tmp_path / "docs" / "guide"
        docs_dir.mkdir(parents=True)
        docs_md_path = docs_dir / "x.md"

        cases = [
            "",
            "no links here",
            "![図](assets/img.png)",
            "[text](assets/data.csv) と ![alt](assets/photo.jpg) 混在",
            "![](assets/nested/deep/file.png)",
            "![a](assets/a.png)\n\n![b](assets/b.png)\n",
            "[link](https://example.com) 外部は変換しない",
            "`assets/literal-in-backticks` は非対象ではないかもしれない",
        ]
        for content in cases:
            v = _apply_asset_link_rewrite(content, docs_md_path, kdir)
            d = _rewrite_asset_links(content, docs_md_path, kdir)
            assert v == d, f"drift for input {content!r}: verify={v!r} docs={d!r}"

    def test_fail_assets_link_rewrite_missing_from_docs(self, tmp_path):
        """If docs MD lacks the rewritten asset link, QO2 must FAIL
        (no silent skip for assets/ content)."""
        kdir = tmp_path / "knowledge"
        kdir.mkdir()
        docs_dir = tmp_path / "docs" / "guide"
        docs_dir.mkdir(parents=True)
        docs_md_path = docs_dir / "x.md"
        data = {
            "id": "f", "title": "T", "content": "", "sections": [
                {"id": "s1", "title": "図", "content": "![図](assets/img.png)"},
            ]
        }
        docs = "# T\n\n## 図\n\n画像なし\n"
        from scripts.verify.verify import check_json_docs_md_consistency
        issues = check_json_docs_md_consistency(
            data, docs, docs_md_path=docs_md_path, knowledge_dir=kdir,
        )
        assert any("QO2" in i and "図" in i for i in issues)

    # --- QO2 Z-1 gap fill -------------------------------------------------

    def test_fail_whitespace_only_diff(self):
        """Whitespace-only deviation between JSON and docs MD is still a
        FAIL: spec §3-3 requires verbatim match."""
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": "本文\n\n続き"},
        ]}
        docs = "# T\n\n## 概要\n\n本文 続き\n"  # newline replaced by space
        issues = self._check(data, docs)
        assert any("QO2" in i for i in issues)

    def test_pass_top_content_with_fenced_h2_inside_top_region(self):
        """Z-1 r7 QO2 F3: top-level content contains `## not a heading`
        inside a fenced code block. The region-bounding (first `##`
        after the H1) must mask fences so the fake `##` does not
        truncate the region prematurely — otherwise the top content
        appears to be missing."""
        data = {"id": "f", "title": "T",
                "content": "説明:\n\n```md\n## not a heading\n```",
                "sections": [{"id": "s1", "title": "本題", "content": "本文。"}]}
        docs = (
            "# T\n\n"
            "説明:\n\n"
            "```md\n"
            "## not a heading\n"
            "```\n\n"
            "## 本題\n\n"
            "本文。\n"
        )
        assert self._check(data, docs) == []

    def test_fail_top_content_with_fenced_h2_missing_from_docs(self):
        """FAIL twin of the above: if the fenced block is missing from
        docs MD, QO2 must flag the top content as not verbatim."""
        data = {"id": "f", "title": "T",
                "content": "説明:\n\n```md\n## not a heading\n```",
                "sections": [{"id": "s1", "title": "本題", "content": "本文。"}]}
        docs = (
            "# T\n\n"
            "説明:\n\n"
            "## 本題\n\n"
            "本文。\n"
        )
        issues = self._check(data, docs)
        assert any("QO2" in i for i in issues)

    def test_pass_section_content_with_fenced_code(self):
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "コード例", "content": "```java\nint x;\n```"},
        ]}
        docs = "# T\n\n## コード例\n\n```java\nint x;\n```\n"
        assert self._check(data, docs) == []

    def test_pass_section_content_with_md_special_chars(self):
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "式", "content": "評価式は `a | b` と記述。"},
        ]}
        docs = "# T\n\n## 式\n\n評価式は `a | b` と記述。\n"
        assert self._check(data, docs) == []

    def test_fail_multiple_sections_one_content_wrong(self):
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "A", "content": "A content"},
            {"id": "s2", "title": "B", "content": "B content"},
            {"id": "s3", "title": "C", "content": "C content"},
        ]}
        # middle section content differs
        docs = "# T\n\n## A\n\nA content\n\n## B\n\nWRONG\n\n## C\n\nC content\n"
        issues = self._check(data, docs)
        assert any("QO2" in i and "B" in i for i in issues)


# ---------------------------------------------------------------------------
# QO4: index.toon 網羅性
# ---------------------------------------------------------------------------

class TestCheckIndexCoverage:
    """QO4: every JSON without no_knowledge_content must be in index.toon."""

    def _check(self, knowledge_dir, index_path):
        from scripts.verify.verify import check_index_coverage
        return check_index_coverage(knowledge_dir, index_path)

    def _write_toon(self, idx_path, entries):
        """Write index.toon in real TOON format: comma-separated, indented rows."""
        lines = [
            "# Nabledge-6 Knowledge Index",
            "",
            f"files[{len(entries)},]{{title,type,category,processing_patterns,path}}:",
        ]
        for e in entries:
            lines.append(f"  {', '.join(e)}")
        idx_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    def test_pass_all_files_indexed(self, tmp_path):
        kdir = tmp_path / "knowledge"
        kdir.mkdir()
        (kdir / "a.json").write_text(json.dumps({"id": "a", "title": "A", "no_knowledge_content": False}))
        idx = tmp_path / "index.toon"
        self._write_toon(idx, [["A", "", "", "", "a.json"]])
        assert self._check(kdir, idx) == []

    def test_fail_json_not_in_index(self, tmp_path):
        kdir = tmp_path / "knowledge"
        kdir.mkdir()
        (kdir / "missing.json").write_text(json.dumps({"id": "m", "title": "M", "no_knowledge_content": False}))
        idx = tmp_path / "index.toon"
        self._write_toon(idx, [])
        issues = self._check(kdir, idx)
        assert any("QO4" in i and "missing.json" in i for i in issues)

    def test_pass_no_knowledge_content_excluded(self, tmp_path):
        kdir = tmp_path / "knowledge"
        kdir.mkdir()
        (kdir / "toc.json").write_text(json.dumps({"id": "t", "title": "T", "no_knowledge_content": True}))
        idx = tmp_path / "index.toon"
        self._write_toon(idx, [])
        assert self._check(kdir, idx) == []

    def test_pass_nested_path_indexed(self, tmp_path):
        """Nested JSON with commas in title must parse correctly from TOON."""
        kdir = tmp_path / "knowledge"
        (kdir / "sub").mkdir(parents=True)
        (kdir / "sub" / "b.json").write_text(json.dumps({"id": "b", "title": "B"}))
        idx = tmp_path / "index.toon"
        self._write_toon(idx, [["B", "about", "sub", "", "sub/b.json"]])
        assert self._check(kdir, idx) == []

    def test_fail_missing_index_file(self, tmp_path):
        kdir = tmp_path / "knowledge"
        kdir.mkdir()
        (kdir / "a.json").write_text(json.dumps({"id": "a", "title": "A", "no_knowledge_content": False}))
        issues = self._check(kdir, tmp_path / "nonexistent.toon")
        assert any("QO4" in i for i in issues)

    # --- QO4 Z-1 gap fill -------------------------------------------------

    def test_fail_missing_index_lists_every_content_json(self, tmp_path):
        """Spec §3-3: when index.toon is absent, every content JSON is FAIL."""
        kdir = tmp_path / "knowledge"
        kdir.mkdir()
        (kdir / "a.json").write_text(json.dumps({"id": "a", "title": "A"}))
        (kdir / "b.json").write_text(json.dumps({"id": "b", "title": "B"}))
        (kdir / "no.json").write_text(json.dumps({"id": "no", "title": "N", "no_knowledge_content": True}))
        issues = self._check(kdir, tmp_path / "missing.toon")
        # Two content JSONs must both be reported
        assert any("a.json" in i for i in issues)
        assert any("b.json" in i for i in issues)
        # no_knowledge file is not reported
        assert not any("no.json" in i for i in issues)

    def test_fail_dangling_entry_in_index(self, tmp_path):
        """Spec §3-3: an index entry without a matching JSON on disk FAILs."""
        kdir = tmp_path / "knowledge"
        kdir.mkdir()
        (kdir / "real.json").write_text(json.dumps({"id": "r", "title": "R"}))
        idx = tmp_path / "index.toon"
        self._write_toon(idx, [
            ["R", "", "", "", "real.json"],
            ["Phantom", "", "", "", "ghost.json"],  # no file on disk
        ])
        issues = self._check(kdir, idx)
        assert any("ghost.json" in i and "missing JSON" in i for i in issues)

    def test_empty_knowledge_dir_without_index_passes(self, tmp_path):
        """No content JSONs and no index.toon — nothing to verify."""
        kdir = tmp_path / "knowledge"
        kdir.mkdir()
        assert self._check(kdir, tmp_path / "missing.toon") == []

    def test_cjk_filename_indexed(self, tmp_path):
        kdir = tmp_path / "knowledge"
        (kdir / "sub").mkdir(parents=True)
        (kdir / "sub" / "日本語.json").write_text(json.dumps({"id": "j", "title": "日本語"}))
        idx = tmp_path / "index.toon"
        self._write_toon(idx, [["日本語", "", "", "", "sub/日本語.json"]])
        assert self._check(kdir, idx) == []

    def test_fail_broken_json_surfaces_qo4(self, tmp_path):
        """Spec §3-3 point 4: a JSON that cannot be parsed is QO4 FAIL
        (no silent skip — zero-tolerance)."""
        kdir = tmp_path / "knowledge"
        kdir.mkdir()
        (kdir / "valid.json").write_text(json.dumps({"id": "v", "title": "V"}))
        (kdir / "broken.json").write_text("{ not json")
        idx = tmp_path / "index.toon"
        self._write_toon(idx, [["V", "", "", "", "valid.json"]])
        issues = self._check(kdir, idx)
        assert any("QO4" in i and "broken.json" in i and "parse failed" in i for i in issues)

    # --- Z-1 r7 QO4 Findings -------------------------------------------

    def test_fail_no_knowledge_json_listed_in_index_has_distinct_message(self, tmp_path):
        """Z-1 r7 QO4 F1: a no_knowledge_content JSON that is erroneously
        listed in index.toon must NOT be reported as 'missing JSON' — the
        file exists, the defect is that it was indexed. Distinct message."""
        kdir = tmp_path / "knowledge"
        kdir.mkdir()
        (kdir / "toc.json").write_text(json.dumps({"id": "t", "title": "T", "no_knowledge_content": True}))
        idx = tmp_path / "index.toon"
        self._write_toon(idx, [["T", "", "", "", "toc.json"]])
        issues = self._check(kdir, idx)
        assert any("QO4" in i and "no_knowledge" in i and "toc.json" in i for i in issues), issues
        assert not any("missing JSON" in i and "toc.json" in i for i in issues), issues

    def test_fail_broken_json_in_index_not_double_reported(self, tmp_path):
        """Z-1 r7 QO4 F2: broken JSON listed in index.toon must produce
        exactly one FAIL (the parse error), not a second misleading
        'missing JSON' message — the file exists on disk."""
        kdir = tmp_path / "knowledge"
        kdir.mkdir()
        (kdir / "broken.json").write_text("{ not json")
        idx = tmp_path / "index.toon"
        self._write_toon(idx, [["X", "", "", "", "broken.json"]])
        issues = self._check(kdir, idx)
        broken_issues = [i for i in issues if "broken.json" in i]
        assert any("parse failed" in i for i in broken_issues), broken_issues
        assert not any("missing JSON" in i for i in broken_issues), broken_issues

    def test_pass_toon_backslash_path_normalised(self, tmp_path):
        """Z-1 r7 QO4 F5: if a TOON writer emits a backslash path, verify
        normalises to forward slash on both sides so equality holds."""
        kdir = tmp_path / "knowledge"
        (kdir / "sub").mkdir(parents=True)
        (kdir / "sub" / "a.json").write_text(json.dumps({"id": "a", "title": "A"}))
        idx = tmp_path / "index.toon"
        # Write a row with a backslash separator in the path.
        idx.write_text(
            "files[1,]{title,type,category,processing_patterns,path}:\n"
            "  A, , , , sub\\a.json\n",
            encoding="utf-8",
        )
        assert self._check(kdir, idx) == []

    def test_fail_missing_index_lists_every_content_json_strict(self, tmp_path):
        """Z-1 r7 QO4 F7: pin the spec requirement that when index.toon
        is absent, EVERY content JSON appears as a FAIL — not just the
        header. A test that asserted only 'any QO4 in issues' would pass
        even if the per-file enumeration regressed to nothing."""
        kdir = tmp_path / "knowledge"
        kdir.mkdir()
        (kdir / "a.json").write_text(json.dumps({"id": "a", "title": "A"}))
        (kdir / "b.json").write_text(json.dumps({"id": "b", "title": "B"}))
        issues = self._check(kdir, tmp_path / "no-such.toon")
        per_file = [i for i in issues if "not registered" in i]
        assert len(per_file) >= 2
        assert any("a.json" in i for i in per_file)
        assert any("b.json" in i for i in per_file)


# ---------------------------------------------------------------------------
# QO3: docs MD 存在確認 (via check_docs_coverage)
# ---------------------------------------------------------------------------

class TestCheckDocsCoverage:
    """QO3: each JSON file has a corresponding docs MD (1:1 existence)."""

    def _check(self, knowledge_dir, docs_dir):
        from scripts.verify.verify import check_docs_coverage
        return check_docs_coverage(knowledge_dir, docs_dir)

    def _write_json(self, kdir: Path, rel: str, data: dict | None = None) -> Path:
        path = kdir / rel
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(data or {"title": "t"}), encoding="utf-8")
        return path

    def _write_md(self, ddir: Path, rel: str) -> Path:
        path = ddir / rel
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text("content")
        return path

    def test_pass_each_json_has_docs_md(self, tmp_path):
        kdir = tmp_path / "knowledge"; kdir.mkdir()
        ddir = tmp_path / "docs"; ddir.mkdir()
        # JSON layout mirrors docs MD layout (same relative path, .json -> .md)
        self._write_json(kdir, "about/nablarch/a.json")
        self._write_md(ddir, "about/nablarch/a.md")
        (ddir / "README.md").write_text("1ページ\n")
        assert self._check(kdir, ddir) == []

    def test_fail_json_without_matching_docs_md(self, tmp_path):
        kdir = tmp_path / "knowledge"; kdir.mkdir()
        ddir = tmp_path / "docs"; ddir.mkdir()
        self._write_json(kdir, "about/nablarch/a.json")
        # Note: no corresponding a.md in docs/about/nablarch/
        (ddir / "README.md").write_text("0ページ\n")
        issues = self._check(kdir, ddir)
        assert any("QO3" in i for i in issues)

    def test_pass_no_knowledge_json_still_requires_docs_md(self, tmp_path):
        """no_knowledge_content JSONs still get a minimal docs MD (see docs.py)."""
        kdir = tmp_path / "knowledge"; kdir.mkdir()
        ddir = tmp_path / "docs"; ddir.mkdir()
        self._write_json(kdir, "about/nablarch/a.json", {"no_knowledge_content": True, "title": "a"})
        self._write_md(ddir, "about/nablarch/a.md")
        (ddir / "README.md").write_text("1ページ\n")
        assert self._check(kdir, ddir) == []

    def test_fail_readme_missing(self, tmp_path):
        kdir = tmp_path / "knowledge"; kdir.mkdir()
        ddir = tmp_path / "docs"; ddir.mkdir()
        issues = self._check(kdir, ddir)
        assert any("README" in i for i in issues)

    # --- QO3 Z-1 gap fill -------------------------------------------------

    def test_fail_docs_md_at_wrong_nested_path(self, tmp_path):
        """JSON at knowledge/a/b/c.json requires docs MD at docs/a/b/c.md.
        A docs MD at a different path (e.g. top level) must still FAIL."""
        kdir = tmp_path / "knowledge"; kdir.mkdir()
        ddir = tmp_path / "docs"; ddir.mkdir()
        self._write_json(kdir, "a/b/c.json")
        self._write_md(ddir, "c.md")  # wrong location
        (ddir / "README.md").write_text("1ページ\n")
        issues = self._check(kdir, ddir)
        assert any("QO3" in i and "a/b/c.md" in i for i in issues)

    def test_pass_cjk_filename(self, tmp_path):
        kdir = tmp_path / "knowledge"; kdir.mkdir()
        ddir = tmp_path / "docs"; ddir.mkdir()
        self._write_json(kdir, "about/日本語.json")
        self._write_md(ddir, "about/日本語.md")
        (ddir / "README.md").write_text("1ページ\n")
        assert self._check(kdir, ddir) == []

    def test_pass_empty_knowledge_dir(self, tmp_path):
        kdir = tmp_path / "knowledge"; kdir.mkdir()
        ddir = tmp_path / "docs"; ddir.mkdir()
        (ddir / "README.md").write_text("0ページ\n")
        assert self._check(kdir, ddir) == []

    def test_fail_readme_page_count_mismatch(self, tmp_path):
        kdir = tmp_path / "knowledge"; kdir.mkdir()
        ddir = tmp_path / "docs"; ddir.mkdir()
        self._write_json(kdir, "a.json")
        self._write_md(ddir, "a.md")
        self._write_md(ddir, "b.md")
        (ddir / "README.md").write_text("99ページ\n")  # wrong
        issues = self._check(kdir, ddir)
        assert any("count mismatch" in i for i in issues)


# ---------------------------------------------------------------------------
# QC5: 形式純粋性
# ---------------------------------------------------------------------------

class TestVerifyFileQC5:
    """QC5: No format-specific syntax remnants in JSON output."""

    def _check(self, data, fmt):
        from scripts.verify.verify import verify_file
        import json, tempfile, os
        with tempfile.NamedTemporaryFile(mode='w', suffix=f'.{fmt}', delete=False, encoding='utf-8') as sf:
            sf.write("dummy source\n")
            src = sf.name
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as jf:
            json.dump(data, jf, ensure_ascii=False)
            jpath = jf.name
        try:
            from scripts.verify.verify import verify_file
            return verify_file(src, jpath, fmt)
        finally:
            os.unlink(src)
            os.unlink(jpath)

    # RST format
    def test_fail_rst_role_in_content(self):
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": ":ref:`something`"}
        ]}
        issues = self._check(data, "rst")
        assert any("QC5" in i and "RST role" in i for i in issues)

    def test_fail_rst_directive_in_content(self):
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": ".. note::"}
        ]}
        issues = self._check(data, "rst")
        assert any("QC5" in i and "directive" in i for i in issues)

    def test_fail_rst_label_in_content(self):
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": ".. _my-label:"}
        ]}
        issues = self._check(data, "rst")
        assert any("QC5" in i and "label" in i for i in issues)

    def test_fail_rst_heading_underline_in_title(self):
        data = {"id": "f", "title": "====", "content": "", "sections": []}
        issues = self._check(data, "rst")
        assert any("QC5" in i and "underline" in i for i in issues)

    def test_pass_rst_clean_content(self):
        data = {"id": "f", "title": "概要", "content": "普通の本文です。", "sections": [
            {"id": "s1", "title": "詳細", "content": "詳細説明。"}
        ]}
        assert [i for i in self._check(data, "rst") if "QC5" in i] == []

    # MD format
    def test_fail_md_raw_html_in_content(self):
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": "<details>内容</details>"}
        ]}
        issues = self._check(data, "md")
        assert any("QC5" in i and "HTML" in i for i in issues)

    def test_fail_md_backslash_escape_in_content(self):
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": r"これは\*エスケープ\*です"}
        ]}
        issues = self._check(data, "md")
        assert any("QC5" in i and "backslash" in i for i in issues)

    def test_pass_md_clean_content(self):
        data = {"id": "f", "title": "概要", "content": "普通の説明文。", "sections": [
            {"id": "s1", "title": "詳細", "content": "詳細です。`code` も含む。"}
        ]}
        assert [i for i in self._check(data, "md") if "QC5" in i] == []

    def test_pass_xlsx_no_qc5(self):
        """xlsx format: QC5 is not applicable — _check_format_purity returns []."""
        from scripts.verify.verify import _check_format_purity
        data = {"id": "f", "title": "T", "content": ":ref:`role`", "sections": []}
        assert _check_format_purity(data, "xlsx") == []

    def test_pass_no_knowledge_content_skipped(self):
        data = {"id": "f", "title": "T", "no_knowledge_content": True, "sections": []}
        assert self._check(data, "rst") == []

    # --- QC5 Z-1 gap fill -------------------------------------------------

    def test_pass_rst_heading_underline_in_code_block_content(self):
        """Heading underline `====` can legitimately appear inside code
        blocks in content; QC5 restricts the underline check to titles only."""
        content_with_code = "```\n===== \n```"
        data = {"id": "f", "title": "概要", "content": "",
                "sections": [{"id": "s1", "title": "詳細", "content": content_with_code}]}
        assert [i for i in self._check(data, "rst") if "QC5" in i and "underline" in i] == []

    def test_pass_rst_field_list_syntax_is_not_qc5_role(self):
        """Spec §3-1 QC5 defines the RST role pattern as `:role:\\`text\\``.
        An RST field list marker `:name:` (no backtick-delimited argument)
        is legitimate inline syntax that survives verbatim into the normalised
        MD; it is not an unprocessed role, so QC5 must not FAIL."""
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": "フィールド :name: 値 のような行は許容"}
        ]}
        assert [i for i in self._check(data, "rst") if "QC5" in i and "role" in i] == []

    def test_pass_rst_japanese_punctuation_not_confused_with_role(self):
        """Japanese colon-like punctuation must not be mistaken for a role."""
        data = {"id": "f", "title": "概要", "content": "", "sections": [
            {"id": "s1", "title": "詳細", "content": "注意：ここは普通の文章です。"}
        ]}
        assert [i for i in self._check(data, "rst") if "QC5" in i] == []

    # --- Z-1 r7 Findings: QC5 regex strictness ---------------------------

    def test_pass_rst_role_name_without_closing_backtick_not_flagged(self):
        """Spec §3-1 QC5 writes the RST role pattern as `:role:\\`text\\``
        — BOTH backticks required. rbkc.md: 'Yes. Spec §3-1 QC5 writes
        `:role:\\`text\\`` (both backticks).' An opening-backtick-only
        sequence is malformed and not the named pattern; QC5 must not FAIL
        on it. (Z-1 r7 QC5 F1)"""
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": "文中に :foo:` のような破片"}
        ]}
        assert [i for i in self._check(data, "rst") if "QC5" in i and "role" in i] == []

    def test_fail_rst_role_with_both_backticks_flagged(self):
        """Complement to the pass test: a real role `:ref:\\`x\\`` must
        still FAIL QC5 (the converter should have resolved it)."""
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": "未処理 :ref:`label` 残"}
        ]}
        assert any("QC5" in i and "role" in i for i in self._check(data, "rst"))

    def test_pass_rst_label_tokens_in_prose_not_flagged(self):
        """Spec §3-1 QC5 defines the label pattern as `.. _label:` — by
        the RST explicit-markup spec this construct must begin at line
        start. A mid-sentence occurrence is not a label definition and
        must not FAIL QC5. (Z-1 r7 QC5 F2)"""
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": "see .. _foo: below in text"}
        ]}
        assert [i for i in self._check(data, "rst") if "QC5" in i and "label" in i] == []

    def test_fail_rst_label_on_its_own_line_flagged(self):
        """A label definition on its own line must still FAIL QC5."""
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": "前文\n.. _my-label:\n後文"}
        ]}
        assert any("QC5" in i and "label" in i for i in self._check(data, "rst"))

    def test_fail_md_summary_tag(self):
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": "<summary>タイトル</summary>"}
        ]}
        assert any("QC5" in i and "HTML" in i for i in self._check(data, "md"))

    def test_fail_md_br_tag(self):
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": "改行<br>入り"}
        ]}
        assert any("QC5" in i and "HTML" in i for i in self._check(data, "md"))

    def test_fail_md_a_tag(self):
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": "<a href='x'>link</a>"}
        ]}
        assert any("QC5" in i and "HTML" in i for i in self._check(data, "md"))

    def test_fail_md_escaped_underscore(self):
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": r"word\_word"}
        ]}
        assert any("QC5" in i and "backslash" in i for i in self._check(data, "md"))

    def test_fail_md_escaped_bracket(self):
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": r"\[not a link\]"}
        ]}
        assert any("QC5" in i and "backslash" in i for i in self._check(data, "md"))

    def test_fail_md_raw_html_inside_inline_code_still_detected(self):
        """QC5 is an independent check on JSON text. Raw HTML anywhere in
        content FAILs — a literal `<br>` inside backticks is still HTML
        from QC5's perspective because the converter should have escaped
        or transformed it before emission. This pins the spec (§3-1 QC5
        RST/MD 'raw HTML' pattern)."""
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": "使い方: `<br>` を挿入"}
        ]}
        assert any("QC5" in i and "HTML" in i for i in self._check(data, "md"))

    def test_fail_md_self_closing_br(self):
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": "line<br/>break"}
        ]}
        assert any("QC5" in i and "HTML" in i for i in self._check(data, "md"))

    def test_fail_md_self_closing_hr(self):
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": "sep<hr/>done"}
        ]}
        assert any("QC5" in i and "HTML" in i for i in self._check(data, "md"))

    def test_fail_md_self_closing_img(self):
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "概要", "content": "see <img/>"}
        ]}
        assert any("QC5" in i and "HTML" in i for i in self._check(data, "md"))


# ---------------------------------------------------------------------------
# QL2: 外部URL一致
# ---------------------------------------------------------------------------

class TestVerifyFileQL2:
    """QL2: External URLs in source must appear verbatim in JSON."""

    def _check_ql2(self, source_text, data, fmt):
        from scripts.verify.verify import check_external_urls
        return check_external_urls(source_text, data, fmt)

    def test_pass_url_in_json(self):
        src = "詳細は https://example.com を参照。\n"
        data = {"id": "f", "title": "T", "content": "https://example.com を参照。", "sections": []}
        assert self._check_ql2(src, data, "rst") == []

    def test_fail_url_missing_from_json(self):
        src = "詳細は https://example.com を参照。\n"
        data = {"id": "f", "title": "T", "content": "説明。", "sections": []}
        issues = self._check_ql2(src, data, "rst")
        assert any("QL2" in i and "example.com" in i for i in issues)

    def test_pass_duplicate_url_reported_once(self):
        src = "https://example.com と https://example.com が重複。\n"
        data = {"id": "f", "title": "T", "content": "https://example.com を参照。", "sections": []}
        issues = self._check_ql2(src, data, "rst")
        # URL present in JSON → no FAIL even if duplicated in source
        assert all("QL2" not in i for i in issues)

    def test_pass_rst_target_def_url_excluded(self):
        src = ".. _Name: https://only-in-target.com\n通常テキスト\n"
        data = {"id": "f", "title": "T", "content": "通常テキスト", "sections": []}
        # RST target definition URLs are dropped by converter, so no QL2 FAIL
        assert self._check_ql2(src, data, "rst") == []

    def test_pass_no_source_urls(self):
        src = "URLなしのテキスト\n"
        data = {"id": "f", "title": "T", "content": "説明。", "sections": []}
        assert self._check_ql2(src, data, "rst") == []

    def test_pass_xlsx_skipped(self):
        src = "https://should-be-ignored.com\n"
        data = {"id": "f", "title": "T", "content": "説明。", "sections": []}
        assert self._check_ql2(src, data, "xlsx") == []

    def test_pass_no_knowledge_content_skipped(self):
        src = "https://example.com\n"
        data = {"id": "f", "title": "T", "no_knowledge_content": True, "sections": []}
        assert self._check_ql2(src, data, "rst") == []

    def test_fail_url_in_section_content_missing(self):
        src = "詳細は https://example.com を参照。\n"
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "参照", "content": "リンクなし"}
        ]}
        issues = self._check_ql2(src, data, "rst")
        assert any("QL2" in i for i in issues)

    def test_pass_url_in_section_content(self):
        src = "詳細は https://example.com を参照。\n"
        data = {"id": "f", "title": "T", "content": "", "sections": [
            {"id": "s1", "title": "参照", "content": "https://example.com を参照。"}
        ]}
        assert self._check_ql2(src, data, "rst") == []

    def test_pass_rst_inline_code_url_trailing_backtick_trimmed(self):
        """RST ``http://...`` must not leak trailing backticks into the URL."""
        src = "起動したら ``http://localhost:9080/`` にアクセス。\n"
        data = {"id": "f", "title": "T", "content": "起動したら `http://localhost:9080/` にアクセス。", "sections": []}
        assert self._check_ql2(src, data, "rst") == []

    def test_pass_rst_substitution_only_url_skipped(self):
        """URLs appearing only inside a substitution directive body are skipped."""
        src = (
            "通常テキスト。\n\n"
            ".. |jsr317| raw:: html\n\n"
            "   <a href=\"https://only-in-subst.example\" target=\"_blank\">Text</a>\n"
        )
        data = {"id": "f", "title": "T", "content": "通常テキスト。", "sections": []}
        assert self._check_ql2(src, data, "rst") == []

    # --- QL2 Z-1 gap fill: MD path + URL-edge-cases ----------------------

    def test_fail_md_inline_link_url_missing(self):
        src = "# T\n\nsee [docs](https://example.com/a/b) please.\n"
        data = {"id": "f", "title": "T", "content": "see docs please.", "sections": []}
        issues = self._check_ql2(src, data, "md")
        assert any("QL2" in i and "example.com/a/b" in i for i in issues)

    def test_pass_md_inline_link_url_present(self):
        src = "# T\n\nsee [docs](https://example.com/a/b)\n"
        data = {"id": "f", "title": "T", "content": "[docs](https://example.com/a/b)", "sections": []}
        assert self._check_ql2(src, data, "md") == []

    def test_fail_md_autolink_url_missing(self):
        """CommonMark autolink `<https://...>` must be collected via AST."""
        src = "# T\n\nvisit <https://auto.example.com/path>\n"
        data = {"id": "f", "title": "T", "content": "visit", "sections": []}
        issues = self._check_ql2(src, data, "md")
        assert any("QL2" in i and "auto.example.com/path" in i for i in issues)

    def test_pass_md_url_with_query_and_fragment(self):
        """Query string and fragment must not be truncated by the extractor."""
        src = "# T\n\n[link](https://example.com/path?x=1&y=2#frag)\n"
        data = {"id": "f", "title": "T",
                "content": "[link](https://example.com/path?x=1&y=2#frag)", "sections": []}
        assert self._check_ql2(src, data, "md") == []

    def test_pass_md_url_with_parentheses_in_path(self):
        """URL containing parentheses in the path — spec requires the
        source-literal URL to appear verbatim in JSON. The expected URL
        is pinned to the source literal (NOT derived from AST output),
        so the test catches extractor regressions that truncate the URL
        at the `)`. (Z-1 r7 QL2 F1 — previously circular.)"""
        src = "# T\n\n[api](https://example.com/foo(bar))\n"
        data = {"id": "f", "title": "T",
                "content": "api https://example.com/foo(bar)", "sections": []}
        assert self._check_ql2(src, data, "md") == []

    def test_fail_md_url_with_parentheses_truncated_in_json(self):
        """Complement to above: if JSON has the URL truncated at `)`
        (what a buggy extractor might produce), QL2 must FAIL."""
        src = "# T\n\n[api](https://example.com/foo(bar))\n"
        data = {"id": "f", "title": "T",
                "content": "api https://example.com/foo(bar", "sections": []}
        issues = self._check_ql2(src, data, "md")
        assert any("QL2" in i for i in issues)

    def test_pass_md_url_followed_by_japanese_punct_not_absorbed(self):
        """Trailing 」 / 。 must not be absorbed into the URL (AST-only
        principle — the old regex approach was prone to this)."""
        src = "# T\n\n参照「[docs](https://example.com/x)」です。\n"
        data = {"id": "f", "title": "T",
                "content": "[docs](https://example.com/x)", "sections": []}
        assert self._check_ql2(src, data, "md") == []

    def test_fail_md_http_vs_https_distinguished(self):
        """http:// and https:// are different URLs; substituting one for
        the other must FAIL QL2 (exact match required)."""
        src = "# T\n\n[link](http://example.com/a)\n"
        data = {"id": "f", "title": "T",
                "content": "[link](https://example.com/a)", "sections": []}
        issues = self._check_ql2(src, data, "md")
        assert any("QL2" in i and "http://example.com/a" in i for i in issues)

    def test_pass_md_autolink_url_present(self):
        """Z-1 r7 QL2 F3: PASS counterpart for autolink. When the
        autolinked URL is present in JSON content, QL2 must not FAIL.
        (Proves the check is bidirectional — the FAIL case is vacuous
        without a matching PASS.)"""
        src = "# T\n\nvisit <https://auto.example.com/path>\n"
        data = {"id": "f", "title": "T",
                "content": "visit https://auto.example.com/path", "sections": []}
        assert self._check_ql2(src, data, "md") == []

    def test_pass_rst_url_with_parens_javadoc_anchor(self):
        """Z-1 r7 QL2 F4: RST coverage for URL containing balanced parens
        (Javadoc anchors are the most common case in the v6 corpus)."""
        src = (
            "概要\n====\n\n"
            "`Javadoc <https://example.com/Class.html#m(java.lang.String)>`_ を参照\n"
        )
        data = {"id": "f", "title": "T",
                "content": "Javadoc https://example.com/Class.html#m(java.lang.String)",
                "sections": []}
        assert self._check_ql2(src, data, "rst") == []

    def test_fail_md_trailing_slash_distinguished(self):
        """Z-1 r7 QL2 F5: source URL WITH trailing slash, JSON without
        → source URL is not a substring of JSON → QL2 FAIL. Guards
        against extractor trailing-slash-stripping regressions."""
        src = "# T\n\n[link](https://example.com/a/)\n"
        data = {"id": "f", "title": "T",
                "content": "link https://example.com/a here", "sections": []}
        issues = self._check_ql2(src, data, "md")
        assert any("QL2" in i and "https://example.com/a/" in i for i in issues)


# ---------------------------------------------------------------------------
# QC1-QC4: content completeness (RST/MD sequential-delete)
# ---------------------------------------------------------------------------

class TestCheckContentCompleteness:
    """QC1-QC4: sequential-delete algorithm via check_content_completeness."""

    def _check(self, source_text, data, fmt="rst", label_map=None):
        from scripts.verify.verify import check_content_completeness
        return check_content_completeness(source_text, data, fmt, label_map)

    def _data(self, title="", content="", sections=None):
        return {
            "id": "f", "title": title, "content": content,
            "sections": sections or []
        }

    # --- QC2: fabricated content (not in source) ---

    def test_fail_qc2_fabricated_title(self):
        src = "概要\n====\n\n本文。\n"
        data = self._data(sections=[
            {"id": "s1", "title": "存在しないセクション", "content": "本文。"}
        ])
        issues = self._check(src, data)
        assert any("QC2" in i and "fabricated" in i for i in issues)

    def test_fail_qc2_fabricated_content(self):
        src = "概要\n====\n\n本文。\n"
        data = self._data(sections=[
            {"id": "s1", "title": "概要", "content": "捏造されたテキスト。"}
        ])
        issues = self._check(src, data)
        assert any("QC2" in i and "fabricated" in i for i in issues)

    # --- QC3: duplicate content ---

    def test_fail_qc3_duplicate_title(self):
        src = "概要\n====\n\n本文。\n\n詳細\n====\n\n詳細内容。\n"
        data = self._data(sections=[
            {"id": "s1", "title": "概要", "content": "本文。"},
            {"id": "s2", "title": "概要", "content": "詳細内容。"},  # duplicate title
        ])
        issues = self._check(src, data)
        assert any("[QC3]" in i for i in issues)
        assert not any("[QC2]" in i or "[QC4]" in i for i in issues)

    # --- QC4: misplaced content ---

    def test_fail_qc4_misplaced_title(self):
        src = "詳細\n====\n\n詳細内容。\n\n概要\n====\n\n概要内容。\n"
        data = self._data(sections=[
            {"id": "s1", "title": "概要", "content": "概要内容。"},   # appears later in source
            {"id": "s2", "title": "詳細", "content": "詳細内容。"},   # appears earlier in source
        ])
        issues = self._check(src, data)
        assert any("[QC4]" in i and "s2" in i for i in issues), issues

    # --- QC1: source content not captured ---

    def test_fail_qc1_residual_content(self):
        src = "概要\n====\n\n本文。\n\n追加情報はここにあります。\n"
        data = self._data(sections=[
            {"id": "s1", "title": "概要", "content": "本文。"}
            # "追加情報はここにあります。" not captured
        ])
        issues = self._check(src, data)
        assert any("QC1" in i for i in issues)

    def test_fail_qc1_rst_reports_every_residue_fragment(self):
        """Spec §3-1 判定分岐 row 4 names residue as non-whitespace text
        remaining; per `.claude/rules/rbkc.md` ('RST one-snippet vs MD
        all-fragments — All fragments'), every fragment must be reported.
        A prior implementation truncated RST residue to a single 80-char
        snippet, hiding additional gaps. (Z-1 r7 QC1 F2)"""
        src = "概要\n====\n\n本文。\n\nalpha\n\nbravo\n\ncharlie\n"
        data = self._data(sections=[
            {"id": "s1", "title": "概要", "content": "本文。"}
            # alpha / bravo / charlie all uncaptured — three disjoint
            # residue fragments.
        ])
        issues = self._check(src, data)
        qc1 = [i for i in issues if "[QC1]" in i]
        # Expect one issue per residue fragment, not a single snippet.
        assert any("alpha" in i for i in qc1), qc1
        assert any("bravo" in i for i in qc1), qc1
        assert any("charlie" in i for i in qc1), qc1

    def test_pass_rst_syntax_in_residual_allowed(self):
        # Converter renders `.. note::` as `> **Note:** ...` so JSON content
        # includes the MD admonition header. New Visitor renders note body
        # as nested paragraph inside the blockquote, so the JSON content
        # should match the visitor's output.
        src = "概要\n====\n\n本文。\n\n.. note::\n\n   注記内容。\n"
        data = self._data(sections=[
            {"id": "s1", "title": "概要", "content": "本文。\n\n> **Note:**\n> 注記内容。"}
        ])
        issues = self._check(src, data)
        assert all("QC1" not in i for i in issues)

    def test_pass_md_heading_captured_as_title(self):
        # MD h1 is captured as the JSON top-level title; h2 opens a section.
        # The normalised source contains both, and all are consumed by JSON.
        src = "# タイトル\n\n## セクション\n\n本文。\n"
        data = self._data(title="タイトル", sections=[
            {"id": "s1", "title": "セクション", "content": "本文。"}
        ])
        issues = self._check(src, data, fmt="md")
        assert all("QC1" not in i for i in issues)

    # --- PASS cases ---

    def test_pass_all_content_captured(self):
        src = "概要\n====\n\n本文です。\n"
        data = self._data(sections=[
            {"id": "s1", "title": "概要", "content": "本文です。"}
        ])
        assert self._check(src, data) == []

    def test_pass_no_knowledge_content_skipped(self):
        data = {"id": "f", "title": "T", "no_knowledge_content": True, "sections": []}
        assert self._check("any source", data) == []

    def test_pass_empty_data_no_issues(self):
        data = self._data()
        assert self._check("any source", data) == []

    def test_pass_md_verbatim_match(self):
        """MD source and MD JSON content are same format — verbatim match."""
        src = "## セクション\n\n**重要な**情報があります。\n"
        data = self._data(sections=[
            {"id": "s1", "title": "セクション", "content": "**重要な**情報があります。"}
        ])
        issues = self._check(src, data, fmt="md")
        assert all("QC2" not in i for i in issues)
        assert all("QC1" not in i for i in issues)

    def test_pass_top_level_content_captured(self):
        src = "タイトル\n========\n\nトップレベル本文。\n\nセクション\n==========\n\nセクション本文。\n"
        data = self._data(
            title="タイトル",
            content="トップレベル本文。",
            sections=[{"id": "s1", "title": "セクション", "content": "セクション本文。"}]
        )
        assert self._check(src, data) == []

    # --- RST normalization: false positive prevention ---

    def test_pass_rst_double_backtick_inline_code(self):
        """RST ``code`` is converted to MD `code` — must not trigger QC2."""
        src = "名前空間が ``javax.*`` から ``jakarta.*`` になる。\n"
        data = self._data(content="名前空間が `javax.*` から `jakarta.*` になる。")
        assert self._check(src, data) == []

    def test_pass_rst_ref_label_resolved_text(self):
        """RST :ref:`label` is resolved to the target section title via label_map
        (§3-1b zero-exception: unresolved would FAIL; here we supply it)."""
        src = "詳細は :ref:`doma_dependency` を参照。\n\ndoma_dependency\n===============\n\nDoma 設定。\n"
        data = self._data(
            content="詳細は doma_dependency を参照。",
            sections=[{"id": "s1", "title": "doma_dependency", "content": "Doma 設定。"}]
        )
        assert self._check(src, data, label_map={"doma_dependency": "doma_dependency"}) == []

    def test_pass_rst_external_link_text(self):
        """RST `text <url>`_ is converted to MD [text](url) — must not trigger QC2."""
        src = "`公式サイト <https://example.com>`_ を参照。\n"
        data = self._data(content="[公式サイト](https://example.com) を参照。")
        assert self._check(src, data) == []

    def test_pass_rst_ref_display_form_resolved(self):
        """RST :ref:`display <label>` resolved to display text — must not trigger QC2."""
        src = "詳細は :ref:`Doma設定 <doma_config>` を参照。\n"
        data = self._data(content="詳細は Doma設定 を参照。")
        assert self._check(src, data) == []

    def test_pass_rst_backtick_underscore_literal_in_code(self):
        """RST ``_`` literal underscore inside inline code must not be stripped as
        named-reference marker. Converter emits `_` in MD; both sides must align."""
        src = "区切り文字に ``_`` を使用する。\n"
        data = self._data(content="区切り文字に `_` を使用する。")
        assert self._check(src, data) == []

    def test_pass_rst_comment_line_is_syntax(self):
        """``.. text-without-colons`` lines are RST comments per spec and must be
        treated as allowed syntax residue for QC1."""
        src = "概要\n====\n\n本文。\n\n.. textlint-disable ja/foo\n"
        data = self._data(sections=[
            {"id": "s1", "title": "概要", "content": "本文。"}
        ])
        assert self._check(src, data) == []

    def test_pass_rst_comment_block_with_indented_body(self):
        """``..`` comment with indented body: entire block is RST syntax per spec."""
        src = (
            "概要\n====\n\n本文。\n\n"
            ".. 実装済み\n"
            ".. * 項目1\n"
            "..   * サブ項目\n"
            ".. * 項目2\n"
        )
        data = self._data(sections=[
            {"id": "s1", "title": "概要", "content": "本文。"}
        ])
        assert self._check(src, data) == []

    def test_pass_rst_field_list_with_inline_value(self):
        """RST field list ``:name: value`` — per §3-1a, standalone field_list
        drops the name and preserves the value (recursively visited)."""
        src = "概要\n====\n\n:エスケープ対象文字: ``%`` 、 ``_``\n"
        data = self._data(sections=[
            {"id": "s1", "title": "概要", "content": "`%` 、 `_`"}
        ])
        assert self._check(src, data) == []

    def test_pass_rst_field_list_with_separate_value(self):
        """RST field list ``:name:\\n  value`` — value-only retained."""
        src = (
            "概要\n====\n\n"
            ":Status-Code:\n"
            "  応答電文のステータスコード。\n"
        )
        data = self._data(sections=[
            {"id": "s1", "title": "概要", "content": "応答電文のステータスコード。"}
        ])
        assert self._check(src, data) == []

    # --- QC1 Z-1 gap fill: parse / Visitor error paths -------------------

    def test_fail_qc1_md_unknown_token_surfaces(self):
        """Spec §3-1b zero-exception: an unknown markdown-it token type
        must surface as a QC1 FAIL. We inject a token stream with an
        unregistered block token and route it through verify's MD path."""
        import scripts.common.md_ast as md_ast_mod
        from markdown_it.token import Token

        def _inject_unknown(source):
            # Minimal normal block + one unknown block token
            heading_open = Token("heading_open", "h1", 1); heading_open.tag = "h1"; heading_open.markup = "#"
            inline = Token("inline", "", 0); inline.content = "T"
            inline.children = [Token("text", "", 0)]; inline.children[0].content = "T"
            heading_close = Token("heading_close", "h1", -1); heading_close.tag = "h1"; heading_close.markup = "#"
            unknown = Token("custom_block_xyz", "", 0)
            return [heading_open, inline, heading_close, unknown]

        orig = md_ast_mod.parse
        md_ast_mod.parse = _inject_unknown
        try:
            issues = self._check("# T\n", self._data(title="T"), fmt="md")
        finally:
            md_ast_mod.parse = orig
        assert any("[QC1]" in i and "markdown parse/visitor error" in i for i in issues)

    def test_fail_qc1_rst_unresolved_substitution_surfaces(self):
        """Spec §3-1b: unresolved substitution (`|missing|` with no
        `.. |missing| replace::` definition) → QC1 FAIL. The Visitor
        raises UnresolvedReferenceError which verify reports as QC1."""
        src = "概要\n====\n\n利用バージョン: |undefined_version|\n"
        data = self._data(sections=[{"id": "s1", "title": "概要", "content": "x"}])
        issues = self._check(src, data)
        assert any("[QC1]" in i and "RST parse/visitor error" in i for i in issues)

    def test_fail_qc1_rst_parse_error_level_3(self):
        """Spec §3-1b 原則 4: docutils parse error (level >= 3) → QC1 FAIL.
        An unknown directive triggers `(ERROR/3)` in docutils' warning
        stream; the normaliser scans for that and raises UnknownSyntaxError."""
        src = "概要\n====\n\n.. unknown-directive-xyz::\n\n   body text\n"
        data = self._data(sections=[{"id": "s1", "title": "概要", "content": "x"}])
        issues = self._check(src, data)
        assert any("[QC1]" in i and "RST parse/visitor error" in i for i in issues)

    def test_fail_qc1_rst_unknown_role_surfaces(self):
        """Unknown RST role (not in Sphinx shim list) → UnknownSyntaxError → QC1 FAIL."""
        # `:unknownshim:` is not in _SPHINX_INLINE_ROLES and not a docutils native role
        src = "概要\n====\n\nテキスト :unknownshim:`x` を含む。\n"
        data = self._data(sections=[{"id": "s1", "title": "概要", "content": "テキスト を含む。"}])
        issues = self._check(src, data)
        assert any("[QC1]" in i and "RST parse/visitor error" in i for i in issues)

    # --- QC2 Z-1 gap fill: multiple fabrications + top-level + near-miss ---

    def test_fail_qc2_multiple_fabricated_contents(self):
        src = "概要\n====\n\n本文。\n"
        data = self._data(sections=[
            {"id": "s1", "title": "概要", "content": "捏造 A。"},
            {"id": "s2", "title": "詳細", "content": "捏造 B。"},
        ])
        issues = self._check(src, data)
        qc2 = [i for i in issues if "QC2" in i]
        assert len(qc2) >= 2

    def test_fail_qc2_top_level_fabricated_content(self):
        src = "概要\n====\n\n本文。\n"
        data = self._data(title="概要", content="存在しないトップレベル本文。",
                          sections=[{"id": "s1", "title": "概要", "content": "本文。"}])
        issues = self._check(src, data)
        assert any("QC2" in i and "fabricated content" in i for i in issues)

    def test_fail_qc2_near_miss_one_char_differs(self):
        src = "概要\n====\n\nABCDEFG\n"
        data = self._data(sections=[{"id": "s1", "title": "概要", "content": "ABCXEFG"}])
        issues = self._check(src, data)
        assert any("QC2" in i for i in issues)

    # --- QC3 Z-1 gap fill: all 4 untested paths + CJK short collision ----

    def test_fail_qc3_duplicate_content_rst(self):
        src = "概要\n====\n\n共通テキスト。\n\n詳細\n====\n\n別テキスト。\n"
        data = self._data(sections=[
            {"id": "s1", "title": "概要", "content": "共通テキスト。"},
            {"id": "s2", "title": "詳細", "content": "共通テキスト。"},  # content in JSON not in source twice
        ])
        issues = self._check(src, data)
        # Spec §3-1 distinguishes QC3 from QC2/QC4 by label — assert the exact label.
        assert any("[QC3]" in i for i in issues)
        assert not any("[QC2]" in i or "[QC4]" in i for i in issues)

    def test_fail_qc3_duplicate_title_md(self):
        src = "# T\n\n## 概要\n\n本文 A。\n\n## 詳細\n\n本文 B。\n"
        data = self._data(title="T", sections=[
            {"id": "s1", "title": "概要", "content": "本文 A。"},
            {"id": "s2", "title": "概要", "content": "本文 B。"},  # duplicate title in JSON
        ])
        issues = self._check(src, data, fmt="md")
        assert any("[QC3]" in i for i in issues)
        assert not any("[QC2]" in i or "[QC4]" in i for i in issues)

    def test_pass_qc3_short_cjk_repeated_in_source_and_json(self):
        """Two sections legitimately using the same short CJK title; both
        appear in source and JSON — no false QC3."""
        src = "## 概要\n\nA 本文。\n\n## 概要\n\nB 本文。\n"
        data = self._data(title="", sections=[
            {"id": "s1", "title": "概要", "content": "A 本文。"},
            {"id": "s2", "title": "概要", "content": "B 本文。"},
        ])
        issues = self._check(src, data, fmt="md")
        # Source has "概要" twice → each JSON title consumes a distinct
        # occurrence. No QC3/QC4 for the title pair.
        assert not any("[QC3]" in i and "概要" in i for i in issues)

    def test_fail_qc3_top_level_and_section_content_duplicated(self):
        """Top-level content and a section content both declare the same
        text, but source has it only once → second consumption collides."""
        src = "# T\n\n共通テキスト。\n\n## 詳細\n"
        data = self._data(
            title="T",
            content="共通テキスト。",
            sections=[{"id": "s1", "title": "詳細", "content": "共通テキスト。"}],
        )
        issues = self._check(src, data, fmt="md")
        assert any("[QC3]" in i for i in issues)
        assert not any("[QC2]" in i or "[QC4]" in i for i in issues)

    def test_fail_qc3_duplicate_content_md(self):
        src = "# T\n\n## 概要\n\n本文。\n\n## 詳細\n\n別。\n"
        data = self._data(title="T", sections=[
            {"id": "s1", "title": "概要", "content": "本文。"},
            {"id": "s2", "title": "詳細", "content": "本文。"},  # JSON duplicates source "本文。" substring
        ])
        issues = self._check(src, data, fmt="md")
        assert any("[QC3]" in i for i in issues)
        assert not any("[QC2]" in i or "[QC4]" in i for i in issues)

    # --- QC4 Z-1 gap fill: MD misplacement + content-only swap ------------

    def test_fail_qc4_misplaced_title_md(self):
        src = "# T\n\n## 詳細\n\n詳細内容。\n\n## 概要\n\n概要内容。\n"
        data = self._data(title="T", sections=[
            {"id": "s1", "title": "概要", "content": "概要内容。"},
            {"id": "s2", "title": "詳細", "content": "詳細内容。"},
        ])
        issues = self._check(src, data, fmt="md")
        # Spec §3-1 L84 names the affected section id as part of the defect.
        assert any("[QC4]" in i and "s2" in i for i in issues), issues

    def test_fail_qc4_three_section_middle_swap(self):
        """Three sections A/B/C in source; JSON swaps B and C → QC4 must
        fire for one of them (position regression on 2nd or 3rd)."""
        src = "A\n=\n\na\n\nB\n=\n\nb\n\nC\n=\n\nc\n"
        data = self._data(sections=[
            {"id": "s1", "title": "A", "content": "a"},
            {"id": "s2", "title": "C", "content": "c"},
            {"id": "s3", "title": "B", "content": "b"},
        ])
        issues = self._check(src, data)
        # Spec §3-1 L185 names QC4 specifically; assert the label and the
        # flagged section id.
        assert any("[QC4]" in i and "s3" in i for i in issues), issues

    def test_fail_qc4_three_section_content_only_rotation_rst(self):
        """Spec §3-1 L84: 'セクション A のコンテンツが JSON の異なる
        セクションに配置されている'. Titles in spec order; contents
        rotated A→a, B→c, C→b. The middle section's content (c) is out
        of position. (Z-1 r7 QC4 F3)"""
        src = "A\n=\n\na\n\nB\n=\n\nb\n\nC\n=\n\nc\n"
        data = self._data(sections=[
            {"id": "s1", "title": "A", "content": "a"},
            {"id": "s2", "title": "B", "content": "c"},  # swapped
            {"id": "s3", "title": "C", "content": "b"},  # swapped
        ])
        issues = self._check(src, data)
        assert any("[QC4]" in i for i in issues)

    def test_fail_qc4_three_section_content_only_rotation_md(self):
        """MD mirror of the three-section content rotation above."""
        src = "# T\n\n## A\n\na\n\n## B\n\nb\n\n## C\n\nc\n"
        data = self._data(title="T", sections=[
            {"id": "s1", "title": "A", "content": "a"},
            {"id": "s2", "title": "B", "content": "c"},
            {"id": "s3", "title": "C", "content": "b"},
        ])
        issues = self._check(src, data, fmt="md")
        assert any("[QC4]" in i for i in issues)

    def test_fail_qc4_md_content_swap(self):
        """MD: two sections with swapped content (titles correct)."""
        src = "# T\n\n## A\n\nA の内容。\n\n## B\n\nB の内容。\n"
        data = self._data(title="T", sections=[
            {"id": "s1", "title": "A", "content": "B の内容。"},
            {"id": "s2", "title": "B", "content": "A の内容。"},
        ])
        issues = self._check(src, data, fmt="md")
        assert any("[QC4]" in i and "s2" in i for i in issues), issues

    def test_pass_qc3_single_consumption_of_duplicated_source_text(self):
        """Positive guard — when source has duplicated text but JSON only
        consumes one of them, no QC3/QC4 must fire. (Was part of the old
        boundary test; retained as a dedicated pass-guard.)"""
        src = "A\n=\n\nnote\n\nB\n=\n\nnote\n"
        data = self._data(sections=[
            {"id": "s1", "title": "A", "content": "note"},
        ])
        issues = self._check(src, data)
        assert not any("[QC3]" in i for i in issues)
        assert not any("[QC4]" in i for i in issues)

    def test_fail_qc4_boundary_text_occurs_in_both_positions_misplaced(self):
        """Spec §3-1 boundary between QC3 and QC4:

        Source has 'note' at positions A and B; JSON places what should
        be s1's body into s2 and vice versa. Both occurrences exist in
        source, both are UNCONSUMED at the time of JSON s1's lookup
        (JSON order: s1.title=A, s1.content=note_from_B_position,
        s2.title=B, s2.content=note_from_A_position). Position regression
        → QC4 per spec L185 "削除位置が JSON 順より前に逆行 | QC4".

        This test pins the label explicitly — a prior version of this
        test asserted only 'not QC3' which was vacuous when only one
        consumption occurred.
        """
        src = "A\n=\n\nnote1\n\nB\n=\n\nnote2\n"
        data = self._data(sections=[
            {"id": "s1", "title": "A", "content": "note2"},  # B's content
            {"id": "s2", "title": "B", "content": "note1"},  # A's content
        ])
        issues = self._check(src, data)
        assert any("[QC4]" in i for i in issues)

    def test_fail_qc4_not_qc3_when_middle_occurrence_is_unconsumed(self):
        """Spec §3-1 L184 '先行削除済み' means ALL earlier occurrences are
        consumed. If the earliest occurrence is consumed but a middle
        occurrence (still before current_pos) is NOT consumed, the verdict
        must be QC4 (unconsumed earlier offset exists → position regression),
        not QC3 (先行削除済み requires every earlier position consumed).

        Fixture: source has 'note' three times separated by anchors
        'alpha' and 'beta'. JSON order:
          1. s1.title 'note' consumes the EARLIEST occurrence
          2. s1.content 'beta' advances current_pos past the MIDDLE 'note'
             (without consuming it)
          3. s2.title 'note' consumes the LAST occurrence
          4. s3.title 'note' — find from current_pos fails; earliest is
             consumed, but middle is UNCONSUMED and < current_pos.

        Spec: s3's verdict is QC4 (position regression against the middle
        unconsumed occurrence). Buggy impl that picks only the earliest
        occurrence labels this QC3. Middle occurrence also surfaces as
        QC1 residue, which is orthogonal.
        """
        src = "# H\n\nnote alpha note beta note\n"
        data = self._data(title="H", sections=[
            {"id": "s1", "title": "note", "content": "beta"},
            {"id": "s2", "title": "note", "content": ""},
            {"id": "s3", "title": "note", "content": ""},
        ])
        issues = self._check(src, data, fmt="md")
        assert any("[QC4]" in i for i in issues), issues
        assert not any("[QC3]" in i for i in issues), issues

    def test_fail_qc4_misplaced_content_rst(self):
        src = "概要\n====\n\nA の内容。\n\n詳細\n====\n\nB の内容。\n"
        data = self._data(sections=[
            {"id": "s1", "title": "概要", "content": "B の内容。"},  # swapped
            {"id": "s2", "title": "詳細", "content": "A の内容。"},
        ])
        issues = self._check(src, data)
        assert any("[QC4]" in i and "s2" in i for i in issues), issues


# ---------------------------------------------------------------------------
# Excel QC1/QC2/QC3: verify_file(fmt="xlsx")
# ---------------------------------------------------------------------------

class TestVerifyFileExcel:
    """Excel sequential-delete: source cells must appear in JSON text."""

    def _check(self, source_path, data):
        from scripts.verify.verify import verify_file
        import tempfile, os
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as jf:
            json.dump(data, jf, ensure_ascii=False)
            jpath = jf.name
        try:
            return verify_file(source_path, jpath, "xlsx")
        finally:
            os.unlink(jpath)

    def test_pass_real_xlsx(self, tmp_path):
        """A real .xlsx with one cell 'Hello' whose value appears in JSON title."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Hello"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(xlsx_path)
        data = {"id": "f", "title": "Hello", "content": "", "sections": []}
        assert self._check(str(xlsx_path), data) == []

    def test_fail_cell_missing_from_json(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "必須セル値"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(xlsx_path)
        data = {"id": "f", "title": "別の内容", "content": "", "sections": []}
        issues = self._check(str(xlsx_path), data)
        assert any("QC1" in i for i in issues)

    def test_pass_no_knowledge_content_skipped(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "値"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(xlsx_path)
        data = {"id": "f", "title": "T", "no_knowledge_content": True, "sections": []}
        assert self._check(str(xlsx_path), data) == []

    def test_fail_qc2_fabricated_content_in_json(self, tmp_path):
        """Excel QC2: JSON text contains a string not present in any cell."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "セル値A"
        ws["A2"] = "セル値B"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(xlsx_path)
        # JSON has an extra string that no cell covers — QC2 fabrication.
        data = {"id": "f", "title": "セル値A", "content": "これは捏造された追加本文です。",
                "sections": [{"id": "s1", "title": "セル値B", "content": ""}]}
        issues = self._check(str(xlsx_path), data)
        assert any("QC2" in i for i in issues)

    def test_pass_qc2_standalone_triple_dash_is_tolerance_allowed(self, tmp_path):
        """Z-1 r8 QC2 F-QC2-1: spec §3-1 Excel 節 lists `---` explicitly
        as an allowed residue. A JSON field containing `---` (e.g. from
        a GFM table separator that lost its flanking pipes, or a
        horizontal rule fragment) must NOT trigger QC2."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Hello"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(xlsx_path)
        data = {"id": "f", "title": "Hello", "content": "---", "sections": []}
        issues = self._check(str(xlsx_path), data)
        assert not any("QC2" in i for i in issues), issues

    def test_fail_qc2_one_char_fabrication_detected(self, tmp_path):
        """Spec §3-1 Excel 節 手順 3: 空白・空行以外の残存は QC2.
        1-char residue used to be silently dropped — must FAIL now."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "ABC"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(xlsx_path)
        # JSON contains ABC (consumed) plus an extra single-char fabrication.
        data = {"id": "f", "title": "ABC X", "content": "", "sections": []}
        issues = self._check(str(xlsx_path), data)
        assert any("QC2" in i and "X" in i for i in issues)

    def test_pass_xls_cell_in_json(self, tmp_path):
        """Spec §3-1 Excel 節: `.xls` (xlrd) path must behave identically
        to `.xlsx` (openpyxl). xlwt/xlrd are hard dev deps (setup.sh)."""
        import xlwt
        wb = xlwt.Workbook()
        ws = wb.add_sheet("Sheet1")
        ws.write(0, 0, "Hello")
        xls_path = tmp_path / "test.xls"
        wb.save(str(xls_path))
        data = {"id": "f", "title": "Hello", "content": "", "sections": []}
        assert self._check(str(xls_path), data) == []

    def test_fail_xls_cell_missing_from_json(self, tmp_path):
        import xlwt
        wb = xlwt.Workbook()
        ws = wb.add_sheet("Sheet1")
        ws.write(0, 0, "必須セル値")
        xls_path = tmp_path / "test.xls"
        wb.save(str(xls_path))
        data = {"id": "f", "title": "別の内容", "content": "", "sections": []}
        issues = self._check(str(xls_path), data)
        assert any("QC1" in i for i in issues)

    def test_fail_xls_qc2_fabrication(self, tmp_path):
        """`.xls` path must raise QC2 when JSON contains a string with no
        source cell — mirroring the `.xlsx` QC2 behaviour."""
        import xlwt
        wb = xlwt.Workbook()
        ws = wb.add_sheet("Sheet1")
        ws.write(0, 0, "ABC")
        xls_path = tmp_path / "test.xls"
        wb.save(str(xls_path))
        data = {"id": "f", "title": "ABC 捏造", "content": "", "sections": []}
        issues = self._check(str(xls_path), data)
        assert any("QC2" in i and "捏造" in i for i in issues)

    def test_fail_xls_numeric_cell_missing_from_json(self, tmp_path):
        """xlrd numeric cells must be tokenised and compared. If the cell
        is absent from JSON entirely, QC1 must FAIL regardless of
        float/int representation."""
        import xlwt
        wb = xlwt.Workbook()
        ws = wb.add_sheet("Sheet1")
        ws.write(0, 0, 12345)
        xls_path = tmp_path / "test.xls"
        wb.save(str(xls_path))
        data = {"id": "f", "title": "無関係", "content": "", "sections": []}
        issues = self._check(str(xls_path), data)
        assert any("QC1" in i for i in issues)

    def test_fail_qc3_duplicate_cell_in_json(self, tmp_path):
        """Excel QC3: two source cells with the same value but JSON only
        contains that value once → second match falls into the consumed
        region → QC3 duplicate."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "同じ"
        ws["B1"] = "同じ"
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(xlsx_path)
        data = {"id": "f", "title": "同じ", "content": "",
                "sections": []}
        issues = self._check(str(xlsx_path), data)
        assert any("[QC3]" in i for i in issues)
        assert not any("[QC1]" in i or "[QC2]" in i for i in issues)


# ---------------------------------------------------------------------------
# Phase 22-B: Excel P1 — per-sheet tokens + QO2 one-way containment
# ---------------------------------------------------------------------------


class TestXlsxSourceTokensPerSheet:
    """Phase 22-B-5: sheet-level file split requires per-sheet source tokens.

    `_xlsx_source_tokens(source_path, sheet_name=...)` must scope the token
    extraction to a single worksheet so that a sheet-specific JSON is verified
    only against cells from its own sheet. Calling without `sheet_name` must
    preserve the prior all-sheet behaviour (back-compat for ad-hoc callers —
    the production dispatch always passes a name).
    """

    def test_per_sheet_tokens_restricted_to_named_sheet(self, tmp_path):
        import openpyxl
        from scripts.verify.verify import _xlsx_source_tokens
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Alpha"
        ws1["A1"] = "alpha-only"
        ws2 = wb.create_sheet("Beta")
        ws2["A1"] = "beta-only"
        xlsx_path = tmp_path / "two.xlsx"
        wb.save(xlsx_path)

        alpha_tokens = _xlsx_source_tokens(str(xlsx_path), sheet_name="Alpha")
        beta_tokens = _xlsx_source_tokens(str(xlsx_path), sheet_name="Beta")
        assert "alpha-only" in alpha_tokens
        assert "beta-only" not in alpha_tokens
        assert "beta-only" in beta_tokens
        assert "alpha-only" not in beta_tokens

    def test_all_sheets_when_sheet_name_omitted(self, tmp_path):
        import openpyxl
        from scripts.verify.verify import _xlsx_source_tokens
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Alpha"
        ws1["A1"] = "alpha-only"
        ws2 = wb.create_sheet("Beta")
        ws2["A1"] = "beta-only"
        xlsx_path = tmp_path / "two.xlsx"
        wb.save(xlsx_path)

        tokens = _xlsx_source_tokens(str(xlsx_path), sheet_name=None)
        assert "alpha-only" in tokens
        assert "beta-only" in tokens

    def test_per_sheet_xls_tokens_restricted_to_named_sheet(self, tmp_path):
        import xlwt
        from scripts.verify.verify import _xlsx_source_tokens
        wb = xlwt.Workbook()
        ws1 = wb.add_sheet("Alpha")
        ws1.write(0, 0, "alpha-xls")
        ws2 = wb.add_sheet("Beta")
        ws2.write(0, 0, "beta-xls")
        xls_path = tmp_path / "two.xls"
        wb.save(str(xls_path))

        alpha = _xlsx_source_tokens(str(xls_path), sheet_name="Alpha")
        beta = _xlsx_source_tokens(str(xls_path), sheet_name="Beta")
        assert "alpha-xls" in alpha and "beta-xls" not in alpha
        assert "beta-xls" in beta and "alpha-xls" not in beta


class TestCheckJsonDocsMdConsistency_QO2_ExcelP1:
    """Phase 22-B-5: spec §3-3 QO2 Excel 例外 — for P1 sheets, verify uses
    one-way containment (JSON section.content text ⊂ docs MD), because the
    docs MD table has structural tokens (`|---|`, column headers) that do
    not appear verbatim in JSON. Non-Excel / non-P1 JSON still uses strict
    verbatim containment.
    """

    def _check(self, data, docs_md_text):
        from scripts.verify.verify import check_json_docs_md_consistency
        return check_json_docs_md_consistency(data, docs_md_text)

    def test_pass_p1_section_content_tokens_all_in_md_table(self):
        # JSON section.content is vertical "列名: 値" enumeration; docs MD
        # renders the same row as an MD table. Each non-empty token from
        # section.content must appear in the MD, but the table's pipes /
        # separators are only in MD.
        data = {
            "id": "f",
            "title": "リリースノート Sheet1",
            "content": "",
            "sheet_type": "P1",
            "sections": [
                {
                    "id": "s1",
                    "title": "バグ-001",
                    "content": "No.: 001\n種別: バグ\nタイトル: バグ-001\n対応: 修正済",
                },
            ],
        }
        docs = (
            "# リリースノート Sheet1\n\n"
            "## バグ-001\n\n"
            "| No. | 種別 | タイトル | 対応 |\n"
            "| --- | --- | --- | --- |\n"
            "| 001 | バグ | バグ-001 | 修正済 |\n"
        )
        assert self._check(data, docs) == []

    def test_fail_p1_section_content_token_missing_from_md(self):
        # If one of the "値" tokens from JSON is missing from docs MD, QO2
        # must still FAIL (the one-way check still enforces JSON ⊂ MD).
        data = {
            "id": "f",
            "title": "リリースノート Sheet1",
            "content": "",
            "sheet_type": "P1",
            "sections": [
                {
                    "id": "s1",
                    "title": "バグ-001",
                    "content": "No.: 001\n種別: バグ\nタイトル: バグ-001\n対応: 修正済",
                },
            ],
        }
        # "修正済" is missing from the MD table row.
        docs = (
            "# リリースノート Sheet1\n\n"
            "## バグ-001\n\n"
            "| No. | 種別 | タイトル | 対応 |\n"
            "| --- | --- | --- | --- |\n"
            "| 001 | バグ | バグ-001 |  |\n"
        )
        issues = self._check(data, docs)
        assert any("[QO2]" in i and "修正済" in i for i in issues), issues

    def test_pass_p2_falls_back_to_strict_verbatim(self):
        # P2: sheet-wide text, JSON content appears verbatim in MD.
        data = {
            "id": "f",
            "title": "対応表 分類",
            "content": "カテゴリA\n項目1\n項目2\n",
            "sheet_type": "P2",
            "sections": [],
        }
        docs = "# 対応表 分類\n\nカテゴリA\n項目1\n項目2\n"
        assert self._check(data, docs) == []

    def test_pass_p1_value_containing_colon_preserved(self):
        # QA Finding: the implementation uses `partition(":")` so the
        # value after the first colon is preserved. Pin this: a value
        # containing `:` (e.g. a URL) must be checked whole, not split.
        data = {
            "id": "f",
            "title": "リリースノート",
            "content": "",
            "sheet_type": "P1",
            "sections": [
                {
                    "id": "s1",
                    "title": "row1",
                    "content": "URL: https://example.com/path",
                },
            ],
        }
        # PASS: full URL present in MD.
        docs_ok = "# リリースノート\n\n## row1\n\nhttps://example.com/path\n"
        assert self._check(data, docs_ok) == []
        # FAIL: only the scheme is present — full value missing.
        docs_bad = "# リリースノート\n\n## row1\n\nhttps\n"
        issues = self._check(data, docs_bad)
        assert any(
            "[QO2]" in i and "https://example.com/path" in i for i in issues
        ), issues

    def test_pass_p1_key_with_empty_value_skipped(self):
        # QA Finding: a blank-cell row renders as `列名: ` (empty value).
        # The implementation's `if not value: continue` must silently skip
        # such lines. Pin this so future refactors don't regress.
        data = {
            "id": "f",
            "title": "リリースノート",
            "content": "",
            "sheet_type": "P1",
            "sections": [
                {
                    "id": "s1",
                    "title": "row1",
                    "content": "No.: 001\n対応:",
                },
            ],
        }
        docs = "# リリースノート\n\n## row1\n\n| No. | 対応 |\n| --- | --- |\n| 001 |  |\n"
        assert self._check(data, docs) == []

    def test_pass_no_sheet_type_falls_through_to_strict_verbatim(self):
        # SE Observation: non-Excel JSON has no `sheet_type` field; the
        # P1 gate must not trigger. Pin this with a regular RST-derived
        # JSON shape and assert strict containment still applies.
        data = {
            "id": "f",
            "title": "T",
            "content": "",
            "sections": [
                {"id": "s1", "title": "概要", "content": "本文1\n本文2"},
            ],
        }
        docs = "# T\n\n## 概要\n\n本文1\n本文2\n"
        assert self._check(data, docs) == []

    def test_fail_p2_top_content_not_verbatim(self):
        data = {
            "id": "f",
            "title": "対応表 分類",
            "content": "カテゴリA\n項目1\n項目2\n",
            "sheet_type": "P2",
            "sections": [],
        }
        # Content differs (項目2 missing).
        docs = "# 対応表 分類\n\nカテゴリA\n項目1\n"
        issues = self._check(data, docs)
        assert any("[QO2]" in i for i in issues), issues


# ---------------------------------------------------------------------------
# QL1: 内部リンク (check_source_links)
# ---------------------------------------------------------------------------

class TestCheckSourceLinks:
    """QL1: Internal links in source must be reflected in JSON."""

    def _check(self, source_text, fmt, data, label_map=None):
        from scripts.verify.verify import check_source_links
        return check_source_links(source_text, fmt, data, label_map or {})

    def _data(self, content="", sections=None):
        return {
            "id": "f", "title": "T", "content": content,
            "sections": sections or []
        }

    # RST :ref:
    def test_pass_rst_ref_display_text_in_json(self):
        src = "詳細は :ref:`使い方 <usage>` を参照。\n"
        data = self._data(content="詳細は 使い方 を参照。")
        assert self._check(src, "rst", data) == []

    def test_fail_rst_ref_display_text_missing(self):
        src = "詳細は :ref:`使い方 <usage>` を参照。\n"
        data = self._data(content="詳細を参照。")
        issues = self._check(src, "rst", data)
        assert any("QL1" in i and "使い方" in i for i in issues)

    def test_pass_rst_ref_plain_label_resolved(self):
        src = ":ref:`usage`\n"
        data = self._data(content="使い方セクション")
        label_map = {"usage": "使い方セクション"}
        assert self._check(src, "rst", data, label_map) == []

    def test_fail_rst_ref_plain_label_title_missing(self):
        src = ":ref:`usage`\n"
        data = self._data(content="別の内容")
        label_map = {"usage": "使い方セクション"}
        issues = self._check(src, "rst", data, label_map)
        assert any("QL1" in i and "usage" in i for i in issues)

    def test_pass_rst_ref_unknown_label_skipped(self):
        src = ":ref:`cross-file-label`\n"
        data = self._data(content="内容")
        # Unknown label (not in label_map) → cannot verify cross-file ref → PASS
        assert self._check(src, "rst", data, {}) == []

    # MD internal links
    def test_pass_md_internal_link_text_in_json(self):
        src = "[使い方](./usage.md)\n"
        data = self._data(content="使い方")
        assert self._check(src, "md", data) == []

    def test_fail_md_internal_link_text_missing(self):
        src = "[使い方](./usage.md)\n"
        data = self._data(content="別の内容")
        issues = self._check(src, "md", data)
        assert any("QL1" in i and "使い方" in i for i in issues)

    def test_pass_md_external_link_skipped(self):
        """External links (https://) are QL2, not QL1."""
        src = "[外部サイト](https://example.com)\n"
        data = self._data(content="内容")
        assert self._check(src, "md", data) == []

    def test_pass_no_knowledge_content_skipped(self):
        src = ":ref:`something`\n"
        data = {"id": "f", "title": "T", "no_knowledge_content": True, "sections": []}
        assert self._check(src, "rst", data) == []

    # --- QL1 Z-1 gap fill: RST figure / image + MD image ------------------

    def test_fail_rst_figure_caption_missing(self):
        src = (
            "概要\n====\n\n"
            ".. figure:: images/sample.png\n"
            "\n"
            "   サンプル画像の説明\n"
        )
        data = self._data(content="まったく別の内容")
        issues = self._check(src, "rst", data)
        assert any("QL1" in i and "caption" in i and "サンプル画像の説明" in i for i in issues)

    def test_pass_rst_figure_caption_in_json(self):
        src = (
            "概要\n====\n\n"
            ".. figure:: images/sample.png\n"
            "\n"
            "   サンプル画像の説明\n"
        )
        data = self._data(content="サンプル画像の説明 が JSON にある。")
        assert self._check(src, "rst", data) == []

    def test_pass_rst_substitution_image_body_skipped(self):
        """Spec §3-2 line 268: substitution-body content (e.g. `.. |x|
        image::` body) must be excluded from QL1 by AST-attribute /
        parent-ancestry. The substituted occurrence is the reader-visible
        one.

        Z-1 r8 QL1 F1: this fixture pins the skip without relying on
        coincidental JSON containment. The substitution is NEVER
        referenced in the body, so the only `image` node docutils emits
        lives under the `substitution_definition` subtree. JSON has no
        mention of the alt text or filename. Without `_under_substitution`,
        QL1 would emit a FAIL naming 'アイコン' (or 'icon.png'); with
        the skip, zero QL1 issues fire.
        """
        src = (
            "概要\n====\n\n"
            "本文のみ。\n\n"
            ".. |unused| image:: images/icon.png\n"
            "   :alt: アイコン\n"
        )
        data = self._data(content="本文のみ。")
        issues = self._check(src, "rst", data)
        qc = [i for i in issues if "QL1" in i and ("アイコン" in i or "icon.png" in i)]
        assert qc == [], qc

    def test_pass_rst_figure_dedup_same_caption_not_reported_twice(self):
        """Z-1 r8 QL1 F2: RST figure dedup (mirror of image dedup).
        Two `.. figure::` blocks sharing the same caption text; JSON
        omits the caption — QL1 fires once, not once per figure."""
        src = (
            "概要\n====\n\n"
            ".. figure:: images/a.png\n"
            "\n"
            "   共通キャプション\n"
            "\n"
            ".. figure:: images/b.png\n"
            "\n"
            "   共通キャプション\n"
        )
        data = self._data(content="別の内容")
        issues = self._check(src, "rst", data)
        qc = [i for i in issues if "QL1" in i and "共通キャプション" in i]
        assert len(qc) == 1, qc

    def test_pass_rst_image_dedup_same_alt_not_reported_twice(self):
        """Z-1 r7 QL1 F2: RST image dedup (mirror of MD's seen_images).
        When the same alt text appears on two images in one file and
        JSON omits it, QL1 must fire once — not once per occurrence."""
        src = (
            "概要\n====\n\n"
            ".. image:: images/a.png\n"
            "   :alt: 共通ロゴ\n"
            "\n"
            ".. image:: images/b.png\n"
            "   :alt: 共通ロゴ\n"
        )
        data = self._data(content="別の内容")
        issues = self._check(src, "rst", data)
        qc = [i for i in issues if "QL1" in i and "共通ロゴ" in i]
        assert len(qc) == 1, qc

    def test_fail_rst_figure_inline_only_caption_fallback_to_filename(self):
        """When caption is only RST inline syntax (e.g. [1]_), fall back to
        the image filename (§3-2 table: 'caption が RST inline 構文のみ...')."""
        src = (
            "概要\n====\n\n"
            ".. figure:: images/badge.png\n"
            "\n"
            "   [1]_\n"
        )
        data = self._data(content="別の内容")
        issues = self._check(src, "rst", data)
        assert any("QL1" in i and "badge.png" in i for i in issues)

    def test_fail_rst_image_alt_missing(self):
        src = (
            "概要\n====\n\n"
            ".. image:: images/logo.png\n"
            "   :alt: 会社ロゴ\n"
        )
        data = self._data(content="別の内容")
        issues = self._check(src, "rst", data)
        assert any("QL1" in i and "会社ロゴ" in i for i in issues)

    def test_fail_rst_image_without_alt_falls_back_to_filename(self):
        src = (
            "概要\n====\n\n"
            ".. image:: images/diagram.png\n"
        )
        data = self._data(content="別の内容")
        issues = self._check(src, "rst", data)
        assert any("QL1" in i and "diagram.png" in i for i in issues)

    def test_fail_md_image_alt_missing(self):
        src = "# T\n\n![会社ロゴ](./logo.png)\n"
        data = self._data(content="別の内容")
        issues = self._check(src, "md", data)
        assert any("QL1" in i and "会社ロゴ" in i for i in issues)

    def test_fail_md_image_without_alt_falls_back_to_filename(self):
        src = "# T\n\n![](./diagram.png)\n"
        data = self._data(content="別の内容")
        issues = self._check(src, "md", data)
        assert any("QL1" in i and "diagram.png" in i for i in issues)

    def test_pass_md_image_alt_in_json(self):
        src = "# T\n\n![会社ロゴ](./logo.png)\n"
        data = self._data(content="会社ロゴ が載っています。")
        assert self._check(src, "md", data) == []

    # --- QL1 RST native named reference (r2 critical fix 3) ---------------

    def test_fail_rst_named_reference_target_title_missing(self):
        """Real RST: `\\`Target\\`_` references a `.. _label:` defined in
        the same doc. docutils resolves it to a reference node with refid
        pointing at the target section; the resolved title must appear
        in JSON (spec §3-2 row 1)."""
        src = (
            "See `Detailed Usage`_ for more.\n\n"
            ".. _Detailed Usage:\n\n"
            "Detailed Usage\n"
            "==============\n\n"
            "Section body text.\n"
        )
        data = self._data(content="Totally unrelated content.")
        issues = self._check(src, "rst", data, {})
        assert any("QL1" in i and "Detailed Usage" in i for i in issues)

    def test_pass_rst_named_reference_target_title_in_json(self):
        """Same fixture but JSON contains the resolved title — no FAIL."""
        src = (
            "See `Detailed Usage`_ for more.\n\n"
            ".. _Detailed Usage:\n\n"
            "Detailed Usage\n"
            "==============\n\n"
            "Body.\n"
        )
        data = self._data(content="Detailed Usage and body appear here.")
        issues = self._check(src, "rst", data, {})
        assert not any("QL1" in i and "Detailed Usage" in i for i in issues)

    def test_pass_md_mailto_link_not_internal(self):
        """`mailto:` hrefs are not document-to-document internal links —
        QL1 must not demand their 'link text' in JSON (r3 QL1 High gap)."""
        src = "# T\n\n連絡先 [メール](mailto:dev@example.com) 宛。\n"
        data = self._data(content="連絡先 宛。")  # "メール" not in JSON, but OK
        assert self._check(src, "md", data) == []

    def test_pass_md_anchor_only_link_not_internal(self):
        """In-document anchor links (`#section`) are navigation, not
        cross-document references; QL1 scope excludes them."""
        src = "# T\n\n[上部へ](#top) に戻る\n"
        data = self._data(content="に戻る")
        assert self._check(src, "md", data) == []

    def test_pass_md_tel_link_not_internal(self):
        src = "# T\n\n[電話](tel:+81-3-1234-5678) はこちら\n"
        data = self._data(content="はこちら")
        assert self._check(src, "md", data) == []

    def test_fail_md_image_title_missing_from_json(self):
        """Spec §3-2 row 6: MD image alt / title / src filename must be
        in JSON. An image with only a title attribute must FAIL when the
        title is absent from JSON."""
        src = '# T\n\n![](./img.png "会社ロゴ")\n'
        data = self._data(content="別の内容")
        issues = self._check(src, "md", data)
        assert any("QL1" in i and "会社ロゴ" in i for i in issues)

    def test_pass_rst_plain_sections_without_named_references(self):
        """Vanilla sections (no user-defined label, no named reference)
        must not emit spurious QL1 — docutils gives each section an
        auto-id like `section-N`, which QL1 must ignore."""
        src = "Alpha\n=====\n\nA body\n\nBeta\n====\n\nB body\n"
        data = self._data(content="Alpha Beta A body B body")
        assert self._check(src, "rst", data, {}) == []

    def test_pass_rst_named_reference_scheme_mailto_untouched(self):
        """A bare `.. _Contact: mailto:ops@example.com` is not a named
        section reference; QL1 must not raise on it. (Sanity check that
        refuri-bearing references stay in QL2's lane.)"""
        src = (
            "Contact_ us.\n\n"
            ".. _Contact: mailto:ops@example.com\n"
        )
        data = self._data(content="us.")
        assert self._check(src, "rst", data, {}) == []


# ---------------------------------------------------------------------------
# Phase 22-B-5a-r3a: P1 header row expansion
# ---------------------------------------------------------------------------


class TestVerifyFileExcelP1HeaderExpansion:
    """Spec §3-1 Excel 節 (Phase 22-B): P1 シートでは header 行のセル値
    (列名) が、JSON の section.content に「データ行数」回現れる。sequential-
    delete の前に header 行トークンをデータ行数ぶん複製してから照合する。

    - converter を参照せず、§8-2 header 検出 + §8-4 データ行数算出のみ
      に依拠すること
    - P2 シートは展開なし (従来通り 1:1 照合)
    - converter が header 行を多重展開した場合は QC2 で検出される
      (複製数 = データ行数 = JSON section 数 + 何かが食い違えば残存)
    """

    def _check(self, source_path, data, sheet_name=None):
        from scripts.verify.verify import verify_file
        import tempfile, os
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as jf:
            json.dump(data, jf, ensure_ascii=False)
            jpath = jf.name
        try:
            return verify_file(source_path, jpath, "xlsx", sheet_name=sheet_name)
        finally:
            os.unlink(jpath)

    def _p1_sheet(self, tmp_path):
        """Build a minimal 3-column P1 sheet: header + 2 data rows."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "No."
        ws["B1"] = "タイトル"
        ws["C1"] = "概要"
        ws["A2"] = "1"
        ws["B2"] = "値A"
        ws["C2"] = "概要A"
        ws["A3"] = "2"
        ws["B3"] = "値B"
        ws["C3"] = "概要B"
        xlsx_path = tmp_path / "p1.xlsx"
        wb.save(xlsx_path)
        return xlsx_path

    def _p1_json(self):
        """JSON that correctly mirrors the P1 sheet: 2 sections, each
        containing the full {列名}: {値} line-listing.  Title defaults
        to the sheet name (spec §8-4 fallback) since the sheet has no
        ``■...`` row — verify injects this as a synthetic source token."""
        return {
            "id": "p1",
            "title": "Sheet1",
            "content": "",
            "sheet_type": "P1",
            "sections": [
                {
                    "id": "s1",
                    "title": "値A",
                    "content": "No.: 1\nタイトル: 値A\n概要: 概要A",
                },
                {
                    "id": "s2",
                    "title": "値B",
                    "content": "No.: 2\nタイトル: 値B\n概要: 概要B",
                },
            ],
        }

    def test_pass_p1_column_names_repeat_per_data_row(self, tmp_path):
        """P1 converter output where each header cell appears 1 time in
        source but `data_rows` times in JSON must PASS — header expansion
        converts the 1 source token into N tokens matching JSON."""
        xlsx_path = self._p1_sheet(tmp_path)
        data = self._p1_json()
        issues = self._check(str(xlsx_path), data, sheet_name="Sheet1")
        assert issues == [], issues

    def test_fail_p1_data_value_missing_from_json(self, tmp_path):
        """A data-row cell value absent from JSON must still raise QC1.
        Header expansion applies only to header cells, not data cells."""
        xlsx_path = self._p1_sheet(tmp_path)
        data = self._p1_json()
        # Drop section s2's 値B — the cell '値B' is still in title, but
        # '概要B' is not anywhere → QC1.
        data["sections"][1]["content"] = "No.: 2\nタイトル: 値B"
        issues = self._check(str(xlsx_path), data, sheet_name="Sheet1")
        assert any("[QC1]" in i and "概要B" in i for i in issues), issues

    def test_fail_p1_column_name_over_expanded_in_json(self, tmp_path):
        """Converter bug: a header cell value appears (data_rows + 1) times
        in JSON. After header expansion (data_rows = 2 tokens) + 1 normal
        occurrence = 3 source tokens, sequential-delete consumes 3 but JSON
        has 4 → 1 residue → QC2 (or QC3 if the extra sits in a consumed
        region)."""
        xlsx_path = self._p1_sheet(tmp_path)
        data = self._p1_json()
        # Inject an extra 'タイトル:' line in s1 — column name now appears
        # 3 times in JSON but source (with header expansion = 2) + 0
        # elsewhere = 2 source tokens → 1 residue.
        data["sections"][0]["content"] += "\nタイトル: 重複"
        issues = self._check(str(xlsx_path), data, sheet_name="Sheet1")
        # '重複' is never in the source → QC1 fires for that specifically,
        # AND 'タイトル' is now over-count. Accept any FAIL (QC1 / QC2 / QC3).
        assert issues, "Expected at least one FAIL for over-expanded header"

    def test_fail_p1_column_name_under_represented(self, tmp_path):
        """If JSON does not include a column name at all (converter bug
        dropped the `{列名}: ` prefix), header expansion produces
        data_rows tokens but none match → data_rows QC1 FAILs for that
        column."""
        xlsx_path = self._p1_sheet(tmp_path)
        data = self._p1_json()
        # Remove "タイトル: " prefix entirely from both sections.
        data["sections"][0]["content"] = "No.: 1\n値A\n概要: 概要A"
        data["sections"][1]["content"] = "No.: 2\n値B\n概要: 概要B"
        issues = self._check(str(xlsx_path), data, sheet_name="Sheet1")
        # 'タイトル' has source=1 → after expansion = 2 tokens. Neither
        # appears in JSON now → 2 QC1 (or QC3 depending on implementation).
        qc_issues = [i for i in issues if "タイトル" in i]
        assert len(qc_issues) >= 1, issues

    def test_pass_p2_no_header_expansion(self, tmp_path):
        """P2 sheets (段落主体) must NOT get header expansion — 1:1 check
        applies. A header-like row appearing once in source must match
        once in JSON, not N times."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # P2 indicator: only 1 non-empty cell per row (paragraph-like).
        ws["A1"] = "■概要"
        ws["A2"] = "これは段落のシートです。"
        xlsx_path = tmp_path / "p2.xlsx"
        wb.save(xlsx_path)
        data = {
            "id": "p2",
            # Title includes the ■ prefix verbatim (spec §8-4 lets the
            # ■... cell stand as title; stripping would break 1:1 check).
            "title": "■概要",
            "content": "これは段落のシートです。",
            "sheet_type": "P2",
            "sections": [],
        }
        issues = self._check(str(xlsx_path), data, sheet_name="Sheet1")
        assert issues == [], issues

    def test_fail_p2_duplicated_value_triggers_qc3(self, tmp_path):
        """P2 1:1 check: if JSON duplicates a value the source only has
        once, QC3 must fire (no header expansion to paper over it)."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "唯一の値"
        xlsx_path = tmp_path / "p2.xlsx"
        wb.save(xlsx_path)
        data = {
            "id": "p2",
            "title": "",
            "content": "唯一の値 唯一の値",  # written twice
            "sheet_type": "P2",
            "sections": [],
        }
        issues = self._check(str(xlsx_path), data, sheet_name="Sheet1")
        assert any("[QC3]" in i or "[QC2]" in i for i in issues), issues

    def test_pass_p1_multi_row_header_merged(self, tmp_path):
        """§8-3 multi-row header: sub-header rows merge as 'メイン/副'.
        verify must detect the merge using the same rule as converter
        (§8-2), then expand merged-column names by data-row count."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # Main header (5 contiguous non-empty to satisfy run-length ≥ 3)
        ws["A1"] = "No."
        ws["B1"] = "タイトル"
        ws["C1"] = "バージョン"  # parent with two subs
        ws["E1"] = "備考"
        # Sub-header (strict subset of main, narrower)
        ws["C2"] = "モジュール"
        ws["D2"] = "Nablarch"
        # Data rows
        ws["A3"] = "1"
        ws["B3"] = "初回対応"
        ws["C3"] = "1.0"
        ws["D3"] = "5"
        ws["E3"] = "初回"
        ws["A4"] = "2"
        ws["B4"] = "二回目"
        ws["C4"] = "2.0"
        ws["D4"] = "6"
        ws["E4"] = "追加対応"
        xlsx_path = tmp_path / "multihdr.xlsx"
        wb.save(xlsx_path)
        data = {
            "id": "mh",
            # Title = sheet name per spec §8-4 fallback.
            "title": "Sheet1",
            "content": "",
            "sheet_type": "P1",
            "sections": [
                {
                    # section.title = タイトル column value (§8-4 rule)
                    "id": "s1", "title": "初回対応",
                    "content": "No.: 1\nタイトル: 初回対応\nバージョン/モジュール: 1.0\nバージョン/Nablarch: 5\n備考: 初回",
                },
                {
                    "id": "s2", "title": "二回目",
                    "content": "No.: 2\nタイトル: 二回目\nバージョン/モジュール: 2.0\nバージョン/Nablarch: 6\n備考: 追加対応",
                },
            ],
        }
        issues = self._check(str(xlsx_path), data, sheet_name="Sheet1")
        assert issues == [], issues


# ---------------------------------------------------------------------------
# Phase 22-B-5a-r3b: QP — P1 column-value pairing (spec §3-4)
# ---------------------------------------------------------------------------


class TestCheckXlsxP1Pairing:
    """Spec §3-4 (new): For each P1 sheet, JSON section N must pair up
    every `{列名}: {値}` with Excel data-row N's matching cell at the
    same column. Bag matching (QC1–QC4) cannot catch swapped values —
    QP does.
    """

    def _call(self, source_path, data, sheet_name):
        from scripts.verify.verify import check_xlsx_p1_pairing
        return check_xlsx_p1_pairing(source_path, data, sheet_name)

    def _sheet(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "No."
        ws["B1"] = "タイトル"
        ws["C1"] = "概要"
        ws["A2"] = "1"
        ws["B2"] = "値A"
        ws["C2"] = "概要A"
        ws["A3"] = "2"
        ws["B3"] = "値B"
        ws["C3"] = "概要B"
        xlsx_path = tmp_path / "p1.xlsx"
        wb.save(xlsx_path)
        return xlsx_path

    def _good_json(self):
        return {
            "id": "p1",
            "title": "",
            "content": "",
            "sheet_type": "P1",
            "sections": [
                {"id": "s1", "title": "値A",
                 "content": "No.: 1\nタイトル: 値A\n概要: 概要A"},
                {"id": "s2", "title": "値B",
                 "content": "No.: 2\nタイトル: 値B\n概要: 概要B"},
            ],
        }

    def test_pass_aligned_pairs(self, tmp_path):
        xlsx = self._sheet(tmp_path)
        assert self._call(str(xlsx), self._good_json(), "Sheet1") == []

    def test_fail_swapped_values_across_rows(self, tmp_path):
        """Row-1 value ends up in row-2 section — the exact gap QP exists
        to close. QC1–QC4 would PASS this (tokens are present somewhere)."""
        xlsx = self._sheet(tmp_path)
        data = self._good_json()
        # Swap タイトル values between sections.
        data["sections"][0]["content"] = "No.: 1\nタイトル: 値B\n概要: 概要A"
        data["sections"][1]["content"] = "No.: 2\nタイトル: 値A\n概要: 概要B"
        issues = self._call(str(xlsx), data, "Sheet1")
        assert any("[QP]" in i for i in issues), issues

    def test_fail_section_count_mismatch(self, tmp_path):
        """If JSON has fewer/more sections than Excel has data rows,
        QP raises section_count_mismatch (independent of QC1)."""
        xlsx = self._sheet(tmp_path)
        data = self._good_json()
        # Drop the second section.
        data["sections"] = data["sections"][:1]
        issues = self._call(str(xlsx), data, "Sheet1")
        assert any("[QP]" in i and "section" in i.lower() for i in issues), issues

    def test_fail_pair_missing_expected_column(self, tmp_path):
        """A non-empty source cell omitted from the section's content
        must FAIL — it means the section silently dropped a column."""
        xlsx = self._sheet(tmp_path)
        data = self._good_json()
        # Drop 概要 from s1.
        data["sections"][0]["content"] = "No.: 1\nタイトル: 値A"
        issues = self._call(str(xlsx), data, "Sheet1")
        assert any("[QP]" in i and "概要" in i for i in issues), issues

    def test_pass_empty_source_cell_omitted_from_json(self, tmp_path):
        """Spec §3-4: 空セルは期待ペアに含めない。空セル列を section に
        書かないのは許容。converter が空セルを出さないのは設計通り。"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "No."
        ws["B1"] = "タイトル"
        ws["C1"] = "概要"
        ws["A2"] = "1"
        ws["B2"] = "値A"
        # C2 intentionally empty.
        ws["A3"] = "2"
        ws["B3"] = "値B"
        ws["C3"] = "概要B"
        xlsx_path = tmp_path / "p1.xlsx"
        wb.save(xlsx_path)
        data = {
            "id": "p1",
            "title": "",
            "content": "",
            "sheet_type": "P1",
            "sections": [
                {"id": "s1", "title": "値A",
                 "content": "No.: 1\nタイトル: 値A"},  # no 概要 line
                {"id": "s2", "title": "値B",
                 "content": "No.: 2\nタイトル: 値B\n概要: 概要B"},
            ],
        }
        assert self._call(str(xlsx_path), data, "Sheet1") == []

    def test_skip_non_p1(self, tmp_path):
        """P2 JSON must not be checked — return no issues immediately."""
        xlsx = self._sheet(tmp_path)
        data = self._good_json()
        data["sheet_type"] = "P2"
        assert self._call(str(xlsx), data, "Sheet1") == []

    def test_fail_value_contains_colon(self, tmp_path):
        """Spec §8-4 + existing QO2 P1: partition on FIRST `:` so values
        containing `:` are preserved verbatim. QP must use the same rule.
        """
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # 3 columns to escape the §8-2 ≤2-col P2 cap.
        ws["A1"] = "No."
        ws["B1"] = "URL"
        ws["C1"] = "備考"
        ws["A2"] = "1"
        ws["B2"] = "https://example.com/x"
        ws["C2"] = "備考X"
        ws["A3"] = "2"
        ws["B3"] = "https://example.com/y"
        ws["C3"] = "備考Y"
        xlsx_path = tmp_path / "p1.xlsx"
        wb.save(xlsx_path)
        data = {
            "id": "p1",
            "title": "", "content": "",
            "sheet_type": "P1",
            "sections": [
                # Swap URLs between rows → QP must FAIL on pair mismatch.
                {"id": "s1", "title": "https://example.com/y",
                 "content": "No.: 1\nURL: https://example.com/y\n備考: 備考X"},
                {"id": "s2", "title": "https://example.com/x",
                 "content": "No.: 2\nURL: https://example.com/x\n備考: 備考Y"},
            ],
        }
        issues = self._call(str(xlsx_path), data, "Sheet1")
        assert any("[QP]" in i for i in issues), issues


# ---------------------------------------------------------------------------
# Phase 22-B-5a-r3 QA review Findings (7 blockers)
# ---------------------------------------------------------------------------


class TestVerifyP1HeaderBoundary:
    """QA F1: P2 boundary cases (run-length=2, useful_width=2).

    §8-2 is explicit: header requires contiguous non-empty ≥ 3 AND useful
    width > 2.  Boundaries below must be classified P2 (tokens raw 1:1,
    no expansion).
    """

    def _tokens(self, source_path, sheet_name, sheet_type):
        from scripts.verify.verify import _xlsx_source_tokens
        return _xlsx_source_tokens(source_path, sheet_name=sheet_name, sheet_type=sheet_type)

    def test_run_length_2_classified_p2(self, tmp_path):
        """Header-row with contiguous non-empty = 2 (≠ P1 since < 3).
        Sheet has ≥ 3 useful columns overall (so it is not the width cap)."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # Row 1: 2 contiguous non-empty + gap + 1 more → run length = 2
        ws["A1"] = "No."
        ws["B1"] = "タイトル"
        # C1 empty
        ws["D1"] = "備考"
        # Data rows (3 useful columns overall)
        ws["A2"] = "1"
        ws["B2"] = "値A"
        ws["D2"] = "r1"
        ws["A3"] = "2"
        ws["B3"] = "値B"
        ws["D3"] = "r2"
        xlsx_path = tmp_path / "rl2.xlsx"
        wb.save(xlsx_path)
        raw = self._tokens(str(xlsx_path), "Sheet1", None)
        as_p1 = self._tokens(str(xlsx_path), "Sheet1", "P1")
        # run_length < 3 → P2 fall-through even with sheet_type=P1.
        assert raw == as_p1, (
            "run_length=2 sheet must not be expanded; tokens should equal raw 1:1"
        )

    def test_useful_width_2_classified_p2(self, tmp_path):
        """≤ 2 useful columns: spec §8-2 forces P2 even if header detection
        would otherwise succeed."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # Only 2 cols used, 3 contiguous in header row would fail anyway.
        ws["A1"] = "No."
        ws["B1"] = "タイトル"
        ws["A2"] = "1"
        ws["B2"] = "値A"
        ws["A3"] = "2"
        ws["B3"] = "値B"
        xlsx_path = tmp_path / "w2.xlsx"
        wb.save(xlsx_path)
        raw = self._tokens(str(xlsx_path), "Sheet1", None)
        as_p1 = self._tokens(str(xlsx_path), "Sheet1", "P1")
        assert raw == as_p1, (
            "useful_width=2 sheet must not be P1-expanded"
        )


class TestVerifyP1InvalidSubHeader:
    """QA F2: Row 2 wider than or equal to row 1 is NOT a sub-header.
    §8-3 specifies only "副列がある列のみ" — a row equal/wider than the
    main header is a second header candidate or data, not a sub-header.
    """

    def test_sub_header_wider_than_parent_rejected(self, tmp_path):
        """Row 2 has MORE non-empty cells than row 1 → must NOT merge."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # Row 1: 3 contiguous non-empty (qualifies as main header)
        ws["A1"] = "No."
        ws["B1"] = "バージョン"
        ws["C1"] = "備考"
        # Row 2: wider/equal — 4 non-empty cells; NOT a sub-header
        ws["A2"] = "1"
        ws["B2"] = "1.0"
        ws["C2"] = "初回"
        ws["D2"] = "追加"
        # Row 3: 2nd data row
        ws["A3"] = "2"
        ws["B3"] = "2.0"
        ws["C3"] = "二回目"
        ws["D3"] = "追加2"
        xlsx_path = tmp_path / "invsub.xlsx"
        wb.save(xlsx_path)
        from scripts.verify.verify import _detect_header_row, _read_sheet_matrix
        rows = _read_sheet_matrix(str(xlsx_path), "Sheet1")[0]
        detected = _detect_header_row(rows)
        assert detected is not None, "Row 1 should still qualify as single-row header"
        header_start, data_start, columns = detected
        # data_start must be header_start + 1 (single-row header),
        # NOT header_start + 2 (would indicate wrongly merged sub-header).
        assert data_start == header_start + 1, (
            f"Row 2 was erroneously merged as sub-header: data_start={data_start}"
        )
        # Column names must be the row-1 values, not `メイン/副` forms.
        assert "/" not in " ".join(columns), f"columns contain / suggesting merge: {columns}"


class TestVerifyP1ColumnNameSubstringSwap:
    """QA F3: A cell value that contains its column name as substring must
    still allow QC3 / QP to catch a swap bug.  Without this test, a
    converter that swaps two sections that happen to contain the column
    name as substring could silently PASS."""

    def test_swap_with_substring_cell_value_detected(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "No."
        ws["B1"] = "タイトル"
        ws["C1"] = "概要"
        # Row 1: 概要 column value literally contains "概要"
        ws["A2"] = "1"
        ws["B2"] = "値A"
        ws["C2"] = "概要をAで記述する"
        # Row 2: different 概要 value also contains "概要"
        ws["A3"] = "2"
        ws["B3"] = "値B"
        ws["C3"] = "概要をBで記述する"
        xlsx_path = tmp_path / "subswap.xlsx"
        wb.save(xlsx_path)
        # JSON with values for 概要 column swapped between sections.
        data = {
            "id": "x", "title": "", "content": "",
            "sheet_type": "P1",
            "sections": [
                {"id": "s1", "title": "値A",
                 "content": "No.: 1\nタイトル: 値A\n概要: 概要をBで記述する"},
                {"id": "s2", "title": "値B",
                 "content": "No.: 2\nタイトル: 値B\n概要: 概要をAで記述する"},
            ],
        }
        from scripts.verify.verify import verify_file
        import tempfile, os
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as jf:
            json.dump(data, jf, ensure_ascii=False)
            jpath = jf.name
        try:
            issues = verify_file(str(xlsx_path), jpath, "xlsx", sheet_name="Sheet1")
        finally:
            os.unlink(jpath)
        # QP must FAIL for the swap regardless of substring overlap.
        assert any("[QP]" in i for i in issues), issues


class TestVerifyP1DuplicateColumnNames:
    """QA F4: Duplicate Excel column names must not silently pass QP.

    Under zero-tolerance the spec gap ("what if two columns have the same
    name?") must be resolved to FAIL, not to the permissive "last wins"
    behaviour hidden by a dict overwrite.
    """

    def test_duplicate_header_columns_fail(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "No."
        ws["B1"] = "備考"
        ws["C1"] = "備考"   # duplicate header name
        ws["A2"] = "1"
        ws["B2"] = "x1"
        ws["C2"] = "y1"
        ws["A3"] = "2"
        ws["B3"] = "x2"
        ws["C3"] = "y2"
        xlsx_path = tmp_path / "dupcol.xlsx"
        wb.save(xlsx_path)
        data = {
            "id": "x", "title": "", "content": "",
            "sheet_type": "P1",
            "sections": [
                {"id": "s1", "title": "x1",
                 "content": "No.: 1\n備考: x1\n備考: y1"},
                {"id": "s2", "title": "x2",
                 "content": "No.: 2\n備考: x2\n備考: y2"},
            ],
        }
        from scripts.verify.verify import check_xlsx_p1_pairing
        issues = check_xlsx_p1_pairing(str(xlsx_path), data, "Sheet1")
        # Under zero-tolerance, duplicate header names must be flagged
        # explicitly — silent "last wins" hides ambiguity.
        assert any("[QP]" in i and ("duplicate" in i.lower() or "重複" in i)
                   for i in issues), issues


class TestVerifyP1EmptyHeaderSpacer:
    """QA F5: Empty header cell (spacer column) with data underneath must
    not silently drop the data-cell value.  Either the value must be
    checked, or the behaviour must be locked with a test that pins the
    chosen spec interpretation.
    """

    def test_spacer_column_with_data_flagged(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "No."
        ws["B1"] = ""           # spacer: empty header
        ws["C1"] = "タイトル"
        ws["D1"] = "備考"
        # Data row has value under the spacer column (column B).
        ws["A2"] = "1"
        ws["B2"] = "こぼれ値1"   # must not be silently dropped
        ws["C2"] = "値A"
        ws["D2"] = "r1"
        ws["A3"] = "2"
        ws["B3"] = "こぼれ値2"
        ws["C3"] = "値B"
        ws["D3"] = "r2"
        xlsx_path = tmp_path / "spacer.xlsx"
        wb.save(xlsx_path)
        # JSON omits the spacer-column values entirely (converter decided
        # the header is empty so no line is emitted).
        data = {
            "id": "x", "title": "", "content": "",
            "sheet_type": "P1",
            "sections": [
                {"id": "s1", "title": "値A",
                 "content": "No.: 1\nタイトル: 値A\n備考: r1"},
                {"id": "s2", "title": "値B",
                 "content": "No.: 2\nタイトル: 値B\n備考: r2"},
            ],
        }
        from scripts.verify.verify import verify_file
        import tempfile, os
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as jf:
            json.dump(data, jf, ensure_ascii=False)
            jpath = jf.name
        try:
            issues = verify_file(str(xlsx_path), jpath, "xlsx", sheet_name="Sheet1")
        finally:
            os.unlink(jpath)
        # The spacer-column cell values ("こぼれ値1", "こぼれ値2") are
        # genuine source cells.  Whether through QC1 (missing from JSON)
        # or an explicit spacer message, verify must FAIL — silent
        # acceptance is a zero-tolerance violation.
        assert issues, (
            "spacer-column data values must not be silently dropped; expected FAIL"
        )
        assert any("こぼれ値" in i for i in issues), issues


class TestVerifyP1ValueContainsColonPass:
    """QA F6: Positive regression guard for `_MD_SYNTAX_RE` `:` addition.
    Correctly-aligned sheet with URL and time-string values must PASS.
    """

    def test_pass_aligned_url_and_time_values(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "No."
        ws["B1"] = "URL"
        ws["C1"] = "時刻"
        ws["A2"] = "1"
        ws["B2"] = "https://example.com/a"
        ws["C2"] = "12:00"
        ws["A3"] = "2"
        ws["B3"] = "https://example.com/b"
        ws["C3"] = "13:30"
        xlsx_path = tmp_path / "colon.xlsx"
        wb.save(xlsx_path)
        data = {
            "id": "x", "title": "Sheet1", "content": "",
            "sheet_type": "P1",
            "sections": [
                {"id": "s1", "title": "https://example.com/a",
                 "content": "No.: 1\nURL: https://example.com/a\n時刻: 12:00"},
                {"id": "s2", "title": "https://example.com/b",
                 "content": "No.: 2\nURL: https://example.com/b\n時刻: 13:30"},
            ],
        }
        from scripts.verify.verify import verify_file
        import tempfile, os
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as jf:
            json.dump(data, jf, ensure_ascii=False)
            jpath = jf.name
        try:
            issues = verify_file(str(xlsx_path), jpath, "xlsx", sheet_name="Sheet1")
        finally:
            os.unlink(jpath)
        assert issues == [], issues


class TestVerifyP1FabricatedColonLine:
    """QA F7: Fabricated `列: 値` fragment must still trigger QC1/QC2.
    The `:` scrubbing must not mask colon-separated fabrications whose
    key/value are not in the source.
    """

    def test_fabricated_colon_pair_detected(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "No."
        ws["B1"] = "タイトル"
        ws["C1"] = "備考"
        ws["A2"] = "1"
        ws["B2"] = "値A"
        ws["C2"] = "r1"
        ws["A3"] = "2"
        ws["B3"] = "値B"
        ws["C3"] = "r2"
        xlsx_path = tmp_path / "fab.xlsx"
        wb.save(xlsx_path)
        # JSON injects an extra 偽列: 偽値 line in section s1 —
        # neither token is in the source sheet.
        data = {
            "id": "x", "title": "", "content": "",
            "sheet_type": "P1",
            "sections": [
                {"id": "s1", "title": "値A",
                 "content": "No.: 1\nタイトル: 値A\n備考: r1\n偽列: 偽値"},
                {"id": "s2", "title": "値B",
                 "content": "No.: 2\nタイトル: 値B\n備考: r2"},
            ],
        }
        from scripts.verify.verify import verify_file
        import tempfile, os
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as jf:
            json.dump(data, jf, ensure_ascii=False)
            jpath = jf.name
        try:
            issues = verify_file(str(xlsx_path), jpath, "xlsx", sheet_name="Sheet1")
        finally:
            os.unlink(jpath)
        # 偽列 / 偽値 must surface as QC2 or QP (unexpected column).
        assert issues, "fabricated colon pair must FAIL"
        assert any("偽" in i for i in issues), issues
