"""Common utilities for knowledge-creator steps"""

import json
import subprocess
import os
from typing import Any, Tuple, Optional
from datetime import datetime
from openpyxl import load_workbook


def load_json(path: str) -> dict:
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def write_json(path: str, data: Any):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def read_excel_as_markdown(path: str) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    markdown_parts = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        markdown_parts.append(f"## {sheet_name}\n")
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            markdown_parts.append("(Empty sheet)\n")
            continue

        max_cols = max(len(row) for row in rows) if rows else 0
        if max_cols == 0:
            markdown_parts.append("(Empty sheet)\n")
            continue

        for i, row in enumerate(rows):
            padded_row = list(row) + [None] * (max_cols - len(row))
            cells = [str(cell) if cell is not None else "" for cell in padded_row]
            markdown_parts.append("| " + " | ".join(cells) + " |")
            if i == 0:
                markdown_parts.append("| " + " | ".join(["---"] * max_cols) + " |")
        markdown_parts.append("")

    workbook.close()
    return "\n".join(markdown_parts)


def read_file(path: str) -> str:
    if path.endswith('.xlsx'):
        return read_excel_as_markdown(path)
    with open(path, 'r', encoding='utf-8') as f:
        return f.read()


def write_file(path: str, content: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        f.write(content)


def run_claude_with_metrics(
    prompt: str,
    timeout: int = 600,
    json_schema: dict = None,
    max_turns: int = None,
    log_dir: str = None,
    file_id: str = None
) -> Tuple[subprocess.CompletedProcess, dict]:
    """Run claude -p with metrics tracking and optional logging.

    Args:
        prompt: Prompt text
        timeout: Timeout in seconds
        json_schema: JSON Schema for structured output (optional)
        max_turns: Maximum number of turns (optional)
        log_dir: Directory to save logs (optional)
        file_id: File identifier for log filename (optional)

    Returns:
        (result, metrics) where:
        - result: CompletedProcess with stdout containing response/structured_output
        - metrics: dict with turn_count, duration_ms, cost_usd, log_file (if saved)
    """
    cmd = ["claude", "-p", "--output-format", "json"]

    if json_schema:
        cmd.extend(["--json-schema", json.dumps(json_schema)])

    if max_turns:
        cmd.extend(["--max-turns", str(max_turns)])

    # Remove CLAUDECODE to prevent Claude CLI from detecting agent context
    # This ensures prompts run in standard mode, not code agent mode
    env = os.environ.copy()
    env.pop('CLAUDECODE', None)

    result = subprocess.run(
        cmd, input=prompt, capture_output=True, text=True, timeout=timeout, env=env
    )

    metrics = {
        "turn_count": 0,
        "duration_ms": 0,
        "cost_usd": 0.0,
        "log_file": None
    }

    if result.returncode == 0:
        try:
            response = json.loads(result.stdout)

            # Save log if log_dir provided
            if log_dir:
                os.makedirs(log_dir, exist_ok=True)
                timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
                log_filename = f"{file_id}_{timestamp}.json" if file_id else f"claude_{timestamp}.json"
                log_path = os.path.join(log_dir, log_filename)
                with open(log_path, 'w', encoding='utf-8') as f:
                    json.dump(response, f, ensure_ascii=False, indent=2)
                metrics["log_file"] = log_path

            # Extract metrics from Claude CLI JSON output
            if "num_turns" in response:
                metrics["turn_count"] = response["num_turns"]
            if "duration_ms" in response:
                metrics["duration_ms"] = response["duration_ms"]
            if "total_cost_usd" in response:
                metrics["cost_usd"] = response["total_cost_usd"]

            # Handle structured output
            if json_schema:
                subtype = response.get("subtype", "")
                if subtype == "success":
                    structured_output = response.get("structured_output")
                    if structured_output is not None:
                        result = subprocess.CompletedProcess(
                            args=result.args, returncode=0,
                            stdout=json.dumps(structured_output, ensure_ascii=False),
                            stderr=""
                        )
                    else:
                        result = subprocess.CompletedProcess(
                            args=result.args, returncode=1,
                            stdout="", stderr="structured_output field is missing"
                        )
                elif subtype == "error_max_structured_output_retries":
                    error_msg = response.get("result", "Failed to generate valid structured output")
                    result = subprocess.CompletedProcess(
                        args=result.args, returncode=1,
                        stdout="", stderr=f"Structured output error: {error_msg}"
                    )
                else:
                    result = subprocess.CompletedProcess(
                        args=result.args, returncode=1,
                        stdout="", stderr=f"Unknown response subtype: {subtype}"
                    )
            else:
                # No json_schema: return the result text
                result_text = response.get("result", "")
                result = subprocess.CompletedProcess(
                    args=result.args, returncode=0,
                    stdout=result_text,
                    stderr=""
                )

        except json.JSONDecodeError as e:
            result = subprocess.CompletedProcess(
                args=result.args, returncode=1,
                stdout="", stderr=f"Failed to parse claude response JSON: {e}"
            )

    return result, metrics


def run_claude(prompt: str, timeout: int = 600, json_schema: dict = None) -> subprocess.CompletedProcess:
    """Run claude -p via stdin (backward compatibility wrapper).

    Args:
        prompt: Prompt text
        timeout: Timeout in seconds
        json_schema: JSON Schema for structured output (optional)

    Returns:
        CompletedProcess. When json_schema is provided, stdout contains the structured_output JSON.
    """
    result, _ = run_claude_with_metrics(prompt, timeout, json_schema)
    return result
