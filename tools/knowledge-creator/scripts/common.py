"""Common utilities for knowledge-creator steps"""

import json
import re
import subprocess
import os
from typing import Any
from datetime import datetime
from openpyxl import load_workbook


def load_json(path: str) -> dict:
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def write_json(path: str, data: Any):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def read_excel_as_markdown(path: str) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    markdown_parts = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        markdown_parts.append(f"## {sheet_name}\n")
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            markdown_parts.append("(Empty sheet)\n")
            continue

        max_cols = max(len(row) for row in rows) if rows else 0
        if max_cols == 0:
            markdown_parts.append("(Empty sheet)\n")
            continue

        for i, row in enumerate(rows):
            padded_row = list(row) + [None] * (max_cols - len(row))
            cells = [str(cell) if cell is not None else "" for cell in padded_row]
            markdown_parts.append("| " + " | ".join(cells) + " |")
            if i == 0:
                markdown_parts.append("| " + " | ".join(["---"] * max_cols) + " |")
        markdown_parts.append("")

    workbook.close()
    return "\n".join(markdown_parts)


def read_file(path: str) -> str:
    if path.endswith('.xlsx'):
        return read_excel_as_markdown(path)
    with open(path, 'r', encoding='utf-8') as f:
        return f.read()


def write_file(path: str, content: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        f.write(content)


def run_claude(prompt: str, json_schema: dict, log_dir: str, file_id: str) -> subprocess.CompletedProcess:
    """Run claude -p with full logging (always-on verbose mode).

    Args:
        prompt: Prompt text
        json_schema: JSON Schema for structured output
        log_dir: Directory to save execution logs (e.g., ctx.phase_b_executions_dir)
        file_id: File identifier for log filename (e.g., "libraries-tag")

    Returns:
        CompletedProcess with stdout containing structured_output JSON

    Output files (all saved to log_dir with shared {file_id}_{timestamp} prefix):
        .json      - metadata: file_id, timestamp, subtype, cc_metrics, stop_reason, tool_calls
        .in.txt    - prompt text (IN)
        .out.json  - structured_output JSON (OUT)
        .ndjson    - raw stream-json output from Claude CLI
    """
    disallowed = "Read,Edit,Write,Glob,Grep,LS,ToolSearch"

    cmd = [
        "claude", "-p",
        "--output-format", "stream-json",
        "--verbose",
        "--json-schema", json.dumps(json_schema),
        "--max-turns", "10",
        "--disallowedTools", disallowed
    ]

    # Remove CLAUDECODE to prevent Claude CLI from detecting agent context
    # This ensures prompts run in standard mode, not code agent mode
    env = os.environ.copy()
    env.pop('CLAUDECODE', None)

    result = subprocess.run(
        cmd, input=prompt, capture_output=True, text=True, env=env
    )

    if result.returncode == 0:
        try:
            raw_output = result.stdout
            response = None
            ndjson_lines = []
            for line in raw_output.splitlines():
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                    ndjson_lines.append(obj)
                    if obj.get("type") == "result":
                        response = obj
                except json.JSONDecodeError:
                    continue

            if response is None:
                return subprocess.CompletedProcess(
                    args=result.args, returncode=1,
                    stdout="", stderr="No result line found in stream-json output"
                )

            # Save execution logs
            os.makedirs(log_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            base = os.path.join(log_dir, f"{file_id}_{timestamp}")

            # Extract tool calls from stream
            tool_calls = []
            for obj in ndjson_lines:
                content_items = []
                if isinstance(obj.get("message"), dict):
                    content_items = obj["message"].get("content", [])
                elif isinstance(obj.get("content"), list):
                    content_items = obj["content"]
                for item in content_items:
                    if isinstance(item, dict) and item.get("type") == "tool_use":
                        input_data = item.get("input", {})
                        input_summary = {}
                        for k, v in input_data.items():
                            sv = str(v)
                            input_summary[k] = (sv[:200] + "...") if len(sv) > 200 else sv
                        tool_calls.append({
                            "tool_name": item.get("name", "unknown"),
                            "input_summary": input_summary,
                        })

            cc_metrics = {
                "duration_ms":     response.get("duration_ms"),
                "duration_api_ms": response.get("duration_api_ms"),
                "num_turns":       response.get("num_turns"),
                "total_cost_usd":  response.get("total_cost_usd"),
                "usage":           response.get("usage", {}),
            }

            # {base}.json - metadata
            with open(f"{base}.json", 'w', encoding='utf-8') as f:
                json.dump({
                    "file_id":     file_id,
                    "timestamp":   timestamp,
                    "subtype":     response.get("subtype"),
                    "cc_metrics":  cc_metrics,
                    "stop_reason": response.get("stop_reason"),
                    "tool_calls":  tool_calls,
                }, f, ensure_ascii=False, indent=2)

            # {base}.in.txt - prompt (IN)
            with open(f"{base}.in.txt", 'w', encoding='utf-8') as f:
                f.write(prompt)

            # {base}.out.json - structured_output (OUT)
            structured_output = response.get("structured_output")
            with open(f"{base}.out.json", 'w', encoding='utf-8') as f:
                json.dump(structured_output, f, ensure_ascii=False, indent=2)

            # {base}.ndjson - raw stream-json output
            with open(f"{base}.ndjson", 'w', encoding='utf-8') as f:
                f.write(raw_output)

            # Handle structured output
            subtype = response.get("subtype", "")
            if subtype == "success":
                if structured_output is not None:
                    return subprocess.CompletedProcess(
                        args=result.args, returncode=0,
                        stdout=json.dumps(structured_output, ensure_ascii=False),
                        stderr=""
                    )
                else:
                    return subprocess.CompletedProcess(
                        args=result.args, returncode=1,
                        stdout="", stderr="structured_output field is missing"
                    )
            elif subtype == "error_max_structured_output_retries":
                error_msg = response.get("result", "Failed to generate valid structured output")
                return subprocess.CompletedProcess(
                    args=result.args, returncode=1,
                    stdout="", stderr=f"Structured output error: {error_msg}"
                )
            else:
                return subprocess.CompletedProcess(
                    args=result.args, returncode=1,
                    stdout="", stderr=f"Unknown response subtype: {subtype}"
                )

        except json.JSONDecodeError as e:
            return subprocess.CompletedProcess(
                args=result.args, returncode=1,
                stdout="", stderr=f"Failed to parse claude response JSON: {e}"
            )

    return result


def count_source_headings(content: str, fmt: str) -> int:
    """Count split-level headings in source content.

    Args:
        content: Source file content
        fmt: Source format ('rst', 'md', 'xlsx')

    Returns:
        Number of headings found
    """
    if fmt == "rst":
        return len(re.findall(r'\n[^\n]+\n-{3,}\n', content))
    elif fmt == "md":
        return len(re.findall(r'^## (?!#)', content, re.MULTILINE))
    elif fmt == "xlsx":
        return 1
    return 0


def aggregate_cc_metrics(executions_dir: str) -> dict:
    """executions ディレクトリの execution log を走査してメトリクスを集計する。

    Returns:
        {
          "count": int,
          "tokens": {"input": int, "cache_creation": int, "cache_read": int, "output": int},
          "cost_usd": float,
          "avg_turns": float,        # ターンデータがある場合のみ含まれる
          "avg_duration_sec": float, # duration データがある場合のみ含まれる
          "p95_duration_sec": float, # duration データがある場合のみ含まれる
        }
    """
    import glob

    tokens = {"input": 0, "cache_creation": 0, "cache_read": 0, "output": 0}
    cost_usd = 0.0
    turns = []
    durations = []
    count = 0

    for path in glob.glob(os.path.join(executions_dir, "*.json")):
        if path.endswith(".out.json"):
            continue
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except (json.JSONDecodeError, OSError):
            continue

        cc = data.get("cc_metrics", {})
        usage = cc.get("usage", {})
        tokens["input"]          += usage.get("input_tokens", 0)
        tokens["cache_creation"] += usage.get("cache_creation_input_tokens", 0)
        tokens["cache_read"]     += usage.get("cache_read_input_tokens", 0)
        tokens["output"]         += usage.get("output_tokens", 0)
        cost_usd                 += cc.get("total_cost_usd") or 0.0
        if cc.get("num_turns"):
            turns.append(cc["num_turns"])
        if cc.get("duration_ms"):
            durations.append(cc["duration_ms"] / 1000.0)
        count += 1

    result = {"count": count, "tokens": tokens, "cost_usd": round(cost_usd, 4)}
    if turns:
        result["avg_turns"] = round(sum(turns) / len(turns), 1)
    if durations:
        sorted_d = sorted(durations)
        result["avg_duration_sec"] = round(sum(sorted_d) / len(sorted_d), 1)
        p95_idx = max(0, int(len(sorted_d) * 0.95) - 1)
        result["p95_duration_sec"] = round(sorted_d[p95_idx], 1)

    return result
