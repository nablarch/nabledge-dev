"""
Prototype script for Task 3: generate expected JSON + docs MD for security-check-2 sheet
with P1-merged grouping applied.

This is a one-off script — not production code.
"""
import json
import openpyxl
from pathlib import Path

EXCEL_PATH = Path(__file__).parents[2] / ".lw/nab-official/v6/nablarch-system-development-guide/Sample_Project/設計書/Nablarch機能のセキュリティ対応表.xlsx"
SHEET_NAME = "2.チェックリスト"
TITLE_COL = 3  # column C (1-indexed)
DATA_START_ROW = 9  # row 9 is first data row (rows 7-8 are header)

def get_column_headers(ws, header_rows):
    """Build column header map from multi-row header."""
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        parts = []
        for row_idx in header_rows:
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                parts.append(str(val).strip())
        if parts:
            headers[col_idx] = " / ".join(parts)
    return headers

def get_merged_groups(ws, title_col, data_start_row):
    """Return list of (start_row, end_row) for each vulnerability group."""
    # Collect merge ranges in title column
    merge_ranges = []
    for m in ws.merged_cells.ranges:
        if m.min_col <= title_col <= m.max_col and m.min_row >= data_start_row:
            merge_ranges.append((m.min_row, m.max_row))

    # Find the last data row
    last_data_row = data_start_row
    for row_idx in range(data_start_row, ws.max_row + 1):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None for v in row_vals):
            last_data_row = row_idx

    # Build groups: start with all merge ranges, add single rows for unmerged title cells
    merged_rows = set()
    for start, end in merge_ranges:
        for r in range(start, end + 1):
            merged_rows.add(r)

    groups = list(merge_ranges)
    for row_idx in range(data_start_row, last_data_row + 1):
        title_val = ws.cell(row=row_idx, column=title_col).value
        if row_idx not in merged_rows and title_val is not None:
            groups.append((row_idx, row_idx))

    groups.sort(key=lambda x: x[0])
    return groups, last_data_row

def flatten_cell(val):
    if val is None:
        return ""
    return str(val).replace("\n", " ").strip()

def build_data_row(ws, row_idx, col_count):
    """Build a data_row list (raw cell values, LF preserved) as current JSON has."""
    return [ws.cell(row=row_idx, column=c).value or "" for c in range(1, col_count + 1)]

def build_section_content(ws, start_row, end_row, headers):
    """Build section content string for a group of rows."""
    lines = []
    for row_idx in range(start_row, end_row + 1):
        for col_idx, header in sorted(headers.items()):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                flat = flatten_cell(val)
                if flat:
                    lines.append(f"{header}: {flat}")
    return "\n".join(lines)

def get_section_title(ws, start_row, title_col):
    return flatten_cell(ws.cell(row=start_row, column=title_col).value)

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # Header rows are 7 and 8
    headers = get_column_headers(ws, [7, 8])
    groups, last_data_row = get_merged_groups(ws, TITLE_COL, DATA_START_ROW)

    print(f"Found {len(groups)} vulnerability groups (data rows {DATA_START_ROW}-{last_data_row})")
    print()

    # Generate MD
    md_lines = ["# 2.チェックリスト", ""]

    # Preamble: rows 1-6
    preamble_lines = []
    for row_idx in range(1, 7):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        for v in row_vals:
            if v is not None:
                preamble_lines.append(str(v).strip())
    if preamble_lines:
        md_lines.append("\n".join(preamble_lines))
        md_lines.append("")

    # Build sections
    sections = []
    for start_row, end_row in groups:
        title = get_section_title(ws, start_row, TITLE_COL)
        content = build_section_content(ws, start_row, end_row, headers)
        sections.append((title, content))

    # MD table (full data)
    col_list = sorted(headers.items())
    header_cells = [h for _, h in col_list]
    md_lines.append("| " + " | ".join(header_cells) + " |")
    md_lines.append("| " + " | ".join(["---"] * len(header_cells)) + " |")
    for row_idx in range(DATA_START_ROW, last_data_row + 1):
        row_cells = []
        for col_idx, _ in col_list:
            val = ws.cell(row=row_idx, column=col_idx).value
            row_cells.append(flatten_cell(val).replace("|", "\\|"))
        md_lines.append("| " + " | ".join(row_cells) + " |")
    md_lines.append("")

    # --- Build JSON preview ---
    col_count = ws.max_column

    # columns list: same as current P1 (flatten header)
    columns = []
    for col_idx in range(1, col_count + 1):
        columns.append(headers.get(col_idx, ""))

    # data_rows: one entry per group (representative/first row only, per P1-merged spec)
    json_data_rows = []
    for start_row, _ in groups:
        json_data_rows.append(build_data_row(ws, start_row, col_count))

    # sections: one per group
    json_sections = []
    for i, (title, content) in enumerate(sections, 1):
        json_sections.append({
            "id": f"s{i}",
            "title": title,
            "content": content,
        })

    # index: __file__ entry + one per section
    json_index = [{"id": "__file__", "title": "2.チェックリスト"}]
    for s in json_sections:
        json_index.append({"id": s["id"], "title": s["title"]})

    # Preamble text for content field
    preamble_text = "\n".join(preamble_lines)

    json_out = {
        "id": "security-check-2.チェックリスト",
        "title": "2.チェックリスト",
        "content": preamble_text,
        "no_knowledge_content": False,
        "sections": json_sections,
        "sheet_type": "P1",
        "sheet_subtype": "P1-merged",
        "columns": columns,
        "data_rows": json_data_rows,
        "index": json_index,
    }

    json_path = Path(__file__).parent / "preview-security-check-2-checklist.json"
    json_path.write_text(json.dumps(json_out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Written JSON to: {json_path}")

    # Sections summary
    print(f"\n=== JSON sections ({len(json_sections)}) ===")
    for s in json_sections:
        line_count = s['content'].count("\n") + 1
        print(f"  {s['id']}: '{s['title']}' ({line_count} content lines)")

    # --- Build MD preview ---
    preview_lines = list(md_lines)
    for title, content in sections:
        preview_lines.append(f"## {title}")
        preview_lines.append("")
        preview_lines.append(content)
        preview_lines.append("")

    md_path = Path(__file__).parent / "preview-security-check-2-checklist.md"
    md_path.write_text("\n".join(preview_lines), encoding="utf-8")
    print(f"Written MD  to: {md_path}")

if __name__ == "__main__":
    main()
