"""Survey all RBKC-relevant Excel files (releasenote + security).

Outputs (per sheet):
- title_row_idx: row index of first cell starting with '■' (or -1)
- header_row_idx: row index of first "dense" row (>=3 consecutive non-empty cells starting at a non-None position) below title
- preamble_cells: list of non-empty cells between title row (exclusive) and header row (exclusive)
- header_rows: list of rows used as multi-row header (1 or more, consecutive above first data row)
- data_rows_count: count of non-empty data rows after header
- sheet_type_predicted: P1 if header detectable and >=2 data rows, else P2
- duplicate_column_names_in_last_header_row: if top header row has duplicates
- col_count: #non-empty cells in last header row

Prints a CSV-ish summary + distribution counts.
"""
from __future__ import annotations

import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

from openpyxl import load_workbook  # noqa: E402
import xlrd  # type: ignore  # noqa: E402


ROOT = Path("/home/tie303177/work/nabledge/work2/.lw/nab-official")


def _iter_excel_files() -> list[Path]:
    result = []
    for p in sorted(ROOT.rglob("*releasenote*.xls*")):
        result.append(p)
    for p in sorted(ROOT.rglob("*security*.xlsx")):
        result.append(p)
    return result


def _load_rows(path: Path) -> list[tuple[str, list[list[object]]]]:
    """Return [(sheet_name, rows)] where rows is list of lists (cell values)."""
    out = []
    if path.suffix.lower() == ".xlsx":
        wb = load_workbook(path, data_only=True, read_only=True)
        for ws in wb.worksheets:
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            out.append((ws.title, rows))
        wb.close()
    elif path.suffix.lower() == ".xls":
        book = xlrd.open_workbook(str(path), on_demand=True)
        for sn in book.sheet_names():
            sh = book.sheet_by_name(sn)
            rows = []
            for r in range(sh.nrows):
                rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
            out.append((sn, rows))
    return out


def _non_empty(cell) -> bool:
    if cell is None:
        return False
    if isinstance(cell, str) and not cell.strip():
        return False
    return True


def _first_title_row(rows: list[list[object]]) -> int:
    for i, row in enumerate(rows):
        if row and isinstance(row[0], str) and row[0].lstrip().startswith("■"):
            return i
    return -1


def _first_dense_row(rows: list[list[object]], start: int) -> int:
    """First row with >=3 consecutive non-empty cells."""
    for i in range(start, len(rows)):
        row = rows[i]
        max_run = cur = 0
        for c in row:
            if _non_empty(c):
                cur += 1
                max_run = max(max_run, cur)
            else:
                cur = 0
        if max_run >= 3:
            return i
    return -1


def _is_header_like(row: list[object]) -> bool:
    """A row that is non-empty but has short text (likely a header row)."""
    cells = [c for c in row if _non_empty(c)]
    if not cells:
        return False
    # headers typically have short cell values
    short = sum(1 for c in cells if isinstance(c, str) and len(c) <= 40)
    return short >= len(cells) * 0.6


def main() -> None:
    files = _iter_excel_files()
    print(f"# {len(files)} files")

    stats = {
        "total_sheets": 0,
        "p1_pred": 0,
        "p2_pred": 0,
        "preamble_nonempty": 0,
        "multi_row_header": 0,
        "duplicate_columns": 0,
        "no_title_row": 0,
    }

    multi_row_examples = []
    dup_col_examples = []
    preamble_examples = []

    for path in files:
        rel = path.relative_to(ROOT)
        try:
            sheets = _load_rows(path)
        except Exception as e:
            print(f"LOAD_FAIL {rel}: {e}")
            continue
        for sn, rows in sheets:
            stats["total_sheets"] += 1
            t = _first_title_row(rows)
            if t < 0:
                stats["no_title_row"] += 1
            h = _first_dense_row(rows, max(t + 1, 0))
            if h < 0:
                stats["p2_pred"] += 1
                continue
            # Look upward from h to see if the previous row is also header-like
            header_rows = [h]
            i = h - 1
            while i > t and _is_header_like(rows[i]):
                header_rows.insert(0, i)
                i -= 1
            # preamble: non-empty cells between (t, header_rows[0])
            preamble = []
            for ri in range(t + 1, header_rows[0]):
                for ci, c in enumerate(rows[ri]):
                    if _non_empty(c):
                        preamble.append((ri, ci, str(c)[:60]))
            # data rows = non-empty rows after h
            data_count = 0
            for ri in range(h + 1, len(rows)):
                if any(_non_empty(c) for c in rows[ri]):
                    data_count += 1
            if data_count < 2:
                stats["p2_pred"] += 1
                continue
            # last header row's non-empty cells
            last_header = [c for c in rows[h] if _non_empty(c)]
            has_dup = len(last_header) != len(set(str(c) for c in last_header))
            if has_dup:
                stats["duplicate_columns"] += 1
                if len(dup_col_examples) < 10:
                    dup_col_examples.append((str(rel), sn, last_header))
            if len(header_rows) >= 2:
                stats["multi_row_header"] += 1
                if len(multi_row_examples) < 10:
                    multi_row_examples.append((str(rel), sn, [rows[r] for r in header_rows]))
            if preamble:
                stats["preamble_nonempty"] += 1
                if len(preamble_examples) < 10:
                    preamble_examples.append((str(rel), sn, preamble))
            stats["p1_pred"] += 1

    print("\n## Stats")
    for k, v in stats.items():
        print(f"{k}: {v}")

    print("\n## Multi-row header examples (up to 10)")
    for rel, sn, hrs in multi_row_examples:
        print(f"- {rel} :: {sn}")
        for r in hrs:
            print(f"   {[str(c)[:30] if c is not None else '' for c in r]}")

    print("\n## Duplicate column examples (up to 10)")
    for rel, sn, row in dup_col_examples:
        print(f"- {rel} :: {sn}: {row}")

    print("\n## Preamble examples (up to 10)")
    for rel, sn, pre in preamble_examples:
        print(f"- {rel} :: {sn}: {len(pre)} cells")
        for ri, ci, v in pre[:5]:
            print(f"   r{ri} c{ci}: {v}")


if __name__ == "__main__":
    main()
