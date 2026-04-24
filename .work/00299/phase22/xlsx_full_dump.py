"""Exhaustive dump of every xlsx/xls sheet in RBKC scope.

For each sheet, emit a JSON record:

{
  "file": "relative path",
  "sheet": "sheet name",
  "title_row": int (-1 if none),
  "title": str,
  "dense_row": int (-1),       # first row with >=3 consecutive non-empty cells
  "header_rows": [int, ...],   # rows treated as header (walk upward from dense_row)
  "preamble_cells": [(row, col, value), ...],  # non-empty cells between title and header
  "column_raw_rows": [[...], ...],  # raw cell rows used for header composition
  "columns_composed": [...],   # after parent/child span composition
  "composed_separator": " / ", # separator used
  "data_rows_count": int,
  "duplicate_after_compose": bool,
  "col_endings": { ": ": n, " / ": n, ... }   # counts of suffix patterns in composed columns
}
"""
from __future__ import annotations

import json
import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

from openpyxl import load_workbook
try:
    import xlrd
except ImportError:
    xlrd = None


ROOT = Path("/home/tie303177/work/nabledge/work2/.lw/nab-official")
OUT = Path("/home/tie303177/work/nabledge/work2/.work/00299/phase22/xlsx-full-dump.json")
SEP = " / "


def _iter_excel_files() -> list[Path]:
    out: list[Path] = []
    out.extend(sorted(ROOT.rglob("*releasenote*.xls*")))
    out.extend(sorted(ROOT.rglob("*security*.xlsx")))
    return out


def _load_rows(path: Path) -> list[tuple[str, list[list]]]:
    out = []
    if path.suffix.lower() == ".xlsx":
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            for ws in wb.worksheets:
                rows = [list(r) for r in ws.iter_rows(values_only=True)]
                out.append((ws.title, rows))
        finally:
            wb.close()
    elif path.suffix.lower() == ".xls":
        if xlrd is None:
            return []
        book = xlrd.open_workbook(str(path), on_demand=True)
        for sn in book.sheet_names():
            sh = book.sheet_by_name(sn)
            rows = []
            for r in range(sh.nrows):
                rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
            out.append((sn, rows))
    return out


def _non_empty(cell) -> bool:
    if cell is None:
        return False
    if isinstance(cell, str) and not cell.strip():
        return False
    return True


def _first_title_row(rows: list[list]) -> int:
    for i, row in enumerate(rows):
        if row and isinstance(row[0], str) and row[0].lstrip().startswith("■"):
            return i
    return -1


def _first_dense_row(rows: list[list], start: int) -> int:
    for i in range(start, len(rows)):
        row = rows[i]
        max_run = cur = 0
        for c in row:
            if _non_empty(c):
                cur += 1
                max_run = max(max_run, cur)
            else:
                cur = 0
        if max_run >= 3:
            return i
    return -1


def _is_header_like(row: list) -> bool:
    cells = [c for c in row if _non_empty(c)]
    if not cells:
        return False
    # headers typically have short text cells (<= 40 chars)
    short = sum(1 for c in cells if isinstance(c, str) and len(c) <= 40)
    return short >= max(1, int(len(cells) * 0.6))


def _compose_columns(header_rows: list[list], col_count: int) -> list[str]:
    """Compose columns from N header rows using parent/child span logic.

    A "parent" row value at col c spans rightward until the next non-empty
    cell in the same row (Excel merge-like behaviour even when the source
    is not using actual merged cells — parent label appears only at span
    start). The last header row is the "leaf" and never spans.

    Rule:
    - For each column c, walk header_rows top-down.
    - For rows above the leaf: use the most recent non-empty cell at or
      left of col c in that row (span-inherit).
    - For the leaf row: use cell at col c verbatim (no inherit).
    - Compose = SEP.join(non-empty parts).
    """
    if not header_rows:
        return [""] * col_count
    # Pre-compute span-inherited values for each parent row.
    parent_rows = header_rows[:-1]
    leaf_row = header_rows[-1]
    inherited = []  # list of lists, each length col_count
    for pr in parent_rows:
        row_inh: list[str] = [""] * col_count
        last = ""
        for c in range(col_count):
            if c < len(pr) and _non_empty(pr[c]):
                last = str(pr[c]).strip()
            row_inh[c] = last
        # A parent cell only "spans" while leaf has non-empty cells under it.
        # But we keep the inherited value per-column; empty leaf cells get
        # filtered out in the final compose step below.
        inherited.append(row_inh)
    composed: list[str] = []
    for c in range(col_count):
        leaf = str(leaf_row[c]).strip() if c < len(leaf_row) and _non_empty(leaf_row[c]) else ""
        parts: list[str] = []
        for row_inh in inherited:
            if row_inh[c]:
                parts.append(row_inh[c])
        if leaf:
            parts.append(leaf)
        # Dedup consecutive identical parts (e.g. single-row header)
        dedup: list[str] = []
        for p in parts:
            if not dedup or dedup[-1] != p:
                dedup.append(p)
        composed.append(SEP.join(dedup))
    return composed


def _walk_header_rows(rows: list[list], dense_row: int, title_row: int) -> list[int]:
    """Return ordered (top-down) row indices used as header, including dense_row."""
    idxs = [dense_row]
    i = dense_row - 1
    while i > title_row and _is_header_like(rows[i]):
        # A row is treated as part of the header only if its non-empty cells
        # align above non-empty cells in the already-collected header rows.
        above = rows[i]
        below = rows[idxs[0]]
        # A header row should have multiple short cells that align over
        # distinct columns of the leaf header. A single-cell row or a row
        # whose only alignment is at column 0 is preamble, not header.
        above_nonempty_cols = [c for c, v in enumerate(above) if _non_empty(v)]
        aligns = 0
        for c in range(min(len(above), len(below))):
            if _non_empty(above[c]) and _non_empty(below[c]):
                aligns += 1
        # Require at least 2 non-empty cells in the candidate header row AND
        # at least 2 alignments with the leaf row.
        if len(above_nonempty_cols) >= 2 and aligns >= 2:
            idxs.insert(0, i)
            i -= 1
        else:
            break
    return idxs


def main() -> None:
    files = _iter_excel_files()
    results = []
    for path in files:
        rel = str(path.relative_to(ROOT))
        try:
            sheets = _load_rows(path)
        except Exception as e:
            results.append({"file": rel, "sheet": None, "error": f"load_fail: {e!r}"})
            continue
        for sn, rows in sheets:
            title_row = _first_title_row(rows)
            # Title: first non-empty cell of title_row (or "")
            title = ""
            if title_row >= 0:
                for c in rows[title_row]:
                    if _non_empty(c):
                        title = str(c).strip()
                        break
            dense = _first_dense_row(rows, max(title_row + 1, 0))
            if dense < 0:
                results.append({
                    "file": rel, "sheet": sn, "title_row": title_row, "title": title,
                    "dense_row": -1, "header_rows": [], "preamble_cells": [],
                    "columns_raw": [], "columns_composed": [],
                    "data_rows_count": 0, "duplicate_after_compose": False,
                    "classification_guess": "P2",
                })
                continue
            header_rows_idx = _walk_header_rows(rows, dense, title_row)
            # preamble = non-empty cells between (title_row, first header row)
            pre_start = title_row + 1 if title_row >= 0 else 0
            pre_end = header_rows_idx[0]
            preamble = []
            for ri in range(pre_start, pre_end):
                for ci, c in enumerate(rows[ri]):
                    if _non_empty(c):
                        preamble.append([ri, ci, str(c)])
            col_count = max(len(rows[r]) for r in header_rows_idx)
            header_raw = [rows[r] for r in header_rows_idx]
            composed = _compose_columns(header_raw, col_count)
            # Data rows: non-empty after dense
            data_count = 0
            for ri in range(dense + 1, len(rows)):
                if any(_non_empty(c) for c in rows[ri]):
                    data_count += 1
            non_empty_composed = [c for c in composed if c]
            dup = len(non_empty_composed) != len(set(non_empty_composed))
            classification = "P1" if data_count >= 2 and col_count >= 3 else "P2"
            results.append({
                "file": rel, "sheet": sn,
                "title_row": title_row, "title": title,
                "dense_row": dense,
                "header_rows": header_rows_idx,
                "preamble_cells": preamble,
                "columns_raw": [[str(c) if _non_empty(c) else "" for c in r] for r in header_raw],
                "columns_composed": composed,
                "data_rows_count": data_count,
                "duplicate_after_compose": dup,
                "classification_guess": classification,
            })
    OUT.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUT} ({len(results)} records)")


if __name__ == "__main__":
    main()
