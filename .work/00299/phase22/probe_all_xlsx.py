#!/usr/bin/env python3
"""Probe ALL target xlsx/xls (release notes + security) across versions.

Phase 22-B-1 cross-version sweep.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

ROOT = Path("/home/tie303177/work/nabledge/work2/.lw/nab-official")


def truncate(s, n=30):
    s = str(s) if s is not None else ""
    return (s[:n] + "…") if len(s) > n else s


def probe_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        top_rows = []
        for r in range(1, min(ws.max_row, 6) + 1):
            vals = [truncate(ws.cell(r, c).value, 20) for c in range(1, min(ws.max_column, 15) + 1)]
            top_rows.append([v for v in vals if v])  # non-empty only
        sheets.append({"name": sname, "max_row": ws.max_row, "max_col": ws.max_column, "top_rows": top_rows})
    return sheets


def probe_xls(path: Path) -> list[dict]:
    if not HAS_XLRD:
        return [{"name": "<xlrd unavailable>"}]
    wb = xlrd.open_workbook(path)
    sheets = []
    for s in wb.sheets():
        top_rows = []
        for r in range(min(s.nrows, 6)):
            vals = [truncate(s.cell(r, c).value, 20) for c in range(min(s.ncols, 15))]
            top_rows.append([v for v in vals if v])
        sheets.append({"name": s.name, "max_row": s.nrows, "max_col": s.ncols, "top_rows": top_rows})
    return sheets


def main() -> int:
    targets = []
    targets += sorted(ROOT.rglob("*-releasenote.xlsx"))
    targets += sorted(ROOT.rglob("*-releasenote-detail.xls"))
    targets += sorted(ROOT.rglob("Nablarch機能のセキュリティ対応表.xlsx"))

    out_lines = []
    for path in targets:
        rel = path.relative_to(ROOT)
        out_lines.append(f"\n===== {rel} =====")
        try:
            if path.suffix == ".xlsx":
                sheets = probe_xlsx(path)
            else:
                sheets = probe_xls(path)
        except Exception as e:
            out_lines.append(f"  ERROR: {e}")
            continue
        for sh in sheets:
            out_lines.append(f"  sheet: {sh['name']!r}  rows={sh.get('max_row')}  cols={sh.get('max_col')}")
            for i, row in enumerate(sh.get("top_rows", []), 1):
                if row:
                    out_lines.append(f"    r{i}: {row}")

    output = "\n".join(out_lines)
    print(output)
    return 0


if __name__ == "__main__":
    sys.exit(main())
