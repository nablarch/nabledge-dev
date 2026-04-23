#!/usr/bin/env python3
"""Probe v6 xlsx structure for Phase 22-B design."""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

V6_XLSX = [
    Path("/home/tie303177/work/nabledge/work2/.lw/nab-official/v6/nablarch-document/ja/releases/nablarch6-releasenote.xlsx"),
    Path("/home/tie303177/work/nabledge/work2/.lw/nab-official/v6/nablarch-document/ja/releases/nablarch6u1-releasenote.xlsx"),
    Path("/home/tie303177/work/nabledge/work2/.lw/nab-official/v6/nablarch-document/ja/releases/nablarch6u2-releasenote.xlsx"),
    Path("/home/tie303177/work/nabledge/work2/.lw/nab-official/v6/nablarch-document/ja/releases/nablarch6u3-releasenote.xlsx"),
    Path("/home/tie303177/work/nabledge/work2/.lw/nab-official/v6/nablarch-system-development-guide/Sample_Project/設計書/Nablarch機能のセキュリティ対応表.xlsx"),
]


def truncate(s, n=40):
    s = str(s) if s is not None else ""
    return (s[:n] + "…") if len(s) > n else s


def probe(path: Path) -> None:
    print(f"\n===== {path.name} =====")
    wb = openpyxl.load_workbook(path, data_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\n  sheet: {sname!r}")
        print(f"    max_row={ws.max_row}, max_col={ws.max_column}")
        # Print first 5 rows compactly to see title / header positions
        for r in range(1, min(ws.max_row, 8) + 1):
            row_vals = [truncate(ws.cell(r, c).value, 30) for c in range(1, min(ws.max_column, 15) + 1)]
            non_empty = [v for v in row_vals if v]
            if non_empty:
                print(f"    row{r:>2}: {row_vals}")
            else:
                print(f"    row{r:>2}: (empty)")


def main() -> int:
    for p in V6_XLSX:
        if not p.exists():
            print(f"MISSING: {p}", file=sys.stderr)
            continue
        probe(p)
    return 0


if __name__ == "__main__":
    sys.exit(main())
