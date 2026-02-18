#!/usr/bin/env python3
"""Export mapping to Excel (Step 11).

This script creates an Excel file with:
- Summary sheet: Overview statistics
- Mappings sheet: All entries with source, title, categories, target
- Stats by Category sheet: Category counts
- Stats by Directory sheet: Target directory distribution

Input: mapping-v6.json
Output: mapping-v6.json.xlsx
"""

import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


VERSION = "6"
INPUT_FILE = Path("doc/mapping-creation/work-v6/mapping-v6.json")
OUTPUT_FILE = Path("doc/mapping-creation/work-v6/mapping-v6.json.xlsx")


def create_summary_sheet(wb: Workbook, mapping: dict) -> None:
    """Create Summary sheet.

    Args:
        wb: Workbook
        mapping: Mapping JSON structure
    """
    ws = wb.active
    ws.title = "Summary"

    # Header
    ws.append(["Mapping Statistics"])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    # Basic info
    ws.append(["Version", mapping["version"]])
    ws.append(["Created", mapping["created_at"]])
    ws.append([])

    # Counts
    entries = mapping["mappings"]
    with_targets = sum(1 for e in entries if 'target_files' in e)
    no_content = sum(1 for e in entries if e.get('_no_content'))

    ws.append(["Total Entries", len(entries)])
    ws.append(["With Targets", with_targets])
    ws.append(["Navigation-only", no_content])

    # Style header row
    for row in ws.iter_rows(min_row=3, max_row=8, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40


def create_mappings_sheet(wb: Workbook, mapping: dict) -> None:
    """Create Mappings sheet.

    Args:
        wb: Workbook
        mapping: Mapping JSON structure
    """
    ws = wb.create_sheet("Mappings")

    # Header
    headers = ["ID", "Source File", "Title", "Categories", "Target Files", "No Content", "Reason"]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for entry in mapping["mappings"]:
        ws.append([
            entry['id'],
            entry['source_file'],
            entry['title'],
            ', '.join(entry['categories']),
            ', '.join(entry.get('target_files', [])),
            'Yes' if entry.get('_no_content') else 'No',
            entry.get('_no_content_reason', '')
        ])

    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 50

    # Freeze header row
    ws.freeze_panes = 'A2'


def create_category_stats_sheet(wb: Workbook, mapping: dict) -> None:
    """Create Stats by Category sheet.

    Args:
        wb: Workbook
        mapping: Mapping JSON structure
    """
    ws = wb.create_sheet("Stats by Category")

    # Count by category
    category_counts = {}
    for entry in mapping["mappings"]:
        for cat in entry['categories']:
            category_counts[cat] = category_counts.get(cat, 0) + 1

    # Header
    headers = ["Category", "Count"]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for cat in sorted(category_counts.keys()):
        ws.append([cat, category_counts[cat]])

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15

    # Add total row
    total_row = len(category_counts) + 2
    ws[f'A{total_row}'] = "TOTAL"
    ws[f'A{total_row}'].font = Font(bold=True)
    ws[f'B{total_row}'] = sum(category_counts.values())
    ws[f'B{total_row}'].font = Font(bold=True)


def create_directory_stats_sheet(wb: Workbook, mapping: dict) -> None:
    """Create Stats by Directory sheet.

    Args:
        wb: Workbook
        mapping: Mapping JSON structure
    """
    ws = wb.create_sheet("Stats by Directory")

    # Count by directory
    directory_counts = {}
    for entry in mapping["mappings"]:
        if 'target_files' in entry:
            target = entry['target_files'][0]
            directory = '/'.join(target.split('/')[:2])
            directory_counts[directory] = directory_counts.get(directory, 0) + 1

    # Header
    headers = ["Directory", "Count"]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for dir in sorted(directory_counts.keys()):
        ws.append([dir, directory_counts[dir]])

    # Adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15

    # Add total row
    total_row = len(directory_counts) + 2
    ws[f'A{total_row}'] = "TOTAL"
    ws[f'A{total_row}'].font = Font(bold=True)
    ws[f'B{total_row}'] = sum(directory_counts.values())
    ws[f'B{total_row}'].font = Font(bold=True)


def main():
    """Main execution."""
    print("=" * 80)
    print("Export to Excel - Step 11")
    print("=" * 80)
    print()

    # Load input file
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_FILE}")

    print(f"Loading: {INPUT_FILE}")
    with open(INPUT_FILE, 'r', encoding='utf-8') as f:
        mapping = json.load(f)

    print(f"✓ Loaded mapping with {len(mapping['mappings'])} entries")
    print()

    # Create workbook
    print("Creating Excel workbook...")
    wb = Workbook()

    # Create sheets
    print("  Creating Summary sheet...")
    create_summary_sheet(wb, mapping)

    print("  Creating Mappings sheet...")
    create_mappings_sheet(wb, mapping)

    print("  Creating Stats by Category sheet...")
    create_category_stats_sheet(wb, mapping)

    print("  Creating Stats by Directory sheet...")
    create_directory_stats_sheet(wb, mapping)

    print("✓ All sheets created")
    print()

    # Save workbook
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)

    print(f"✓ Written: {OUTPUT_FILE}")
    print()

    print("=" * 80)
    print("Mapping creation complete!")
    print(f"Final output: {INPUT_FILE}")
    print(f"Excel export: {OUTPUT_FILE}")
    print("=" * 80)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
